#!/usr/bin/env python3
'''
          Script :: nddu.py
         Version :: v1.3.5 (04-26-2026)
          Author :: jason.thomaschefsky@cdw.com
         Purpose :: Document network devices using "show" commands, processed with concurrent threads.
     Information :: See 'README.md' and 'CHANGELOG.md'

MIT License

Copyright (c) 2026 Jason Thomaschefsky

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
'''

# --- Import the modules needed for this script ---
import concurrent.futures
import csv
import ipaddress
import keyring
import logging
import os
import queue
import re
import platform
import shutil
import subprocess
import sys
import json
import threading
import urllib.request
import urllib.error
import time
import zipfile
from packaging import version
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, List, Set, Tuple, Union, Any, NoReturn
from netmiko import ConnectHandler, NetmikoTimeoutException, NetmikoAuthenticationException
from netmiko.ssh_autodetect import SSHDetect
from PySide6.QtCore import QObject, QSettings, QTimer, Qt, QThread, QRect, QSize, Signal
from PySide6.QtGui import QPalette, QPixmap, QPainter, QTextFormat, QColor, QPixmap, QTextCursor
from PySide6.QtWidgets import (
    QApplication, QFrame, QWidget, QLabel, QPushButton, QLineEdit, QCheckBox, QVBoxLayout, QHBoxLayout,
    QFileDialog, QGroupBox, QMessageBox, QRadioButton, QProgressBar, QScrollArea, QDialog, QTextEdit,
    QPlainTextEdit, QComboBox, QListWidget, QTableWidget, QTableWidgetItem, QHeaderView,
    QAbstractItemView, QInputDialog, QSplitter, QTreeWidget, QTreeWidgetItem
)
from urllib.error import URLError, HTTPError
from typing import Optional, Dict

# --- Silence Paramiko and Netmiko logs ---
logging.getLogger("paramiko").setLevel(logging.WARNING)  # Suppresses SSH connection details
logging.getLogger("netmiko").setLevel(logging.WARNING)   # Suppresses Netmiko output

# --- Application Metadata ---
APP_NAME = "Network Device Documentation Utility"
APP_VERSION = "v1.3.5"
VERSION_DATE = "(04-26-2026)"
GITHUB_API_LATEST_RELEASE = "https://api.github.com/repos/RacerJay/nddu/releases/latest"
REPO_URL = "https://github.com/RacerJay/nddu"

# --- Dark mode state ---
DARK_MODE_STATE = True  # Start with dark mode enabled

# --- File Paths (all as Path objects) ---
SCRIPT_DIR = Path(__file__).parent
DEFAULT_INPUT_FOLDER = SCRIPT_DIR / "input"
DEFAULT_OUTPUT_FOLDER = SCRIPT_DIR / "output"
DEFAULT_DEVICE_FILE = DEFAULT_INPUT_FOLDER / "Devices.txt"
DEFAULT_COMMAND_FILE = DEFAULT_INPUT_FOLDER / "Commands.txt"
LOGO_PATH = SCRIPT_DIR / "images" / "nddu.png"
COMBINED_OUTPUT_FILENAME = "Combined.txt"
KEYRING_TOOLS_SCRIPT = "keyring_tools.py"

_TIMESTAMP_RE = re.compile(r'^\d{2}-\d{2}-\d{4}')

def get_existing_clients() -> List[str]:
    """Return sorted list of existing client folders under DEFAULT_OUTPUT_FOLDER.
    Excludes bare timestamp directories (MM-DD-YYYY ...) created when no client was specified."""
    if not DEFAULT_OUTPUT_FOLDER.exists():
        return []
    return sorted(
        d.name for d in DEFAULT_OUTPUT_FOLDER.iterdir()
        if d.is_dir() and not _TIMESTAMP_RE.match(d.name)
    )


def sanitize_folder_name(name: str) -> str:
    """Strip characters that are invalid in folder names across Windows, macOS, and Linux.

    Also rejects '.' and '..' to prevent path traversal when the result is
    joined with a parent path.
    """
    invalid = r'\/:*?"<>|'
    sanitized = "".join(c for c in name if c not in invalid).strip(". ")
    if sanitized in ("", ".", ".."):
        return "Default"
    return sanitized

# --- Logging & Formatting ---
MAX_COMMAND_LENGTH = 256
DIVIDER = '=' * 80
FILLER1 = '!' * 20
FILLER2 = '#' * 20

# --- Command Validation ---
ALLOWED_COMMAND_PREFIXES = {"dir", "mor", "sho", "who"}

# --- Concurrency ---
# Calculate the maximum number of threads to execute concurrently
max_workers = min(20, os.cpu_count() * 2)

# Netmiko device type auto-detection configuration
ENABLE_AUTO_DETECT = False  # Set to False to disable auto-detection
AUTO_DETECT_TIMEOUT = 3  # Seconds to wait for auto-detection

# --- Connection Timeout Settings ---
CONN_TIMEOUT = 10        # Initial connection timeout (seconds)
CONN_TIMEOUT_RETRY = 20  # Retry connection timeout (seconds)
READ_TIMEOUT = 40        # Initial read timeout (seconds)
READ_TIMEOUT_RETRY = 60  # Retry read timeout (seconds)

# Vendors / platforms commonly used with Netmiko
ALLOWED_TYPES = {
    # Cisco families
    "cisco_ios",
    "cisco_xe",
    "cisco_xr",
    "cisco_nxos",
    "cisco_asa",
    "cisco_wlc",
    # Others you might add later
    "arista_eos",
    "juniper_junos",
    "huawei",
}

# --- Built-in Inventory Commands (run when Structured Output is enabled) ---
# These execute alongside the user's Commands.txt. Commands already present in the user's
# file are deduplicated. Failures on a specific platform are silently skipped.
EXCEL_COMMANDS: List[str] = [
    "show version",
    "show switch detail",
    "show interfaces",
    "show interfaces description",   # IOS / IOS-XE
    "show interface description",    # NX-OS
    "show ip interface brief",
    "show interface status",         # NX-OS — all interfaces with status/VLAN/description
    "show interfaces switchport",
    "show mac address-table",
    "show mac address-table count",
    "show arp",                          # IOS / IOS-XE ARP
    "show ip arp",                       # NX-OS ARP
    "show cdp neighbors detail",
    "show vlan brief",               # IOS / IOS-XE
    "show vlan",                     # NX-OS
    "show ip route summary",
    "show spanning-tree summary",
    "show inventory",                # chassis PID / serial (all platforms)
    "show template",                 # IOS / IOS-XE — interface templates
    "show port-profile",             # NX-OS — port profiles
]

# --- Excel Component Registry ---
# Maps each selectable report component to:
#   commands: inventory commands required for this component
#   sheet:    name of the sheet generator function (resolved at report generation time)
# 'Summary' is always included (core device info). The rest are optional.
EXCEL_COMPONENTS: Dict[str, Dict[str, Any]] = {
    'Interfaces':    {'commands': ['show interfaces', 'show interfaces description',
                                   'show interface description', 'show ip interface brief',
                                   'show interface status', 'show interfaces switchport']},
    'Neighbors':     {'commands': ['show cdp neighbors detail']},
    'VLANs':         {'commands': ['show vlan brief', 'show vlan']},
    'Routing':       {'commands': ['show ip route summary']},
    'STP':           {'commands': ['show spanning-tree summary']},
    'MAC Addresses': {'commands': ['show mac address-table', 'show arp', 'show ip arp']},
    'MAC Summary':   {'commands': ['show mac address-table count']},
    'Templates':     {'commands': ['show template', 'show port-profile']},
}

# Commands always collected regardless of component selection (Summary sheet + core metadata)
_CORE_COMMANDS: List[str] = [
    'show version', 'show switch detail', 'show inventory',
]

# All component names in display order
COMPONENT_NAMES: List[str] = list(EXCEL_COMPONENTS.keys())


# Netmiko device-type → ntc-templates platform mapping.
# ntc-templates uses Netmiko platform strings but has no "cisco_xe" templates;
# IOS-XE templates are filed under "cisco_ios".
_NTC_PLATFORM_MAP: Dict[str, str] = {
    "cisco_ios":  "cisco_ios",
    "cisco_xe":   "cisco_ios",      # IOS-XE templates are under cisco_ios
    "cisco_xr":   "cisco_xr",
    "cisco_nxos": "cisco_nxos",
    "cisco_asa":  "cisco_asa",
}

# Commands handled by raw parsers instead of ntc-templates.
# Either no template exists, or the raw parser produces a custom schema
# consumed by downstream code (VLANs sheet, MAC Summary, etc.).
_TEXTFSM_SKIP: Set[str] = {
    'show vlan brief',          # Raw parser: _parse_vlan_brief_raw (custom schema for VLANs sheet + VLAN poll)
    'show vlan',                # NX-OS — ntc-templates has a template but raw parse aligns with downstream
    'show template',            # No ntc-templates coverage — raw parser: _parse_ios_templates_raw
    'show port-profile',        # No ntc-templates coverage — raw parser: _parse_nxos_port_profiles_raw
    'show mac address-table count',  # Raw parser: _parse_mac_count_raw (custom per-VLAN schema)
    'show spanning-tree summary',    # ntc-templates template returns empty — raw parser needed
    'show ip route summary',         # NX-OS template fails on some output; raw parser for consistency
}

def try_textfsm_parse(device_type: str, command: str, raw_output: str) -> Optional[List[Dict[str, Any]]]:
    """
    Attempt TextFSM structured parsing of already-captured command output
    using ntc-templates.  Parses raw text locally — does NOT re-send the
    command to the device.

    Returns:
        List of dicts on success, or None if ntc-templates is not installed,
        no template exists for this platform/command, or parsing fails.
    """
    try:
        from ntc_templates.parse import parse_output  # type: ignore
        platform = _NTC_PLATFORM_MAP.get(device_type)
        if not platform:
            return None
        result = parse_output(platform=platform, command=command, data=raw_output)
        if isinstance(result, list) and result:
            return result
        return None
    except Exception:
        return None

def _parse_nxos_ip_brief_raw(raw: str) -> Dict[str, Any]:
    """
    Raw text fallback for NX-OS 'show ip interface brief'.
    Genie only captures 'unnumbered' entries; this regex catches all L3/SVI interfaces.
    Returns a dict compatible with the Genie 'interface' schema key.
    """
    import re
    result: Dict[str, Any] = {}
    pattern = re.compile(
        r'^(\S+)\s+(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}/\d+|unnumbered)\s+([\w/\-]+)',
        re.MULTILINE
    )
    for m in pattern.finditer(raw):
        intf, ip_addr, status = m.groups()
        result[intf] = {'ip_address': ip_addr, 'interface_status': status}
    return result



def _parse_vlan_brief_raw(raw: str) -> Dict[str, Any]:
    """
    Raw parser for IOS/IOS-XE 'show vlan brief'.
    Replaces Genie parser which suffers from catastrophic regex backtracking
    on long port-list lines (192s+ on macOS ARM vs <1s on Windows).

    Format:
        VLAN Name                             Status    Ports
        ---- -------------------------------- --------- -------------------------------
        1    default                          active    Te5/3, Te5/4
        12   Building_Security                active    Gi1/1, Gi3/48, Gi4/13, ...
                                                        Gi7/12, Gi7/13, ...
        1002 fddi-default                     act/unsup

    Returns dict matching the Genie IOS-XE schema so downstream code
    (Excel report, VLAN poll) works unchanged:
        {vlan: {vlan<id>: {vlan_name, vlan_status, vlan_port: [...]}},
         vlans: {<id>: {name, status, interfaces: [...]}}}
    The 'vlan' key is the Genie IOS-XE schema consumed by the Excel VLANs
    sheet; the 'vlans' key is consumed by the per-VLAN MAC poll lookup.
    """
    genie_vlans: Dict[str, Any] = {}   # vlan<N> keyed (Excel report)
    plain_vlans: Dict[str, Any] = {}   # numeric keyed (VLAN poll)
    current_vlan: Optional[str] = None

    for line in raw.splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        # Skip header / separator lines
        if stripped.startswith('VLAN') and 'Name' in stripped:
            continue
        if stripped.startswith('----'):
            continue

        # New VLAN line: starts with a VLAN number
        m = re.match(r'^(\d+)\s+(\S+)\s+(active|sus(?:pended)?|act/unsup|act/lshut)\s*(.*)', stripped)
        if m:
            vid, name, status, ports_str = m.groups()
            current_vlan = vid
            ports = [p.strip() for p in ports_str.split(',') if p.strip()] if ports_str.strip() else []
            genie_vlans[f'vlan{vid}'] = {
                'vlan_name': name,
                'vlan_status': status,
                'vlan_port': list(ports),
            }
            plain_vlans[vid] = {
                'name': name,
                'status': status,
                'interfaces': list(ports),
            }
        elif current_vlan and f'vlan{current_vlan}' in genie_vlans:
            # Continuation line: more ports for the current VLAN
            ports = [p.strip() for p in stripped.split(',') if p.strip()]
            if ports and all(re.match(r'^[A-Za-z]', p) for p in ports):
                genie_vlans[f'vlan{current_vlan}']['vlan_port'].extend(ports)
                plain_vlans[current_vlan]['interfaces'].extend(ports)
            else:
                current_vlan = None  # Not a continuation line

    if not genie_vlans:
        return {}
    return {'vlan': genie_vlans, 'vlans': plain_vlans}


def _parse_ios_templates_raw(raw: str) -> Dict[str, Any]:
    """
    Raw parser for IOS-XE 'show template'.
    Output is a table: Name  Class  Type, followed by optional indented BOUND: lines.
    Returns {template_name: {type, interfaces: [...]}}
    """
    result: Dict[str, Any] = {}
    current: Optional[str] = None
    in_bound = False

    for line in raw.splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        # Skip header / separator
        if (stripped.startswith('Template') and 'Class' in stripped) or stripped.startswith('---'):
            continue
        # Indented line: BOUND: or continuation interfaces
        if line[0] in (' ', '\t'):
            if current is None:
                continue
            if 'BOUND:' in stripped:
                in_bound = True
                intf_part = stripped.split('BOUND:', 1)[1]
                result[current]['interfaces'].extend(intf_part.split())
            elif in_bound:
                result[current]['interfaces'].extend(stripped.split())
        else:
            # Non-indented: new template line or device prompt
            in_bound = False
            parts = stripped.split()
            if len(parts) >= 3:
                current = parts[0]
                result[current] = {'type': parts[2], 'interfaces': []}
            else:
                current = None  # device prompt or unrecognised line

    return result


def _parse_nxos_port_profiles_raw(raw: str) -> Dict[str, Any]:
    """
    Raw parser for NX-OS 'show port-profile'.
    Handles two output formats:

    Format A (newer NX-OS): property-per-line block
        port-profile <name>
          type: <type>
          status: <status>
          assigned interfaces:
            Ethernet1/1

    Format B (older NX-OS): config-style block + interface inherit stanzas
        port-profile type <type> <name>
          <config lines>
          state enabled|disabled
        interface <range>
          inherit port-profile <name>

    Returns {profile_name: {type, status, interfaces: [...]}}
    """
    import re
    result: Dict[str, Any] = {}
    current: Optional[str] = None
    in_assigned = False
    current_intf: Optional[str] = None  # Format B: active interface block

    for line in raw.splitlines():
        stripped = line.strip()

        if not stripped:
            current_intf = None  # blank line resets interface context
            continue

        # Format B: port-profile type <type> <name>  (more specific — check first)
        m_b = re.match(r'^port-profile\s+type\s+(\S+)\s+(\S+)', line)
        if m_b:
            current = m_b.group(2)
            result[current] = {'type': m_b.group(1).capitalize(), 'status': 'N/A', 'interfaces': []}
            in_assigned = False
            current_intf = None
            continue

        # Format A: port-profile <name>
        m_a = re.match(r'^port-profile\s+(\S+)', line)
        if m_a:
            current = m_a.group(1)
            result[current] = {'type': 'N/A', 'status': 'N/A', 'interfaces': []}
            in_assigned = False
            current_intf = None
            continue

        # Format B: non-indented interface block
        m_intf = re.match(r'^interface\s+(.+)', line)
        if m_intf:
            current_intf = m_intf.group(1).strip()
            continue

        # Format B: lines inside an interface block
        if current_intf is not None:
            m_inherit = re.match(r'\s+inherit\s+port-profile\s+(\S+)', line)
            if m_inherit:
                prof_name = m_inherit.group(1)
                if prof_name in result:
                    result[prof_name]['interfaces'].append(current_intf)
            continue  # all other indented interface-block lines are ignored

        if current is None:
            continue

        # Format A properties / Format B state line
        if stripped.startswith('type:'):
            result[current]['type'] = stripped.split(':', 1)[1].strip().capitalize() or 'N/A'
        elif stripped.startswith('status:'):
            result[current]['status'] = stripped.split(':', 1)[1].strip() or 'N/A'
        elif stripped.startswith('state '):
            result[current]['status'] = stripped.split(None, 1)[1].strip()  # 'enabled'/'disabled'
        elif stripped == 'assigned interfaces:':
            in_assigned = True
        elif stripped.endswith(':'):
            in_assigned = False  # any other section header ends the interfaces block
        elif in_assigned:
            result[current]['interfaces'].append(stripped)

    return result


def _parse_nxos_ip_arp_raw(raw: str) -> Dict[str, str]:
    """
    Raw parser for NX-OS 'show ip arp'.
    Returns {mac_address: ip_address} for use as an ARP lookup map.
    Table format:
        Address         Age       MAC Address     Interface
        10.24.250.3     00:00:36  xxxx.xxxx.xxxx  Vlan1401
    """
    result: Dict[str, str] = {}
    in_table = False
    for line in raw.splitlines():
        stripped = line.strip()
        if stripped.startswith('Address') and 'MAC Address' in stripped:
            in_table = True
            continue
        if not in_table or not stripped:
            continue
        # Skip flag/header lines that start with a capital letter word (not an IP)
        parts = stripped.split()
        if len(parts) >= 3 and re.match(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$', parts[0]):
            ip_addr = parts[0]
            mac     = parts[2]
            result[mac] = ip_addr
    return result


def _parse_mac_count_raw(raw: str, device_type: str) -> Dict[str, Any]:
    """
    Raw parser for 'show mac address-table count'.

    IOS-XE returns per-VLAN blocks plus a grand total:
        Mac Entries for Vlan 14:
          Dynamic Address Count  : 38
          Static  Address Count  : 1
          Total Mac Addresses    : 39
        Total Dynamic Address Count  : 48
        Total Static  Address Count  : 2
        Total Mac Address In Use     : 50
        Total Mac Address Space Available: 32718

    NX-OS returns only a grand total:
        MAC Entries for all vlans :
          Total Address Count:   1846
          Dynamic Address Count: 1846
          Static Address (User-defined) Count: 0
          Secure Address Count:  0

    Returns:
        {
          'vlans': {'14': {'dynamic': 38, 'static': 1, 'total': 39}, ...},   # IOS-XE only
          'totals': {'dynamic': N, 'static': N, 'total': N},
        }
    """
    result: Dict[str, Any] = {'vlans': {}, 'totals': {}}
    current_vlan: Optional[str] = None

    def _extract_int(s: str) -> int:
        m = re.search(r'(\d+)', s)
        return int(m.group(1)) if m else 0

    for line in raw.splitlines():
        stripped = line.strip()
        if not stripped:
            continue

        # IOS-XE standard: "Mac Entries for Vlan N:"
        # Chassis IOS per-VLAN: "MAC Entries for Vlan 254:"
        m_vlan = re.match(r'Mac Entries for Vlan\s+(\d+)\s*:', stripped, re.IGNORECASE)
        if m_vlan:
            current_vlan = m_vlan.group(1)
            result['vlans'][current_vlan] = {'dynamic': 0, 'static': 0, 'total': 0}
            continue

        # "MAC Entries for all vlans" — chassis IOS/NX-OS totals-only aggregate
        if re.match(r'MAC Entries for all', stripped, re.IGNORECASE):
            current_vlan = None
            continue

        val = _extract_int(stripped)
        low = stripped.lower()

        if current_vlan:
            # Per-VLAN counts (IOS-XE standard — single dynamic, static, total line each)
            if 'dynamic' in low:
                result['vlans'][current_vlan]['dynamic'] = val
            elif 'static' in low:
                result['vlans'][current_vlan]['static'] = val
            elif 'total' in low:
                result['vlans'][current_vlan]['total'] = val
                current_vlan = None  # VLAN block ends after its total line
        else:
            # Grand totals — skip non-unicast and non-count lines
            if any(k in low for k in ('multicast', 'available', 'secure', 'overlay')):
                pass
            elif 'total' in low and 'dynamic' in low and 'count' in low:
                # "Total Dynamic Address Count" — IOS-XE explicit grand total
                result['totals']['dynamic'] = val
            elif 'total' in low and 'static' in low and 'count' in low:
                # "Total Static Address Count" — IOS-XE explicit grand total
                result['totals']['static'] = val
            elif 'total' in low and ('in use' in low or
                    ('address count' in low and 'dynamic' not in low and 'static' not in low)):
                # "Total Mac Address In Use" / "Total Address Count" (N77)
                result['totals']['total'] = val
            elif 'dynamic' in low and 'count' in low:
                # Accumulate: handles N9K (Local + Remote) and chassis IOS (single line)
                result['totals']['dynamic'] = result['totals'].get('dynamic', 0) + val
            elif 'static' in low and 'count' in low:
                # Accumulate: handles N9K (Local + Remote) and chassis IOS (User + System defined)
                result['totals']['static'] = result['totals'].get('static', 0) + val

    # Derive total if not explicitly set (e.g. N77 per-VLAN, some NX-OS variants)
    totals = result['totals']
    if not totals.get('total') and (totals.get('dynamic') or totals.get('static')):
        totals['total'] = totals.get('dynamic', 0) + totals.get('static', 0)

    return result


def _parse_mac_count_vlan_raw(raw: str) -> Dict[str, int]:
    """
    Parse 'show mac address-table count vlan N' output for a single VLAN.

    Handles multiple platforms:
        IOS-XE standard:  Dynamic Address Count  : 38 / Static Address Count : 1 / Total Mac Addresses : 39
        Chassis IOS:      Dynamic Unicast Address Count : 90
                          Static Unicast Address (User-defined) Count : 0
                          Static Unicast Address (System-defined) Count : 1
                          Total Unicast MAC Addresses In Use : 91
        N77 NX-OS:        Dynamic Address Count: 137  (no explicit total — derived)
        N9K NX-OS:        Dynamic Local Address Count: 54 / Dynamic Remote Address Count: 0
                          Total MAC Addresses in Use (DLAC+...): 54

    Returns: {'dynamic': N, 'static': N, 'total': N}
    Dynamic = sum of all Dynamic lines (handles Local+Remote on N9K)
    Static  = sum of all Static lines (handles User+System defined on chassis IOS)
    Total   = explicit total-in-use line, or dynamic+static if absent
    """
    dynamic = 0
    static  = 0
    total   = 0

    def _val(s: str) -> int:
        m = re.search(r'(\d+)', s)
        return int(m.group(1)) if m else 0

    for line in raw.splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        low = stripped.lower()

        # Skip non-unicast and non-count lines
        if any(k in low for k in ('multicast', 'available', 'secure', 'overlay')):
            continue

        val = _val(stripped)

        if 'total' in low and ('in use' in low or
                ('address count' in low and 'dynamic' not in low and 'static' not in low)):
            # Explicit total: "Total Unicast MAC Addresses In Use", "Total Address Count"
            total = val
        elif 'dynamic' in low and 'count' in low:
            # Accumulate: handles N9K Local+Remote
            dynamic += val
        elif 'static' in low and 'count' in low:
            # Accumulate: handles chassis IOS User-defined + System-defined
            static += val

    # Derive total when no explicit total line (e.g. N77 per-VLAN output)
    if not total and (dynamic or static):
        total = dynamic + static

    return {'dynamic': dynamic, 'static': static, 'total': total}


def _parse_ip_route_summary_raw(raw: str) -> Optional[Dict[str, Any]]:
    """
    Raw parser for 'show ip route summary' (IOS/IOS-XE and NX-OS).

    IOS/IOS-XE format:
        Route Source    Networks    Subnets     ...
        connected       0           7           ...
        ospf 1          0           45          ...
        Total           3           53          ...

    NX-OS format:
        Total number of routes: 156
        Total number of paths:  162
        Best paths per protocol:
          ospf-10          141
          static           2

    Returns dict with 'protocols' list and 'total' dict per VRF.
    """
    result: Dict[str, Any] = {'protocols': [], 'total': {}}

    # Detect NX-OS format
    if 'Best paths per protocol' in raw or 'Total number of routes' in raw:
        for line in raw.splitlines():
            stripped = line.strip()
            m = re.match(r'^Total number of routes:\s*(\d+)', stripped)
            if m:
                result['total']['routes'] = int(m.group(1))
            m = re.match(r'^Total number of paths:\s*(\d+)', stripped)
            if m:
                result['total']['paths'] = int(m.group(1))
            # Protocol lines: "  ospf-10   141"
            m = re.match(r'^\s{2,}(\S+)\s+(\d+)', line)
            if m:
                proto_key = m.group(1)
                count = int(m.group(2))
                # Split instance from protocol: ospf-10 → ospf, 10
                dash = proto_key.rfind('-')
                if dash > 0 and proto_key[dash + 1:].isdigit():
                    proto, instance = proto_key[:dash], proto_key[dash + 1:]
                else:
                    proto, instance = proto_key, '-'
                result['protocols'].append({
                    'protocol': proto, 'instance': instance, 'routes': count
                })
    else:
        # IOS/IOS-XE tabular format
        in_table = False
        for line in raw.splitlines():
            stripped = line.strip()
            if stripped.startswith('Route Source'):
                in_table = True
                continue
            if not in_table or not stripped:
                continue
            parts = stripped.split()
            if len(parts) < 3:
                continue
            # Handle "ospf 1  0  45 ..." (protocol + instance in first two columns)
            proto_key = parts[0]
            try:
                int(parts[1])
                # parts[1] is numeric → flat protocol (connected, static, etc.)
                networks = int(parts[1])
                subnets = int(parts[2]) if len(parts) > 2 and parts[2].isdigit() else 0
                proto, instance = proto_key, '-'
            except ValueError:
                # parts[1] is not numeric → instance ID (e.g. "ospf 1")
                # Skip lines where the numeric columns are also non-numeric (e.g. VRF/context headers)
                if not parts[2].isdigit():
                    continue
                instance = parts[1]
                proto = proto_key
                networks = int(parts[2])
                subnets = int(parts[3]) if len(parts) > 3 and parts[3].isdigit() else 0
            routes = networks + subnets
            if proto.lower() == 'total':
                result['total']['routes'] = routes
            else:
                if routes > 0 or instance != '-':
                    result['protocols'].append({
                        'protocol': proto, 'instance': instance, 'routes': routes
                    })
    return result if (result['protocols'] or result['total']) else None


def _parse_stp_summary_raw(raw: str) -> Optional[Dict[str, Any]]:
    """
    Raw parser for 'show spanning-tree summary' (IOS/IOS-XE and NX-OS).
    ntc-templates returns empty for this command.

    Returns dict with: mode, root_bridge_for, num_vlans, blocking, forwarding, stp_active
    """
    result: Dict[str, Any] = {}
    for line in raw.splitlines():
        stripped = line.strip()
        low = stripped.lower()
        if low.startswith('switch is in') and 'mode' in low:
            # "Switch is in rapid-pvst mode"
            result['mode'] = stripped.split('in', 1)[1].replace('mode', '').strip()
        elif low.startswith('root bridge for:'):
            result['root_bridge_for'] = stripped.split(':', 1)[1].strip()
        elif re.match(r'^\d+\s+vlans?\s+', low):
            # Summary totals line: "3 vlans    1    0    0    40    41"
            parts = stripped.split()
            if len(parts) >= 7:
                try:
                    result['num_vlans'] = int(parts[0])
                    result['blocking'] = int(parts[2])
                    result['forwarding'] = int(parts[5])
                    result['stp_active'] = int(parts[6])
                except (ValueError, IndexError):
                    pass
    return result if result else None


# --- Add supports for VERBOSE logging level 15 ---
VERBOSE_LEVEL_NUM = 15  # Between INFO(20) and DEBUG(10)
logging.addLevelName(VERBOSE_LEVEL_NUM, "VERBOSE")

def verbose(self, message: str, *args: Any, **kwargs: Any) -> None:
    """Log a message with severity 'VERBOSE' (level 15)."""
    if self.isEnabledFor(VERBOSE_LEVEL_NUM):
        self._log(VERBOSE_LEVEL_NUM, message, args, **kwargs)

logging.Logger.verbose = verbose

def format_time(dt: Optional[datetime] = None) -> str:
    """
    Format a datetime object into a standardized string format.
    
    Args:
        dt: Optional datetime object to format. If None, uses current time.
        
    Returns:
        Formatted datetime string in 'Day MM/DD/YYYY - HH:MM:SS AM/PM' format
    """
    if dt is None:
        dt = datetime.now()
    return dt.strftime('%a %m/%d/%Y - %I:%M:%S %p')

def detect_device_type(host: str, username: str, password: str, enable_password: Optional[str] = None) -> Optional[str]:
    """
    Two-phase device type detection.

    Phase 1 (fast): Connect as generic cisco_ios, run 'show version' once, and
    pattern-match. Handles all Cisco platforms in ~5-8s.

    Phase 2 (fallback): Netmiko SSHDetect for non-Cisco devices or when Phase 1
    connection fails. SSHDetect iterates 40+ device types (~7s each) so this path
    is slow but comprehensive.

    The old SSHDetect-only approach took 108-115s for Catalyst 4500 IOS-XE chassis
    switches because their 'show version' output says "Cisco IOS Software" (not
    "Cisco IOS XE Software"), matching cisco_ios at priority 95 — too low for
    early exit — forcing a full iteration of all device types.
    """
    logger = logging.getLogger(__name__)
    logger.info(f"Detecting device type on device: {host}")
    start_time = time.time()

    # --- Phase 1: Fast Cisco detection via 'show version' ---
    conn = None
    try:
        conn = ConnectHandler(
            device_type='cisco_ios',
            host=host,
            username=username,
            password=password,
            secret=enable_password,
            conn_timeout=CONN_TIMEOUT,
            banner_timeout=CONN_TIMEOUT + 5,
            read_timeout_override=READ_TIMEOUT,
            global_delay_factor=0.5,
            fast_cli=True
        )
        ver = conn.send_command('show version', read_timeout=15)
        elapsed = time.time() - start_time

        # Pattern match in priority order (most specific first)
        detected = None
        if 'NX-OS' in ver or ('Nexus' in ver and 'Cisco' in ver):
            detected = 'cisco_nxos'
        elif 'IOS-XE' in ver or 'IOS XE' in ver:
            detected = 'cisco_xe'
        elif 'IOS-XR' in ver or 'IOS XR' in ver:
            detected = 'cisco_xr'
        elif 'Adaptive Security Appliance' in ver:
            detected = 'cisco_asa'
        elif 'Arista' in ver:
            detected = 'arista_eos'
        elif 'Cisco IOS Software' in ver or 'Cisco Internetwork Operating System' in ver:
            detected = 'cisco_ios'

        if detected:
            if detected not in ALLOWED_TYPES:
                raise ValueError(f"Detected unsupported type '{detected}' for {host}")
            logger.verbose(f"Device {host} detected as type '{detected}' in {elapsed:.2f}s")
            return detected
        else:
            logger.debug(f"Fast detection: 'show version' did not match known patterns for {host}")
    except Exception as e:
        logger.debug(f"Fast device detection failed for {host}: {e}")
    finally:
        if conn:
            try:
                conn.disconnect()
            except Exception:
                pass

    # --- Phase 2: Fallback to Netmiko SSHDetect ---
    # Handles non-Cisco devices (Arista, Juniper, Huawei, etc.) and edge cases
    # where the Phase 1 cisco_ios connection fails.
    try:
        logger.debug(f"Falling back to SSHDetect for {host}")
        detector = SSHDetect(
            device_type='autodetect',
            host=host,
            username=username,
            password=password,
            secret=enable_password,
            conn_timeout=AUTO_DETECT_TIMEOUT,
            banner_timeout=AUTO_DETECT_TIMEOUT,
            read_timeout_override=AUTO_DETECT_TIMEOUT,
            global_delay_factor=0.25
        )

        best_match = detector.autodetect()

        if not best_match:
            raise ValueError(f"Could not auto-detect device type for {host}")

        # NX-OS can occasionally be misidentified as cisco_ios due to SSH fingerprint
        # timing. NX-OS patterns (NX-OS, Nexus) will never appear in real IOS output,
        # so any non-zero NX-OS score is a reliable correction signal.
        if best_match == 'cisco_ios':
            nxos_score = detector.potential_matches.get('cisco_nxos', 0)
            if nxos_score > 0:
                best_match = 'cisco_nxos'
                logger.verbose(f"Corrected cisco_ios -> cisco_nxos for {host} "
                               f"(NX-OS scored {nxos_score} in potential_matches)")

        if best_match not in ALLOWED_TYPES:
            raise ValueError(f"Auto-detected unsupported type '{best_match}' for {host}")

        elapsed = time.time() - start_time
        logger.verbose(f"Device {host} detected as type '{best_match}' in {elapsed:.2f}s (SSHDetect fallback)")
        logger.verbose(f"Potential matches: {detector.potential_matches}")
        return best_match

    except Exception as e:
        logger.debug(f"Device auto-detection failed for {host}: {e}")

    return None

DEVICE_CACHE_FILENAME = "device_cache.json"
_device_cache_lock = threading.Lock()  # Serializes concurrent cache reads/writes


def _load_device_cache(cache_path: Path) -> Dict[str, str]:
    """Load cached device types from a JSON file. Returns {ip: device_type}."""
    if cache_path.exists():
        try:
            with open(cache_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            if isinstance(data, dict):
                return data
        except Exception:
            pass
    return {}


def _save_device_cache(cache_path: Path, cache: Dict[str, str]) -> None:
    """Save device type cache to a JSON file."""
    try:
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        with open(cache_path, 'w', encoding='utf-8') as f:
            json.dump(cache, f, indent=2, sort_keys=True)
    except Exception:
        pass


def get_device_type(host: str, username: str, password: str,
                   enable_password: Optional[str] = None, enable_structured_output: bool = False,
                   cache_path: Optional[Path] = None, force_redetect: bool = False) -> str:
    """
    Get device type with cache and fallback logic.

    Args:
        host: Device hostname or IP
        username: SSH username
        password: SSH password
        enable_password: Enable password (optional)
        enable_structured_output: When True, auto-detection runs to determine the
            correct OS type (required for accurate structured parsing)
        cache_path: Path to device_cache.json for this client. If provided,
            cached types are used and new detections are saved.
        force_redetect: When True, ignore cached type and re-detect.
    """
    logger = logging.getLogger(__name__)

    if not enable_structured_output:
        return 'cisco_ios'

    # Check cache first (unless force re-detect) — lock protects against concurrent reads/writes
    if cache_path and not force_redetect:
        with _device_cache_lock:
            cache = _load_device_cache(cache_path)
            cached_type = cache.get(host)
        if cached_type and cached_type in ALLOWED_TYPES:
            logger.verbose(f"Cache hit: {host} -> '{cached_type}' ({cache_path.name})")
            return cached_type

    # Detection runs outside the lock — it's slow (SSH) and doesn't touch the cache
    detected_type = detect_device_type(host, username, password, enable_password)

    final_type = detected_type if detected_type else 'cisco_ios'
    if not detected_type:
        logger.verbose(f"Auto-detection failed for {host}, using cisco_ios as fallback")

    # Save result to cache — re-read under lock so concurrent writes aren't lost
    if cache_path:
        with _device_cache_lock:
            cache = _load_device_cache(cache_path)
            action = "updated" if host in cache and cache[host] != final_type else "saved"
            cache[host] = final_type
            _save_device_cache(cache_path, cache)
        logger.verbose(f"Cache {action}: {host} -> '{final_type}' ({cache_path.name})")

    return final_type

class AllowedCommands:
    """Class to validate and manage allowed command prefixes."""
    
    def __init__(self) -> None:
        """Initialize with default allowed command prefixes."""
        self.allowed_prefixes: Set[str] = ALLOWED_COMMAND_PREFIXES

    def is_command_allowed(self, command: str) -> bool:
        """
        Check if a command is allowed based on its prefix and length.
        
        Args:
            command: The command string to validate
            
        Returns:
            True if command is allowed, False otherwise
        """
        command = command.strip()
        
        # Check length first
        if len(command) > MAX_COMMAND_LENGTH:
            return False
            
        # Then check prefix
        prefix = command[:3].lower()  # Extract the first 3 characters of the command
        return prefix in self.allowed_prefixes

class CustomFormatter(logging.Formatter):
    """Custom log formatter with different formats for file and GUI output."""
    
    def __init__(self, fmt: Optional[str] = None, datefmt: Optional[str] = None, style: str = '%') -> None:
        """Initialize the formatter with format styles for different destinations."""
        super().__init__(fmt, datefmt, style)
        # Format styles for different destinations and levels
        self.file_formats: Dict[int, str] = {
            VERBOSE_LEVEL_NUM: "%(levelname)s: %(message)s",
            logging.INFO: "%(message)s",
            logging.DEBUG: "%(levelname)s: %(message)s",
            logging.WARNING: "%(levelname)s: %(message)s",
            logging.ERROR: "%(levelname)s: %(message)s",
            logging.CRITICAL: "%(levelname)s: %(message)s"
        }
        self.gui_formats: Dict[int, str] = {
            VERBOSE_LEVEL_NUM: "%(levelname)s: %(message)s",
            logging.INFO: "%(message)s",
            logging.DEBUG: "%(levelname)s: %(message)s",
            logging.WARNING: "%(levelname)s: %(message)s",
            logging.ERROR: "%(levelname)s: %(message)s",
            logging.CRITICAL: "%(levelname)s: %(message)s"
        }

    def format(self, record: logging.LogRecord) -> str:
        """
        Format the specified record according to destination.
        
        Args:
            record: Log record to format
            
        Returns:
            Formatted log message
        """
        # Store original format
        original_format = self._style._fmt
        
        # Choose format based on handler type
        if getattr(record, 'for_gui', False):
            self._style._fmt = self.gui_formats.get(record.levelno, "%(message)s")
        else:
            self._style._fmt = self.file_formats.get(record.levelno, "%(levelname)s: %(message)s")
        
        # Format the message
        result = super().format(record)
        
        # Restore original format
        self._style._fmt = original_format
        
        return result

class LogEmitter(QObject):
    """Qt signal emitter for logging messages to the GUI."""
    log_signal = Signal(str, str)  # (message, levelname)

class GUILogHandler(logging.Handler):
    """Custom logging handler that buffers log records for GUI display.

    Messages are placed into a thread-safe queue rather than emitted as Qt
    signals directly.  A QTimer on the main thread drains the queue at a fixed
    interval (~150 ms), which prevents hundreds of cross-thread signal
    deliveries per second from starving the Qt event loop.
    """

    def __init__(self) -> None:
        super().__init__()
        self.emitter  = LogEmitter()          # kept for any legacy signal uses
        self._queue: queue.SimpleQueue = queue.SimpleQueue()
        self.setFormatter(CustomFormatter())

    def emit(self, record: logging.LogRecord) -> None:
        try:
            record.for_gui = True
            msg = self.format(record)
            self._queue.put_nowait((msg, record.levelname))
        except Exception as e:
            print(f"Logging error: {e}")

class Worker(QThread):
    """Worker thread for executing device documentation tasks."""
    
    progress_signal = Signal(int)  # Progress percentage
    log_signal = Signal(str)       # Log messages
    completion_signal = Signal(str, bool)  # Completion signal with output folder and success flag
    stopped_signal = Signal()      # Signal for clean stop

    def __init__(self, device_file: str, command_file: str, credentials: Dict[str, str],
                 enable_password: str, output_folder: Path, verbose_enabled: bool,
                 create_combined_output: bool = False, enable_structured_output: bool = False,
                 selected_components: Optional[Set[str]] = None,
                 force_redetect: bool = False, save_txt: bool = True) -> None:
        """
        Initialize the worker thread.

        Args:
            device_file: Path to device list file
            command_file: Path to command list file
            credentials: Dictionary with 'username' and 'password'
            enable_password: Enable/privileged password
            output_folder: Directory for output files
            verbose_enabled: Whether verbose logging is enabled
            create_combined_output: Whether to create combined output file
            enable_structured_output: Enable structured parsing (TextFSM), device type
                auto-detection, inventory commands, and per-device JSON output
            selected_components: Set of component names to include in report.
                None or empty means all components.
            force_redetect: When True, ignore cached device types and re-detect all.
            save_txt: When True, write per-device raw .txt output files.
        """
        super().__init__()
        self.device_file = device_file
        self.command_file = command_file
        self.credentials = credentials
        self.enable_password = enable_password
        self.output_folder = output_folder
        self.verbose_enabled = verbose_enabled
        self.create_combined_output = create_combined_output
        self.enable_structured_output = enable_structured_output
        self.selected_components: Set[str] = selected_components or set(COMPONENT_NAMES)
        self.force_redetect = force_redetect
        self.save_txt = save_txt
        self._is_cancelled = False  # Cancellation flag
        self.active_connections: List[Any] = []  # Track active connections
        self._connections_lock = threading.Lock()  # Guard active_connections and output_files
        self.output_files: List[Path] = []  # Track output files (pre-init here for lock coverage)

        # Configure logging
        self.logger = logging.getLogger(__name__)
        if verbose_enabled:
            self.logger.setLevel("VERBOSE")
        else:
            self.logger.setLevel("INFO")

    def validate_credentials(self, devices: List[str]) -> bool:
        """
        Validate credentials by attempting to connect to the first reachable device.
        
        Args:
            devices: List of device IPs to try
            
        Returns:
            True if credentials are valid, False otherwise
        """
        if not devices:
            self.logger.error(f"No valid devices found in the Devices input file.")
            return False

        # Try to connect to each device until a reachable one is found
        for device in devices:
            self.logger.info(f"Testing credentials on device: {device}")
            try:
                # Attempt to connect to the device
                connection = ConnectHandler(
                    device_type='cisco_ios',
                    host=device,
                    username=self.credentials['username'],
                    password=self.credentials['password'],
                    secret=self.enable_password,
                    banner_timeout=60,  # Set a timeout to wait for the SSH banner, 0 to skip banner
                    conn_timeout=5,  # Connection timeout (seconds)
                    read_timeout_override=5,  # Read timeout (seconds)
                    global_delay_factor=0.5,  # Reduce delay factor for faster response
                    fast_cli=True
                )
                connection.enable()
                connection.disconnect()
                self.logger.info(f"Credentials validated successfully on device: {device}")
                return True
            except NetmikoTimeoutException:
                self.logger.warning(f"Device unreachable: {device}")
                continue  # Skip to the next device
            except NetmikoAuthenticationException:
                self.logger.error(f"Invalid credentials for device: {device}")
                return False  # Credentials are invalid, no need to try other devices
            except Exception as e:
                self.logger.error(f"Unexpected error validating credentials on device {device}: {e}")
                continue  # Skip to the next device

        # If no reachable devices were found
        self.logger.error(f"No reachable devices found in the Device(s) input file.")
        return False

    def run(self) -> None:
        """Main execution method for the worker thread."""
        try:
            start_time = datetime.now()
            start_time_str = format_time(start_time)
            self.logger.info(f"***** Script started - {start_time_str} *****")
            self.logger.info(f"{DIVIDER}")
            self.logger.info(f"{Path(__file__).name} {APP_VERSION} {VERSION_DATE}")
            self.logger.info(f"{DIVIDER}")

            # Read and validate devices
            devices = self.read_devices(self.device_file)
            
            # Read and validate commands
            commands = self.read_commands(self.command_file)

            # Check if we have at least one valid device and one valid command
            if not devices or not commands:
                self.logger.error(f"Script cannot proceed - no valid devices or commands found")
                end_time = datetime.now()
                end_time_str = format_time(end_time)
                total_time = end_time - start_time
                
                self.logger.info(f"{DIVIDER}")
                self.logger.info(f"Script Summary:")
                self.logger.info(f"  Valid devices found: {len(devices)}")
                self.logger.info(f"  Valid commands found: {len(commands)}")
                self.logger.info(f"  Script run time (h:mm:ss.ms): {total_time}")
                self.logger.info(f"{DIVIDER}")
                self.logger.info(f"***** Script ended - {end_time_str} *****")
                self.completion_signal.emit(str(self.output_folder), False)
                return

            total_devices = len(devices)
            successful_devices = 0
            failed_devices = 0
            total_commands = len(commands)
            successful_commands = 0
            failed_commands = 0

            # Validate credentials before proceeding
            if not self.validate_credentials(devices):
                self.logger.error(f"Script terminated due to credential validation failure.")
                end_time = datetime.now()
                end_time_str = format_time(end_time)
                total_time = end_time - start_time

                self.logger.info(f"{DIVIDER}")
                self.logger.info(f"Script Summary:")
                self.logger.info(f"  Script run time (h:mm:ss.ms): {total_time}")
                self.logger.info(f"{DIVIDER}")
                self.logger.info(f"***** Script ended - {end_time_str} *****")
                self.completion_signal.emit(str(self.output_folder), False)
                return

            # Log Verbose output
            if self.verbose_enabled:
                self.logger.verbose(f"Verbose Status: {self.verbose_enabled}")
                self.logger.verbose(f"Credentials - Username: {self.credentials['username']}")
                self.logger.verbose(f"Credentials - Password: [hidden]")
                self.logger.verbose(f"Credentials - Enable Password: {'[hidden]' if self.enable_password else '(none)'}")
                self.logger.verbose(f'Input File - Devices: "{self.device_file}"')
                self.logger.verbose(f'Input File - Commands: "{self.command_file}"')
                self.logger.verbose(f"Structured Output: {self.enable_structured_output}")
                self.logger.verbose(f"CPU Count: {os.cpu_count()}, max_workers: {max_workers}")

            # Device type cache lives at the client folder level (parent of run folder).
            # For no-client runs: output/device_cache.json
            # For client runs:    output/{Client}/device_cache.json
            self._cache_path: Optional[Path] = None
            if self.enable_structured_output:
                self._cache_path = self.output_folder.parent / DEVICE_CACHE_FILENAME
                if self.force_redetect:
                    self.logger.info("Force re-detect enabled — ignoring cached device types.")
                elif self._cache_path.exists():
                    cache = _load_device_cache(self._cache_path)
                    if cache:
                        self.logger.info(f"Device type cache loaded: {len(cache)} entries from {self._cache_path.name}")

            # Pre-warm ntc-templates import so the template index loads once in the
            # main thread before workers start parsing concurrently.
            if self.enable_structured_output:
                try:
                    from ntc_templates.parse import parse_output  # type: ignore  # noqa: F401
                    self.logger.verbose("ntc-templates loaded successfully.")
                except ImportError:
                    self.logger.warning("ntc-templates not installed — structured parsing will be limited to raw parsers.")

            with concurrent.futures.ThreadPoolExecutor(max_workers) as executor:
                futures = {}
                for i, device in enumerate(devices, start=1):  # Track device count
                    future = executor.submit(self.process_device, device, commands, i, total_devices)
                    futures[future] = device

                for future in concurrent.futures.as_completed(futures):
                    try:
                        device, success, commands_executed = future.result()
                        if success:
                            successful_devices += 1
                            successful_commands += commands_executed
                        else:
                            failed_devices += 1
                            failed_commands += commands_executed
                        self.progress_signal.emit(int((successful_devices + failed_devices) / total_devices * 100))
                    except Exception as e:
                        self.logger.error(f"Error processing device: {e}")
                        failed_devices += 1

            # Only create combined output if enabled
            if self.create_combined_output and self.output_files:
                self.create_combined_output_file()

            # Generate Excel report from per-device JSON files if structured output enabled
            if self.enable_structured_output:
                generate_excel_report(self.output_folder, self.logger, self.selected_components)

            end_time = datetime.now()
            end_time_str = format_time(end_time)
            total_time = end_time - start_time

            self.logger.info(f"{DIVIDER}")
            self.logger.info(f"Script Summary:")
            self.logger.info(f"  Total commands per device: {total_commands}")
            self.logger.info(f"  Successful devices: {successful_devices}")
            self.logger.info(f"  Failed devices: {failed_devices}")
            self.logger.info(f"  Script run time (h:mm:ss.ms): {total_time}")
            self.logger.info(f"{DIVIDER}")
            self.logger.info(f"***** Script ended - {end_time_str} *****")

            self.completion_signal.emit(str(self.output_folder), failed_devices == 0)

        except Exception as e:
            self.logger.error(f"{e}")
            self.completion_signal.emit(str(self.output_folder), False)

    def read_devices(self, device_file: str) -> List[str]:
        """
        Read and validate the Device(s) input file.
        
        Args:
            device_file: Path to the device list file
            
        Returns:
            List of unique, valid device IPs
        """
        devices: List[str] = []
        seen_devices: Set[str] = set()
        valid_devices_found = False

        try:
            with open(device_file, 'r', encoding='utf-8') as file:
                for line_number, line in enumerate(file, start=1):
                    line = line.strip()
                    if '#' in line:
                        line = line.split('#', 1)[0]
                    
                    line = line.strip()
                    if not line:  # Skip empty lines
                        continue

                    # Check for valid IP address
                    try:
                        ipaddress.ip_address(line)
                        # Check for duplicate IP addresses
                        if line in seen_devices:
                            self.logger.warning(f"Line {line_number}: Duplicate IP address - {line} (ignored)")
                            continue
                        seen_devices.add(line)
                        # If valid, add to the list of devices
                        devices.append(line)
                        valid_devices_found = True
                    except ValueError:
                        self.logger.warning(f"Line {line_number}: Invalid IP address - {line} (ignored)")

            if not valid_devices_found:
                self.logger.error(f"No valid device IP addresses found in the Device(s) input file")
            return devices

        except Exception as e:
            self.logger.error(f"Error reading device file: {e}")
            return []

    def read_commands(self, command_file: str) -> List[str]:
        """
        Read and validate the Command(s) input file.
        
        Args:
            command_file: Path to the command list file
            
        Returns:
            List of unique, valid commands
        """
        commands: List[str] = []
        seen_commands: Set[str] = set()
        valid_commands_found = False
        allowed_commands = AllowedCommands()

        try:
            with open(command_file, 'r', encoding='utf-8') as file:
                for line_number, line in enumerate(file, start=1):
                    line = line.strip()
                    if '#' in line:
                        line = line.split('#', 1)[0]
                    
                    line = line.strip()
                    if not line:  # Skip empty lines
                        continue

                    # Check command length
                    if len(line) > MAX_COMMAND_LENGTH:
                        self.logger.warning(f"Line {line_number}: Command too long (>{MAX_COMMAND_LENGTH} chars) - {line[:20]}... (ignored)")
                        continue

                    # Check for valid command
                    if not allowed_commands.is_command_allowed(line):
                        self.logger.warning(f"Line {line_number}: Invalid command - {line} (ignored)")
                        continue

                    # Check for duplicate commands
                    if line in seen_commands:
                        self.logger.warning(f"Line {line_number}: Duplicate command - {line} (ignored)")
                        continue

                    seen_commands.add(line)
                    commands.append(line)
                    valid_commands_found = True

            if not valid_commands_found:
                self.logger.error(f"No valid commands found in the Command(s) input file")
            return commands

        except Exception as e:
            self.logger.error(f"Error reading command file: {e}")
            return []

    def cancel(self) -> None:
        """Signal the worker to stop execution and clean up connections."""
        self._is_cancelled = True
        self.logger.info(f"Cancellation requested - cleaning up, please wait...")
        self.logger.warning(f"Script execution cancelled by user - output may be incomplete")

        # Disconnect all active connections
        with self._connections_lock:
            connections_snapshot = list(self.active_connections)
        for conn in connections_snapshot:
            try:
                if conn.is_alive():
                    conn.disconnect()
            except Exception as e:
                self.logger.warning(f"Error disconnecting: {e}")
        with self._connections_lock:
            self.active_connections.clear()

    def process_device(self, device: str, commands: List[str],
                      device_count: int, total_devices: int) -> Tuple[str, bool, int]:
        """
        Process a single device with cancellation support and one retry on timeout.

        Args:
            device: Device IP to process
            commands: List of commands to execute
            device_count: Current device number (for progress tracking)
            total_devices: Total number of devices (for progress tracking)

        Returns:
            Tuple of (device_ip, success_flag, commands_executed)
        """
        if self._is_cancelled:
            return device, False, 0

        # Zero-padded prefix keeps log columns aligned across all devices
        width = len(str(total_devices))
        prefix = f"[{device_count:{width}d}/{total_devices}]"
        _t0 = time.time()
        def _elapsed() -> str:
            return f"{time.time() - _t0:.1f}s"
        connection = None

        try:
            # Get device type (auto-detection runs when structured output is enabled)
            auto_device_type = get_device_type(
                device,
                self.credentials['username'],
                self.credentials['password'],
                self.enable_password,
                self.enable_structured_output,
                cache_path=self._cache_path,
                force_redetect=self.force_redetect
            )
            self.logger.verbose(f"{prefix} Processing {device} as type '{auto_device_type}'")

            # --- Connection with one retry on timeout ---
            for attempt in range(1, 3):
                conn_timeout = CONN_TIMEOUT if attempt == 1 else CONN_TIMEOUT_RETRY
                read_timeout = READ_TIMEOUT if attempt == 1 else READ_TIMEOUT_RETRY
                try:
                    connection = ConnectHandler(
                        device_type=auto_device_type,
                        host=device,
                        username=self.credentials['username'],
                        password=self.credentials['password'],
                        secret=self.enable_password,
                        banner_timeout=60,
                        conn_timeout=conn_timeout,
                        read_timeout_override=read_timeout,
                        global_delay_factor=1,
                        fast_cli=True
                    )
                    break  # Connection succeeded
                except NetmikoTimeoutException:
                    if attempt == 1:
                        self.logger.warning(f"{prefix} Timeout connecting to {device} — retrying with extended timeout...")
                        continue
                    self.logger.error(f"{prefix} Failed to connect to {device} (timeout after retry)")
                    return device, False, 0
                except NetmikoAuthenticationException:
                    self.logger.error(f"{prefix} Authentication failed for {device}")
                    return device, False, 0

            connection.enable()
            with self._connections_lock:
                self.active_connections.append(connection)

            # Post-connection type validation — catches devices SSHDetect misidentifies.
            # show version is the only reliable check; reconnect only when necessary.
            if auto_device_type == 'cisco_ios':
                try:
                    ver_check = connection.send_command('show version', read_timeout=15)
                    if 'NX-OS' in ver_check or ('Nexus' in ver_check and 'Cisco' in ver_check):
                        self.logger.info(
                            f"{prefix} NX-OS detected post-connection for {device} "
                            f"— reconnecting with cisco_nxos"
                        )
                        connection.disconnect()
                        with self._connections_lock:
                            if connection in self.active_connections:
                                self.active_connections.remove(connection)
                        auto_device_type = 'cisco_nxos'
                        connection = ConnectHandler(
                            device_type='cisco_nxos',
                            host=device,
                            username=self.credentials['username'],
                            password=self.credentials['password'],
                            secret=self.enable_password or '',
                            banner_timeout=60,
                            conn_timeout=CONN_TIMEOUT,
                            read_timeout_override=READ_TIMEOUT,
                            global_delay_factor=1,
                            fast_cli=True
                        )
                        connection.enable()
                        with self._connections_lock:
                            self.active_connections.append(connection)
                    elif 'IOS-XE' in ver_check or 'IOS XE' in ver_check:
                        # Catalyst 4500/3850/etc. report "IOS-XE Software" in show version
                        # but SSHDetect scores them as cisco_ios. No reconnect needed —
                        # the session is compatible; only the Genie OS mapping changes.
                        self.logger.info(
                            f"{prefix} IOS-XE detected post-connection for {device} "
                            f"— correcting cisco_ios -> cisco_xe (no reconnect needed)"
                        )
                        auto_device_type = 'cisco_xe'
                except Exception as e:
                    self.logger.debug(
                        f"{prefix} Post-connection type validation failed for {device}: {e}"
                    )

            # Get hostname and resolve output file path
            devicename = self.get_device_hostname(connection)
            output_file = self.generate_output_filename(device, devicename)
            with self._connections_lock:
                self.output_files.append(output_file)

            label = f"{devicename} - {device}" if devicename else device
            self.logger.info(f"{prefix} Connected to {label}")
            self.logger.verbose(f"{prefix} [TIMING] Connected at {_elapsed()}")

            structured:    Dict[str, Any] = {}  # Parsed data keyed by command
            raw_outputs:   Dict[str, str] = {}  # Raw text keyed by command

            with open(output_file, 'w', encoding='utf-8') as file:
                file.write(f"***** DOCUMENTATION STARTED - {format_time()} *****")
                commands_executed = 0
                successful_commands = 0
                failed_commands = 0

                # --- User commands from Commands.txt ---
                for command in commands:
                    if self._is_cancelled:
                        raise Exception("Execution cancelled by user")

                    try:
                        output = connection.send_command(command)
                        if "Invalid input detected at '^' marker" in output:
                            output = f"Command not valid (on this platform): {command}"
                            self.logger.warning(f"{prefix} Command not valid (on this platform): {command}")
                            failed_commands += 1
                        else:
                            successful_commands += 1
                            if self.enable_structured_output:
                                raw_outputs[command] = output

                        file.write(f"\n\n\n!  {format_time()}  {FILLER1}  {command}  {FILLER1}\n")
                        commands_executed += 1
                        file.write(f"\n{output}\n")
                        self.logger.verbose(f"{prefix} Executed '{command}' on {label}")

                    except Exception as e:
                        self.logger.error(f"{prefix} Failed to execute '{command}' on {label}: {e}")
                        failed_commands += 1

                self.logger.verbose(f"{prefix} [TIMING] User commands done at {_elapsed()}")

                # --- Inventory commands (Structured Output only) ---
                if self.enable_structured_output:
                    user_command_set = {c.lower().strip() for c in commands}
                    # 'show vlan' (full) is NX-OS only — on IOS/IOS-XE it outputs every
                    # port membership per VLAN which can be thousands of lines and takes
                    # several minutes on large chassis switches; 'show vlan brief' covers
                    # the same data for IOS/IOS-XE.  Similarly, 'show vlan brief' is
                    # redundant on NX-OS where 'show vlan' is the canonical command.
                    _nxos = (auto_device_type == 'cisco_nxos')
                    _skip = {'show vlan brief'} if _nxos else {'show vlan'}
                    # Build inventory command list from core commands + selected components
                    _needed: List[str] = list(_CORE_COMMANDS)
                    for comp_name, comp_info in EXCEL_COMPONENTS.items():
                        if comp_name in self.selected_components:
                            _needed.extend(comp_info['commands'])
                    # Deduplicate while preserving order
                    _seen_cmds: set = set()
                    _unique_needed: List[str] = []
                    for c in _needed:
                        if c not in _seen_cmds:
                            _seen_cmds.add(c)
                            _unique_needed.append(c)
                    inventory_commands = [
                        c for c in _unique_needed
                        if c.lower().strip() not in user_command_set
                        and c.lower().strip() not in _skip
                    ]
                    if inventory_commands:
                        self.logger.info(f"{prefix} Collecting structured data for {label} ({len(inventory_commands)} inventory commands)")
                        if self.save_txt:
                            file.write(f"\n\n\n{DIVIDER}\n")
                            file.write(f"# Inventory Commands (Structured Output)\n")
                            file.write(f"{DIVIDER}\n")
                        for command in inventory_commands:
                            if self._is_cancelled:
                                raise Exception("Execution cancelled by user")
                            try:
                                output = connection.send_command(command)
                                if ("Invalid input detected" in output or
                                        output.strip().startswith("%")):
                                    continue  # Not supported on this platform — skip silently
                                if self.save_txt:
                                    file.write(f"\n\n\n!  {format_time()}  {FILLER1}  {command}  {FILLER1}\n")
                                    file.write(f"\n{output}\n")
                                raw_outputs[command] = output
                            except Exception:
                                pass  # Inventory command failures are always silent

                self.logger.verbose(f"{prefix} [TIMING] Inventory commands done at {_elapsed()}")
                file.write(f"\n\n\n***** DOCUMENTATION ENDED - {format_time()} *****\n")

            # --- Phase 2: TextFSM parsing (deferred — after all SSH I/O for this device) ---
            # Collecting all raw output first (Phase 1) and parsing here (Phase 2)
            # keeps SSH I/O phases across concurrent threads running with true parallelism.
            if self.enable_structured_output:
                for _cmd, _out in raw_outputs.items():
                    if _cmd not in _TEXTFSM_SKIP:
                        _parsed = try_textfsm_parse(auto_device_type, _cmd, _out)
                        if _parsed is not None:
                            structured[_cmd] = _parsed
            self.logger.verbose(f"{prefix} [TIMING] TextFSM parse done at {_elapsed()}")

            # --- NX-OS fallback: raw-parse show ip interface brief for SVIs ---
            # TextFSM may miss some NX-OS SVI entries; supplement with raw parser.
            if auto_device_type == 'cisco_nxos' and self.enable_structured_output:
                ip_brief_key = 'show ip interface brief'
                existing = structured.get(ip_brief_key)
                raw = raw_outputs.get(ip_brief_key, '')
                if raw:
                    fallback = _parse_nxos_ip_brief_raw(raw)
                    if fallback:
                        # Convert raw dict to TextFSM-style list and merge
                        existing_intfs = set()
                        if isinstance(existing, list):
                            existing_intfs = {
                                _normalize_intf_name(e.get('interface', ''))
                                for e in existing if isinstance(e, dict)
                            }
                        else:
                            existing = []
                        for intf, data in fallback.items():
                            if _normalize_intf_name(intf) not in existing_intfs:
                                existing.append({
                                    'interface': intf,
                                    'ip_address': data.get('ip_address', ''),
                                    'status': data.get('interface_status', ''),
                                    'proto': '',
                                })
                        structured[ip_brief_key] = existing

            # --- Raw-parse commands (no ntc-templates coverage or custom format) ---
            if self.enable_structured_output:
                if auto_device_type in ('cisco_ios', 'cisco_xe'):
                    vlan_raw = raw_outputs.get('show vlan brief', '')
                    if vlan_raw and not vlan_raw.strip().startswith('%'):
                        parsed = _parse_vlan_brief_raw(vlan_raw)
                        if parsed:
                            structured['show vlan brief'] = parsed

                if auto_device_type in ('cisco_ios', 'cisco_xe'):
                    tmpl_raw = raw_outputs.get('show template', '')
                    if tmpl_raw and not tmpl_raw.strip().startswith('%') and 'Template' in tmpl_raw:
                        parsed = _parse_ios_templates_raw(tmpl_raw)
                        if parsed:
                            structured['show template'] = parsed
                elif auto_device_type == 'cisco_nxos':
                    pp_raw = raw_outputs.get('show port-profile', '')
                    if pp_raw and not pp_raw.strip().startswith('%') and 'port-profile' in pp_raw:
                        parsed = _parse_nxos_port_profiles_raw(pp_raw)
                        if parsed:
                            structured['show port-profile'] = parsed
                    # NX-OS ARP: 'show ip arp' raw parse (ntc-templates covers 'show arp' for IOS)
                    arp_raw = raw_outputs.get('show ip arp', '')
                    if arp_raw and not arp_raw.strip().startswith('%'):
                        parsed = _parse_nxos_ip_arp_raw(arp_raw)
                        if parsed:
                            structured['show ip arp'] = parsed

                # show ip route summary — raw parser for cross-platform consistency
                route_raw = raw_outputs.get('show ip route summary', '')
                if route_raw and not route_raw.strip().startswith('%'):
                    try:
                        parsed = _parse_ip_route_summary_raw(route_raw)
                        if parsed:
                            structured['show ip route summary'] = parsed
                    except Exception:
                        pass

                # show spanning-tree summary — raw parser (ntc-templates returns empty)
                stp_raw = raw_outputs.get('show spanning-tree summary', '')
                if stp_raw and not stp_raw.strip().startswith('%'):
                    try:
                        parsed = _parse_stp_summary_raw(stp_raw)
                        if parsed:
                            structured['show spanning-tree summary'] = parsed
                    except Exception:
                        pass

                # MAC address-table count — both platforms (raw parser)
                mac_count_raw = raw_outputs.get('show mac address-table count', '')
                if mac_count_raw and not mac_count_raw.strip().startswith('%'):
                    try:
                        parsed = _parse_mac_count_raw(mac_count_raw, auto_device_type)
                        if parsed.get('totals') or parsed.get('vlans'):
                            structured['show mac address-table count'] = parsed
                    except Exception:
                        pass

                # Per-VLAN MAC count polling: if the device only returned grand totals
                # (e.g. chassis IOS or NX-OS), query each VLAN individually.
                mac_counts = structured.get('show mac address-table count')
                if (mac_counts and isinstance(mac_counts, dict)
                        and not mac_counts.get('vlans')):
                    vlan_key = 'show vlan' if auto_device_type == 'cisco_nxos' else 'show vlan brief'
                    # Primary: structured data from raw parser
                    vlan_ids = [
                        v for v in structured.get(vlan_key, {}).get('vlans', {})
                        if str(v).isdigit() and 1 <= int(v) <= 4094
                    ]
                    # Fallback: extract VLAN IDs directly from raw output
                    # (lines starting with a VLAN number)
                    if not vlan_ids:
                        _vlan_raw = raw_outputs.get(vlan_key, '')
                        _seen: set = set()
                        for _ln in _vlan_raw.splitlines():
                            _m = re.match(r'^\s*(\d+)\s+\S', _ln)
                            if _m:
                                _vid = int(_m.group(1))
                                if 1 <= _vid <= 4094 and _vid not in _seen:
                                    _seen.add(_vid)
                                    vlan_ids.append(_vid)
                        vlan_ids.sort()
                    if vlan_ids:
                        self.logger.info(
                            f"{prefix} No per-VLAN MAC counts — querying {len(vlan_ids)} VLANs individually for {label}"
                        )
                        for vlan_id in sorted(vlan_ids, key=int):
                            if self._is_cancelled:
                                break
                            try:
                                vraw = connection.send_command(
                                    f'show mac address-table count vlan {vlan_id}',
                                    read_timeout=10,  # short — simple count command
                                )
                                if (vraw and not vraw.strip().startswith('%')
                                        and 'Invalid input' not in vraw):
                                    vdata = _parse_mac_count_vlan_raw(vraw)
                                    if vdata.get('total', 0) > 0 or vdata.get('dynamic', 0) > 0:
                                        mac_counts['vlans'][str(vlan_id)] = vdata
                            except Exception:
                                pass  # Per-VLAN query failures are always silent

                # show inventory stack members — raw parse for consistent cross-platform format.
                # TextFSM stores show inventory as a list; wrap in a dict to carry
                # both the TextFSM entries and stack member data.
                inv_raw = raw_outputs.get('show inventory', '')
                if inv_raw:
                    inv_members = _parse_inventory_members_raw(inv_raw)
                    if inv_members:
                        existing_inv = structured.get('show inventory')
                        if isinstance(existing_inv, list):
                            structured['show inventory'] = {
                                '_entries': existing_inv,
                                '_stack_members': inv_members,
                            }
                        elif isinstance(existing_inv, dict):
                            existing_inv['_stack_members'] = inv_members
                        else:
                            structured['show inventory'] = {'_stack_members': inv_members}

            if self.enable_structured_output:
                self.logger.verbose(f"{prefix} [TIMING] Post-processing done at {_elapsed()}")

            # --- Write per-device JSON if structured data was captured ---
            if self.enable_structured_output and not structured:
                self.logger.warning(f"{prefix} Structured Output enabled but no commands parsed for {label} (device_type='{auto_device_type}')")
            if structured:
                json_data: Dict[str, Any] = {
                    "meta": {
                        "host": device,
                        "hostname": devicename,
                        "device_type": auto_device_type,
                        "documented_at": format_time(),
                        "nddu_version": APP_VERSION,
                    },
                    "structured": structured,
                }
                json_folder = self.output_folder / "json"
                json_folder.mkdir(exist_ok=True)
                json_file = json_folder / f"{Path(output_file).stem}.json"
                try:
                    with open(json_file, 'w', encoding='utf-8') as jf:
                        json.dump(json_data, jf, indent=2, default=str)
                    self.logger.info(f"{prefix} Structured data saved  ->  json/{json_file.name}")
                except Exception as e:
                    self.logger.warning(f"{prefix} Failed to write JSON for {label}: {e}")

            connection.disconnect()
            with self._connections_lock:
                self.active_connections.remove(connection)

            # Resolve display path for the completion log line
            try:
                display_path = f"./{output_file.relative_to(Path(__file__).parent).as_posix()}"
            except ValueError:
                display_path = str(output_file)

            self.logger.verbose(f"{prefix} [TIMING] Total elapsed {_elapsed()}")
            self.logger.info(
                f"{prefix} Done  {label}  ({successful_commands}/{len(commands)} commands)  ->  {display_path}"
            )

            return device, True, commands_executed

        except Exception as e:
            if str(e) == "Execution cancelled by user":
                self.logger.info(f"{prefix} Execution stopped while processing {device}")
            else:
                self.logger.error(f"{prefix} Unexpected error processing {device}: {e}")

            # Ensure connection is cleaned up
            if connection is not None and connection.is_alive():
                try:
                    connection.disconnect()
                    with self._connections_lock:
                        if connection in self.active_connections:
                            self.active_connections.remove(connection)
                except Exception as cleanup_err:
                    self.logger.warning(f"{prefix} Error during disconnect: {cleanup_err}")

            return device, False, 0

    def get_device_hostname(self, connection: Any) -> Optional[str]:
        """
        Retrieve the device's hostname from the running configuration.
        
        Args:
            connection: Active Netmiko connection
            
        Returns:
            Device hostname if found, None otherwise
        """
        try:
            # Send command to get the full hostname from running configuration
            output = connection.send_command('show running-config | include hostname')
            if output:
                hostname_line = output.strip()
                if hostname_line.startswith('hostname'):
                    hostname = hostname_line.split()[1]
                    return hostname
            hostname = connection.base_prompt  # Fallback: Directly use Netmiko's detected prompt
            return hostname.strip()
        except Exception as e:
            self.logger.warning(f"Failed to retrieve hostname: {e}")
            return None

    def generate_output_filename(self, device: str, devicename: Optional[str]) -> Path:
        """
        Generate the output filename for the device.
        
        Args:
            device: Device IP address
            devicename: Optional device hostname
            
        Returns:
            Path object for the output file
        """
        # Hostname comes from the network device — sanitize before using as a
        # filename so a hostile or misconfigured device cannot write outside
        # the run's output folder.
        if devicename:
            safe_name = sanitize_folder_name(devicename)
            filename = self.output_folder / f"{safe_name} - {device}.txt"
        else:
            filename = self.output_folder / f"{device}.txt"

        # Try to return the relative path, but fall back to the absolute path if it fails
        try:
            relative_path = filename.relative_to(Path(__file__).parent)
            return Path(f"./{relative_path.as_posix()}")
        except ValueError:
            # If the file is not in a subpath of the script's directory, return the absolute path
            return filename

    def create_combined_output_file(self) -> None:
        """Combine all individual output files into one master file in order of processing."""
        if not self.output_files:
            self.logger.warning(f"No output files to combine")
            return

        # Try to use the relative path, but fall back to the absolute path if it fails
        try:
            combined_file = Path(self.output_folder / COMBINED_OUTPUT_FILENAME).relative_to(Path(__file__).parent)
        except ValueError:
            # If the file is not in a subpath of the script's directory, return the absolute path
            combined_file = self.output_folder / COMBINED_OUTPUT_FILENAME

        try:
            with open(combined_file, 'w', encoding='utf-8') as outfile:
                # Write header
                outfile.write(f"***** COMBINED DEVICE OUTPUT - {format_time()} *****\n")
                outfile.write(f"Combined output from {len(self.output_files)} devices\n")
                outfile.write(f"{DIVIDER}\n")

                # Append each device's output file
                for output_file in self.output_files:
                    try:
                        with open(output_file, 'r', encoding='utf-8') as infile:
                            # Write device header
                            outfile.write(f"\n\n{FILLER2} {output_file.stem} {FILLER2}\n\n")
                            # Copy contents
                            outfile.write(infile.read())
                            outfile.write("\n\n")
                    except Exception as e:
                        self.logger.error(f"Failed to combine file {output_file}: {e}")

                # Write footer
                outfile.write(f"{DIVIDER}\n")
                outfile.write(f"***** END OF COMBINED OUTPUT - {format_time()} *****\n")

            self.logger.info(f'Created combined output file: "{combined_file}"')
        except Exception as e:
            self.logger.error(f"Failed to create combined output file: {e}")

class LineNumberArea(QWidget):
    """Widget that displays line numbers for a CodeEditor."""
    
    def __init__(self, editor: 'CodeEditor') -> None:
        """Initialize with a reference to the parent editor."""
        super().__init__(editor)
        self.editor = editor

    def sizeHint(self) -> QSize:
        """Return the recommended size for the line number area."""
        return QSize(self.editor.line_number_area_width(), 0)

    def paintEvent(self, event: Any) -> None:
        """Handle paint events by delegating to the editor."""
        self.editor.line_number_area_paint_event(event)

class CodeEditor(QPlainTextEdit):
    """Enhanced text editor with line numbers and syntax highlighting."""
    
    def __init__(self, parent: Optional[QWidget] = None) -> None:
        """Initialize the editor with line number support."""
        super().__init__(parent)
        self.line_number_area = LineNumberArea(self)

        # Define padding between line numbers and text
        self.TEXT_WINDOW_PADDING = 10

        # Connect signals
        self.blockCountChanged.connect(self.update_line_number_area_width)
        self.updateRequest.connect(self.update_line_number_area)
        self.cursorPositionChanged.connect(self.highlight_current_line)

        # Set up the line number area
        self.update_line_number_area_width()
        self.highlight_current_line()

    def line_number_area_width(self) -> int:
        """Calculate the width required for the line number area."""
        digits = 1
        max_lines = max(1, self.blockCount())
        while max_lines >= 10:
            max_lines /= 10
            digits += 1
        # Add some extra space for padding
        space = 10 + self.fontMetrics().horizontalAdvance('9') * digits
        return space

    def update_line_number_area_width(self) -> None:
        """Update the width of the line number area and add padding to the text window."""
        self.setViewportMargins(
            self.line_number_area_width() + self.TEXT_WINDOW_PADDING,  # Left margin (line numbers + padding)
            0,  # Top margin
            0,  # Right margin
            0   # Bottom margin
        )
        self.line_number_area.update()  # Force a redraw of the line number area

    def update_line_number_area(self, rect: QRect, dy: int) -> None:
        """Update the line number area when the text changes."""
        if dy:
            self.line_number_area.scroll(0, dy)
        else:
            self.line_number_area.update(0, rect.y(), self.line_number_area.width(), rect.height())
        if rect.contains(self.viewport().rect()):
            self.update_line_number_area_width()

    def resizeEvent(self, event: Any) -> None:
        """Handle resize events to adjust the line number area."""
        super().resizeEvent(event)
        cr = self.contentsRect()
        self.line_number_area.setGeometry(
            QRect(
                cr.left(),  # X position
                cr.top(),  # Y position
                self.line_number_area_width(),  # Width
                cr.height()  # Height
            )
        )

    def line_number_area_paint_event(self, event: Any) -> None:
        """Paint the line numbers in the line number area."""
        painter = QPainter(self.line_number_area)
        painter.fillRect(event.rect(), QColor(240, 240, 240))  # Light gray background

        block = self.firstVisibleBlock()
        block_number = block.blockNumber()
        top = self.blockBoundingGeometry(block).translated(self.contentOffset()).top()
        bottom = top + self.blockBoundingRect(block).height()

        while block.isValid() and top <= event.rect().bottom():
            if block.isVisible() and bottom >= event.rect().top():
                number = str(block_number + 1)
                painter.setPen(Qt.GlobalColor.black)
                # Calculate the right-aligned position for the line number
                text_width = self.fontMetrics().horizontalAdvance(number)
                x = self.line_number_area.width() - text_width - 5  # Right-align with 5px padding
                painter.drawText(
                    x,  # Right-aligned X position
                    int(top),
                    text_width,  # Width of the text
                    self.fontMetrics().height(),
                    Qt.AlignmentFlag.AlignRight,
                    number
                )

            block = block.next()
            top = bottom
            bottom = top + self.blockBoundingRect(block).height()
            block_number += 1

    def highlight_current_line(self) -> None:
        """Highlight the current line in the editor."""
        extra_selections = []
        if not self.isReadOnly():
            selection = QTextEdit.ExtraSelection()
            line_color = QColor(255, 255, 0, 50)  # Light yellow highlight
            selection.format.setBackground(line_color)
            selection.format.setProperty(QTextFormat.Property.FullWidthSelection, True)
            selection.cursor = self.textCursor()
            selection.cursor.clearSelection()
            extra_selections.append(selection)
        self.setExtraSelections(extra_selections)

class FileEditorDialog(QDialog):
    """Dialog for editing input files with syntax highlighting."""
    
    file_path_updated = Signal(Path, str)  # Signal to emit when the file path changes (Path, file_type)

    def __init__(self, file_path: Path, file_type: str, parent: Optional[QWidget] = None) -> None:
        """
        Initialize the file editor dialog.

        Args:
            file_path: Path to the file to edit
            file_type: Type of file ('device', 'command', or 'port_map')
            parent: Optional parent widget
        """
        super().__init__(parent)
        self.file_path = file_path
        self.file_type = file_type  # Track whether this is a "device", "command", or "port_map" file
        self.setWindowTitle(f"Editing: {file_path.name}")  # Display only the file name
        self.setModal(True)  # Make the dialog modal
        self.resize(800, 600)
        self.unsaved_changes = False  # Track unsaved changes

        # Create the main layout
        layout = QVBoxLayout(self)

        # Create a CodeEditor for editing the file (with line numbers)
        self.text_edit = CodeEditor(self)
        self.text_edit.textChanged.connect(self.mark_unsaved_changes)  # Connect to textChanged signal
        layout.addWidget(self.text_edit)

        # Create a status label to show messages
        self.status_label = QLabel("Ready", self)
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.status_label.setStyleSheet("font-weight: bold; color: black;")
        layout.addWidget(self.status_label)

        # Create a horizontal layout for the buttons
        button_layout = QHBoxLayout()

        # Add Save button
        self.save_button = QPushButton("Save", self)
        self.save_button.clicked.connect(self.save_file)
        self.save_button.setEnabled(False)  # Disabled by default
        button_layout.addWidget(self.save_button)

        # Add Save As button
        self.save_as_button = QPushButton("Save As...", self)
        self.save_as_button.clicked.connect(self.save_file_as)
        self.save_as_button.setEnabled(True)  # Always enabled
        button_layout.addWidget(self.save_as_button)

        # Add Set as Default button (not applicable for port_map files)
        self.set_default_button = QPushButton("Set as Default", self)
        self.set_default_button.clicked.connect(self.set_as_default)
        self.set_default_button.setToolTip(
            "Make this the default file loaded on startup.\n"
            "If the file cannot be found in a future session, the original default is used."
        )
        button_layout.addWidget(self.set_default_button)
        if file_type == "port_map":
            self.set_default_button.hide()

        # Add Close button
        self.close_button = QPushButton("Close", self)
        self.close_button.clicked.connect(self.close_editor)
        button_layout.addWidget(self.close_button)

        # Add the button layout to the main layout
        layout.addLayout(button_layout)

        # Set the layout
        self.setLayout(layout)

        # Load the file content after setting up the UI
        self.load_file(file_path)

    def load_file(self, file_path: Path) -> None:
        """Load the file contents into the editor."""
        try:
            # Block the textChanged signal while loading the file
            self.text_edit.blockSignals(True)  # Suppress textChanged signal
            with open(file_path, "r") as file:
                self.file_content = file.read()
            self.text_edit.setPlainText(self.file_content)  # Set the text in the CodeEditor
            self.text_edit.blockSignals(False)  # Re-enable textChanged signal

            # Move the cursor to the start of the document
            cursor = self.text_edit.textCursor()
            cursor.movePosition(QTextCursor.MoveOperation.Start)
            self.text_edit.setTextCursor(cursor)

            # Highlight the first line
            self.text_edit.highlight_current_line()

            # Ensure the first line is visible
            self.text_edit.ensureCursorVisible()

            # Ensure the line number area and padding are updated
            self.text_edit.update_line_number_area_width()
            self.text_edit.line_number_area.update()  # Force a redraw of the line number area

            self.status_label.setText("File loaded successfully.")
            self.status_label.setStyleSheet("font-weight: bold; color: green;")
        except Exception as e:
            self.status_label.setText(f"Error reading file: {e}")
            self.status_label.setStyleSheet("font-weight: bold; color: red;")
            self.close()

    def save_file(self) -> None:
        """Save changes to the original file."""
        try:
            with open(self.file_path, "w") as file:
                file.write(self.text_edit.toPlainText())
            self.unsaved_changes = False  # Reset unsaved changes flag
            self.setWindowTitle(f"Editing: {self.file_path.name}")  # Update title
            self.save_button.setEnabled(False)  # Disable Save button after saving
            self.status_label.setText("File saved successfully.")
            self.status_label.setStyleSheet("font-weight: bold; color: green;")
        except Exception as e:
            self.status_label.setText(f"Failed to save file: {e}")
            self.status_label.setStyleSheet("font-weight: bold; color: red;")

    def save_file_as(self) -> None:
        """Save changes to a new file."""
        if self.file_type == "port_map":
            file_filter = "CSV Files (*.csv);;All Files (*)"
        else:
            file_filter = "Text Files (*.txt);;All Files (*)"
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save File As", str(self.file_path.parent), file_filter
        )
        if file_path:
            try:
                with open(file_path, "w") as file:
                    file.write(self.text_edit.toPlainText())
                self.unsaved_changes = False  # Reset unsaved changes flag
                self.file_path = Path(file_path)  # Update the file path
                self.setWindowTitle(f"Editing: {self.file_path.name}")  # Update title
                self.save_button.setEnabled(False)  # Disable Save button after saving
                self.status_label.setText("File saved successfully.")
                self.status_label.setStyleSheet("font-weight: bold; color: green;")
                # Emit the new file path and file type
                self.file_path_updated.emit(self.file_path, self.file_type)
            except Exception as e:
                self.status_label.setText(f"Failed to save file: {e}")
                self.status_label.setStyleSheet("font-weight: bold; color: red;")

    def close_editor(self) -> None:
        """Prompt to save changes before closing if there are unsaved changes."""
        if self.unsaved_changes:
            reply = QMessageBox.question(
                self, "Save Changes", "Do you want to save changes before closing?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No | QMessageBox.StandardButton.Cancel
            )
            if reply == QMessageBox.StandardButton.Yes:
                self.save_file()
                self.accept()  # Close the dialog and return the new file path
            elif reply == QMessageBox.StandardButton.No:
                self.status_label.setText("Changes discarded.")
                self.status_label.setStyleSheet("font-weight: bold; color: orange;")
                self.reject()  # Close the dialog without saving
            # If Cancel, do nothing
        else:
            self.reject()  # Close the dialog without saving

    def set_as_default(self) -> None:
        """Save the current file path as the default for this file type."""
        key = f"defaults/{self.file_type}_file"
        QSettings(SettingsDialog.SETTINGS_FILE, QSettings.Format.IniFormat).setValue(key, str(self.file_path))
        self.status_label.setText(f"Set as default {self.file_type} file: {self.file_path.name}")
        self.status_label.setStyleSheet("font-weight: bold; color: green;")

    def mark_unsaved_changes(self) -> None:
        """Mark that the file has unsaved changes."""
        if not self.unsaved_changes:  # Only update if changes are not already marked
            self.unsaved_changes = True
            self.setWindowTitle(f"Editing: {self.file_path.name} *")  # Add asterisk to indicate unsaved changes
            self.save_button.setEnabled(True)  # Enable Save button
            self.status_label.setText("Unsaved changes.")
            self.status_label.setStyleSheet("font-weight: bold; color: orange;")

    def closeEvent(self, event: Any) -> None:
        """Override the close event to prompt for saving changes."""
        self.close_editor()
        event.ignore()  # Prevent the dialog from closing immediately

class HelpDialog(QDialog):
    """Custom dialog for displaying help information."""
    
    def __init__(self, title: str, version: str, version_date: str, 
                 parent: Optional[QWidget] = None) -> None:
        """
        Initialize the help dialog.
        
        Args:
            title: Application title
            version: Version string
            version_date: Version date string
            parent: Optional parent widget
        """
        super().__init__(parent)
        self.setWindowTitle("Help")
        
        # Inherit dark mode setting from parent if available
        self.dark_mode = parent.dark_mode if parent and hasattr(parent, 'dark_mode') else DARK_MODE_STATE
        
        # Set Help window size (width, height)
        self.setFixedSize(500, 640)

        # Get repo URL and callback from parent if available
        self.repo_url = REPO_URL
        # self.check_update_callback = None
        # if parent and hasattr(parent, 'check_for_updates'):
        #     self.check_update_callback = parent.check_for_updates
        
        # Create main layout with margins
        layout = QVBoxLayout(self)
        layout.setContentsMargins(15, 15, 15, 15)  # Window margins

        # Add the title and version
        title_label = QLabel(f"<h1>{APP_NAME}</h1><h3>{APP_VERSION} {VERSION_DATE}</h3>")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        # Create a widget for repo link and update check button
        repo_widget = QWidget()
        repo_layout = QHBoxLayout(repo_widget)
        repo_layout.setContentsMargins(0, 10, 0, 10)  # Add some top margin

        # Repo link
        repo_label = QLabel(f'<a href="{self.repo_url}" style="color: #4CAF50;">Visit Website</a>')
        repo_label.setOpenExternalLinks(True)
        repo_layout.addWidget(repo_label)
        repo_layout.addStretch()

        # Add title and repo widget
        layout.addWidget(title_label)
        layout.addWidget(repo_widget, alignment=Qt.AlignmentFlag.AlignCenter)

        # Create scroll area (no frame)
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setFrameShape(QFrame.NoFrame)  # Remove default border

        # Create container widget
        container = QWidget()
        container_layout = QVBoxLayout(container)
        container_layout.setContentsMargins(0, 0, 0, 0)

        # Create framed content area
        content_frame = QFrame()
        content_frame.setObjectName("contentFrame")
        frame_layout = QVBoxLayout(content_frame)
        frame_layout.setContentsMargins(15, 15, 15, 15)  # Padding inside frame

        # Add the help text
        help_text = f"""
        This tool helps you document network devices by processing:
        1. A list of device IPs (Device List).
        2. A list of commands to run on each device (Command List).
        3. Credentials for accessing the devices.

        Usage:
        1. Select the Device List and Command List files.
        2. Choose a credential option (Manual or Keyring).
        3. Enter a Client / Job name to organize output.
        4. Click 'Go' to start the process.

        Features:
        - Excel Report: Enable structured parsing to generate detailed
          Excel reports with selectable components (Interfaces, Neighbors,
          VLANs, Routing, STP, MAC Addresses, etc.).
        - Client Manager: Manage client folders, view run history,
          archive/delete clients, and generate run reports.
        - Device Diff: Compare two prior runs for the same client and
          generate a color-coded Excel diff report showing what changed.
        - Device Type Cache: Auto-detected device types are cached per
          client for faster subsequent runs.
        - Keyring Credentials: Store credentials securely in the OS
          keyring instead of entering them each time.
        """

        help_label = QLabel(help_text)
        help_label.setWordWrap(True)
        help_label.setAlignment(Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignLeft)
        
        frame_layout.addWidget(help_label)
        container_layout.addWidget(content_frame)
        scroll_area.setWidget(container)
        layout.addWidget(scroll_area)

        # Create a widget for the Close button (centered)
        close_widget = QWidget()
        close_layout = QHBoxLayout(close_widget)
        close_layout.setContentsMargins(0, 10, 0, 10)
        
        # Add stretch to center the button
        close_layout.addStretch()
        
        # Add Close button
        close_button = QPushButton("Close")
        close_button.clicked.connect(self.close)
        close_button.setMinimumWidth(100)
        close_layout.addWidget(close_button)
        
        # Add stretch to center the button
        close_layout.addStretch()
        
        # Add close widget to main layout
        layout.addWidget(close_widget)

        # Apply theme
        self.apply_theme()

    def apply_theme(self) -> None:
        """Apply dark or light theme based on current mode."""
        if self.dark_mode:
            # Dark theme palette
            dark_palette = QPalette()
            dark_palette.setColor(QPalette.ColorRole.Window, QColor(53, 53, 53))
            dark_palette.setColor(QPalette.ColorRole.WindowText, Qt.GlobalColor.white)
            dark_palette.setColor(QPalette.ColorRole.Base, QColor(25, 25, 25))
            dark_palette.setColor(QPalette.ColorRole.Text, Qt.GlobalColor.white)
            dark_palette.setColor(QPalette.ColorRole.Button, QColor(53, 53, 53))
            dark_palette.setColor(QPalette.ColorRole.ButtonText, Qt.GlobalColor.white)
            self.setPalette(dark_palette)
            
            # Dark theme stylesheet
            self.setStyleSheet("""
                QFrame#contentFrame {
                    background-color: #252525;
                    border: 2px solid #555;
                    border-radius: 5px;
                }
                QLabel {
                    color: white;
                    background-color: transparent;
                }
                QScrollArea {
                    background-color: #353535;
                    border: none;
                }
                QPushButton {
                    background-color: #353535;
                    color: white;
                    border: 1px solid #555;
                    padding: 8px 20px;
                    border-radius: 4px;
                    min-width: 100px;
                }
                QPushButton:hover {
                    background-color: #454545;
                }
                QPushButton:disabled {
                    background-color: #555;
                    color: #999;
                }
            """)
        else:
            # Reset to default light theme
            self.setPalette(QApplication.style().standardPalette())
            self.setStyleSheet("""
                QPushButton:disabled {
                    background-color: #e0e0e0;
                    color: #999;
                }
            """)

class ClientManagerDialog(QDialog):
    """Dialog for managing Client/Job output folders."""

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("Client / Job Manager")
        self.setMinimumSize(720, 620)
        self._init_ui()
        self._refresh_list()

    # ------------------------------------------------------------------
    # UI construction
    # ------------------------------------------------------------------
    def _init_ui(self) -> None:
        root = QHBoxLayout(self)
        root.setSpacing(8)

        # ── Left: client list + client actions ────────────────────────
        left = QVBoxLayout()
        left.setSpacing(4)
        left.addWidget(QLabel("Clients / Jobs:"))
        self.client_list = QListWidget()
        self.client_list.setMinimumWidth(180)
        self.client_list.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.client_list.currentRowChanged.connect(self._on_selection_changed)
        left.addWidget(self.client_list)

        # Client-level action buttons (always visible, selection-gated where needed)
        client_btn_col = QVBoxLayout()
        client_btn_col.setSpacing(4)
        self.new_button            = QPushButton("New")
        self.rename_button         = QPushButton("Rename")
        self.report_button         = QPushButton("Report…")
        self.archive_delete_button = QPushButton("Archive / Delete…")
        self.new_button.setToolTip("Create a new client folder")
        self.rename_button.setToolTip("Rename the selected client folder")
        self.report_button.setToolTip("Export a run history report (XLSX or CSV)")
        self.archive_delete_button.setToolTip("Archive and/or delete the selected client folder")
        for btn in (self.new_button, self.rename_button, self.report_button,
                    self.archive_delete_button):
            client_btn_col.addWidget(btn)
        client_btn_col.addStretch()
        left.addLayout(client_btn_col)

        self.new_button.clicked.connect(self._on_new)
        self.rename_button.clicked.connect(self._on_rename)
        self.report_button.clicked.connect(self._on_report)
        self.archive_delete_button.clicked.connect(self._on_archive_delete)

        root.addLayout(left, 1)

        # ── Right: runs tree + cache section ─────────────────────────
        right = QVBoxLayout()
        right.setSpacing(6)

        # Runs tree — timestamps as top-level, output files as children
        runs_group = QGroupBox("Runs")
        runs_layout = QVBoxLayout()
        runs_layout.setSpacing(4)
        self.runs_tree = QTreeWidget()
        self.runs_tree.setHeaderLabels(["Timestamp / File", "Device File"])
        self.runs_tree.header().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.runs_tree.header().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.runs_tree.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.runs_tree.setMinimumHeight(160)
        self.runs_tree.itemDoubleClicked.connect(self._on_tree_double_click)
        self.runs_tree.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.runs_tree.customContextMenuRequested.connect(self._on_tree_context_menu)
        self.runs_tree.currentItemChanged.connect(self._on_tree_selection_changed)
        runs_layout.addWidget(self.runs_tree)

        # Run-level action buttons
        run_btn_row = QHBoxLayout()
        run_btn_row.setSpacing(5)
        self.open_run_folder_button = QPushButton("Open Folder")
        self.open_run_folder_button.setToolTip("Open the selected run folder in the OS file explorer")
        self.delete_run_button = QPushButton("Delete Run")
        self.delete_run_button.setToolTip("Delete the selected run folder and all its files")
        self.compare_runs_button = QPushButton("Compare Runs…")
        self.compare_runs_button.setToolTip("Generate a diff report comparing two runs")
        self.open_run_folder_button.setEnabled(False)
        self.delete_run_button.setEnabled(False)
        self.compare_runs_button.setEnabled(False)
        self.open_run_folder_button.clicked.connect(self._on_open_run_folder)
        self.delete_run_button.clicked.connect(self._on_delete_run)
        self.compare_runs_button.clicked.connect(self._on_compare_runs)
        run_btn_row.addWidget(self.open_run_folder_button)
        run_btn_row.addWidget(self.delete_run_button)
        run_btn_row.addWidget(self.compare_runs_button)
        run_btn_row.addStretch()
        runs_layout.addLayout(run_btn_row)

        runs_group.setLayout(runs_layout)
        right.addWidget(runs_group, 2)

        # Device type cache section
        cache_group = QGroupBox("Device Type Cache")
        cache_layout = QVBoxLayout()
        cache_layout.setSpacing(4)
        self.cache_table = QTableWidget(0, 2)
        self.cache_table.setHorizontalHeaderLabels(["IP Address", "Device Type"])
        self.cache_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.cache_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.cache_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.cache_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.cache_table.setMaximumHeight(130)
        cache_layout.addWidget(self.cache_table)

        cache_btn_row = QHBoxLayout()
        cache_btn_row.setSpacing(5)
        self.delete_cache_entry_button = QPushButton("Remove Entry")
        self.delete_cache_entry_button.setToolTip("Remove the selected IP entry from the device type cache")
        self.delete_all_cache_button   = QPushButton("Clear All Cache")
        self.delete_all_cache_button.setToolTip("Delete device_cache.json — all devices will be re-detected on the next run")
        cache_btn_row.addWidget(self.delete_cache_entry_button)
        cache_btn_row.addWidget(self.delete_all_cache_button)
        cache_btn_row.addStretch()
        cache_layout.addLayout(cache_btn_row)
        cache_group.setLayout(cache_layout)
        right.addWidget(cache_group, 1)

        self.delete_cache_entry_button.clicked.connect(self._on_delete_cache_entry)
        self.delete_all_cache_button.clicked.connect(self._on_delete_all_cache)

        # Status + close
        self.status_label = QLabel("")
        self.status_label.setWordWrap(True)
        right.addWidget(self.status_label)

        close_row = QHBoxLayout()
        close_row.addStretch()
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.accept)
        close_row.addWidget(close_btn)
        right.addLayout(close_row)

        root.addLayout(right, 2)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    def _selected_client(self) -> Optional[str]:
        item = self.client_list.currentItem()
        return item.text() if item else None

    def _client_path(self, name: str) -> Path:
        return DEFAULT_OUTPUT_FOLDER / name

    def _set_status(self, msg: str, error: bool = False) -> None:
        self.status_label.setText(msg)
        self.status_label.setStyleSheet("color: red;" if error else "color: green;")

    def _refresh_list(self, select: Optional[str] = None) -> None:
        self.client_list.blockSignals(True)
        self.client_list.clear()
        for name in get_existing_clients():
            self.client_list.addItem(name)
        self.client_list.blockSignals(False)
        if select:
            items = self.client_list.findItems(select, Qt.MatchFlag.MatchExactly)
            if items:
                self.client_list.setCurrentItem(items[0])
            else:
                self.client_list.setCurrentRow(0)
        elif self.client_list.count():
            self.client_list.setCurrentRow(0)
        else:
            self._on_selection_changed(-1)

    def _read_run_info(self, run_path: Path) -> Dict[str, str]:
        """Read run_info.json if present; fall back to scanning output files."""
        info_path = run_path / "run_info.json"
        if info_path.exists():
            try:
                with open(info_path, encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, dict):
                    return data
            except (OSError, ValueError):
                pass
        return {}

    def _refresh_runs(self, client_path: Path) -> None:
        self.runs_tree.clear()
        if not client_path.exists():
            return
        runs = sorted(
            (d for d in client_path.iterdir() if d.is_dir() and _TIMESTAMP_RE.match(d.name)),
            key=lambda d: d.name
        )
        for run in runs:
            info = self._read_run_info(run)
            device_file = info.get("device_file", "")
            run_item = QTreeWidgetItem([run.name, device_file])
            # Child rows: one per output file in the run folder
            for f in sorted(run.iterdir()):
                if f.is_file() and f.suffix in (".txt", ".xlsx", ".log", ".json"):
                    child = QTreeWidgetItem([f"  {f.name}", ""])
                    run_item.addChild(child)
            self.runs_tree.addTopLevelItem(run_item)

    def _refresh_cache(self, client_path: Path) -> None:
        self.cache_table.setRowCount(0)
        cache_path = client_path / DEVICE_CACHE_FILENAME
        has_cache = cache_path.exists()
        cache = _load_device_cache(cache_path) if has_cache else {}
        self.delete_all_cache_button.setEnabled(has_cache)
        self.delete_cache_entry_button.setEnabled(False)
        for ip, dtype in sorted(cache.items()):
            row = self.cache_table.rowCount()
            self.cache_table.insertRow(row)
            self.cache_table.setItem(row, 0, QTableWidgetItem(ip))
            self.cache_table.setItem(row, 1, QTableWidgetItem(dtype))
        self.cache_table.itemSelectionChanged.connect(self._on_cache_selection_changed)

    def _on_cache_selection_changed(self) -> None:
        self.delete_cache_entry_button.setEnabled(bool(self.cache_table.selectedItems()))

    def _client_buttons_enabled(self, enabled: bool) -> None:
        for btn in (self.rename_button, self.archive_delete_button, self.report_button):
            btn.setEnabled(enabled)

    def _selected_run_path(self) -> Optional[Path]:
        """Return the Path for the currently selected top-level run item, or None."""
        item = self.runs_tree.currentItem()
        if item is None:
            return None
        # Top-level items are runs; child items are files inside a run
        if item.parent() is not None:
            item = item.parent()
        name = self._selected_client()
        if not name:
            return None
        return self._client_path(name) / item.text(0).strip()

    def _open_path(self, path: Path) -> None:
        """Open a file or folder with the OS default application."""
        try:
            if platform.system() == "Windows":
                os.startfile(path)
            elif platform.system() == "Darwin":
                subprocess.run(["open", str(path)], check=False)
            else:
                subprocess.run(["xdg-open", str(path)], check=False)
        except Exception as e:
            self._set_status(f"Could not open: {e}", error=True)

    def _do_archive(self, name: str) -> Optional[str]:
        """Zip client folder; returns save path on success, None if cancelled/failed."""
        timestamp = datetime.now().strftime("%m-%d-%Y_%I-%M_%p")
        default_zip = DEFAULT_OUTPUT_FOLDER / f"{name}_{timestamp}.zip"
        save_path, _ = QFileDialog.getSaveFileName(
            self, "Save Archive", str(default_zip), "ZIP Files (*.zip)"
        )
        if not save_path:
            return None
        client_path = self._client_path(name)
        try:
            with zipfile.ZipFile(save_path, "w", zipfile.ZIP_DEFLATED) as zf:
                for file in client_path.rglob("*"):
                    if file.is_file():
                        zf.write(file, file.relative_to(client_path.parent))
            return save_path
        except Exception as e:
            self._set_status(f"Archive failed: {e}", error=True)
            return None

    # ------------------------------------------------------------------
    # Slots
    # ------------------------------------------------------------------
    def _on_selection_changed(self, row: int) -> None:
        if row < 0:
            self.runs_tree.clear()
            self.cache_table.setRowCount(0)
            self._client_buttons_enabled(False)
            self.delete_all_cache_button.setEnabled(False)
            self.delete_cache_entry_button.setEnabled(False)
            self.open_run_folder_button.setEnabled(False)
            self.delete_run_button.setEnabled(False)
            self.compare_runs_button.setEnabled(False)
            self.status_label.setText("")
            return
        name = self.client_list.item(row).text()
        path = self._client_path(name)
        self._refresh_runs(path)
        self._refresh_cache(path)
        self._client_buttons_enabled(True)
        self._update_compare_button_state(path)

    def _on_tree_selection_changed(self, current, _previous) -> None:
        has_item = current is not None
        self.open_run_folder_button.setEnabled(has_item)
        # Only enable Delete Run when a top-level (run) row is selected
        is_run = has_item and current.parent() is None
        self.delete_run_button.setEnabled(is_run)
        # Switch button text between Open Folder / Open File
        if has_item and current.parent() is not None:
            self.open_run_folder_button.setText("Open File")
            self.open_run_folder_button.setToolTip("Open the selected file with the OS default application")
        else:
            self.open_run_folder_button.setText("Open Folder")
            self.open_run_folder_button.setToolTip("Open the selected run folder in the OS file explorer")

    def _on_new(self) -> None:
        name, ok = QInputDialog.getText(self, "New Client / Job", "Client / Job name:")
        if not ok or not name.strip():
            return
        sanitized = sanitize_folder_name(name.strip())
        path = self._client_path(sanitized)
        if path.exists():
            self._set_status(f'"{sanitized}" already exists.', error=True)
            return
        try:
            path.mkdir(parents=True)
            self._refresh_list(select=sanitized)
            self._set_status(f'Created "{sanitized}".')
        except OSError as e:
            self._set_status(f"Failed to create folder: {e}", error=True)

    def _on_rename(self) -> None:
        name = self._selected_client()
        if not name:
            return
        new_name, ok = QInputDialog.getText(self, "Rename Client / Job", "New name:", text=name)
        if not ok or not new_name.strip():
            return
        sanitized = sanitize_folder_name(new_name.strip())
        src = self._client_path(name)
        dst = self._client_path(sanitized)
        if dst.exists():
            self._set_status(f'"{sanitized}" already exists.', error=True)
            return
        try:
            src.rename(dst)
            self._refresh_list(select=sanitized)
            self._set_status(f'Renamed "{name}" -> "{sanitized}".')
        except OSError as e:
            self._set_status(f"Rename failed: {e}", error=True)

    def _on_archive_delete(self) -> None:
        name = self._selected_client()
        if not name:
            return

        # Use a custom dialog to guarantee left-to-right button order on all platforms
        dlg = QDialog(self)
        dlg.setWindowTitle("Archive / Delete Client")
        dlg.setFixedWidth(380)
        dlg_layout = QVBoxLayout(dlg)
        dlg_layout.setSpacing(12)
        dlg_layout.addWidget(QLabel(f'What would you like to do with "{name}"?'))
        btn_row = QHBoxLayout()
        btn_row.setSpacing(6)
        archive_btn        = QPushButton("Archive")
        archive_delete_btn = QPushButton("Archive && Delete")
        delete_btn         = QPushButton("Delete")
        cancel_btn         = QPushButton("Cancel")
        for b in (archive_btn, archive_delete_btn, delete_btn, cancel_btn):
            btn_row.addWidget(b)
        dlg_layout.addLayout(btn_row)
        choice = [None]
        archive_btn.clicked.connect(lambda: (choice.__setitem__(0, "archive"), dlg.accept()))
        archive_delete_btn.clicked.connect(lambda: (choice.__setitem__(0, "archive_delete"), dlg.accept()))
        delete_btn.clicked.connect(lambda: (choice.__setitem__(0, "delete"), dlg.accept()))
        cancel_btn.clicked.connect(dlg.reject)
        cancel_btn.setDefault(True)
        dlg.exec()

        if choice[0] in ("archive", "archive_delete"):
            zip_path = self._do_archive(name)
            if zip_path is None:
                return
            if choice[0] == "archive_delete":
                try:
                    shutil.rmtree(self._client_path(name))
                    self._refresh_list()
                    self._set_status(f'Archived and deleted "{name}" -> {Path(zip_path).name}')
                except OSError as e:
                    self._set_status(f'Archived but delete failed: {e}', error=True)
            else:
                self._set_status(f'Archived "{name}" -> {Path(zip_path).name}')
        elif choice[0] == "delete":
            reply = QMessageBox.question(
                self, "Confirm Delete",
                f'Permanently delete "{name}" and all its run data?\nThis cannot be undone.',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel,
                QMessageBox.StandardButton.Cancel,
            )
            if reply != QMessageBox.StandardButton.Yes:
                return
            try:
                shutil.rmtree(self._client_path(name))
                self._refresh_list()
                self._set_status(f'Deleted "{name}".')
            except OSError as e:
                self._set_status(f"Delete failed: {e}", error=True)

    def _on_tree_double_click(self, item: QTreeWidgetItem, _column: int) -> None:
        """Open a file with the OS default app on double-click (child items only)."""
        if item.parent() is None:
            return  # top-level run row — expand/collapse only
        name = self._selected_client()
        if not name:
            return
        run_name = item.parent().text(0).strip()
        file_name = item.text(0).strip()
        file_path = self._client_path(name) / run_name / file_name
        if file_path.exists():
            self._open_path(file_path)

    def _on_tree_context_menu(self, pos) -> None:
        """Right-click context menu on the runs tree."""
        from PySide6.QtWidgets import QMenu  # noqa: PLC0415
        item = self.runs_tree.itemAt(pos)
        if item is None:
            return
        menu = QMenu(self)
        if item.parent() is None:
            # Run-level item
            open_act   = menu.addAction("Open Run Folder")
            delete_act = menu.addAction("Delete Run")
            chosen = menu.exec(self.runs_tree.viewport().mapToGlobal(pos))
            if chosen is open_act:
                self._on_open_run_folder(item)
            elif chosen is delete_act:
                self._on_delete_run(item)
        else:
            # File-level item
            open_act = menu.addAction("Open File")
            chosen = menu.exec(self.runs_tree.viewport().mapToGlobal(pos))
            if chosen is open_act:
                self._on_tree_double_click(item, 0)

    def _on_open_run_folder(self, item: Optional[QTreeWidgetItem] = None) -> None:
        """Open the selected run folder or file depending on what is selected."""
        current = self.runs_tree.currentItem()
        if current is None:
            return
        name = self._selected_client()
        if not name:
            return
        if current.parent() is not None:
            # File-level — open the file
            run_name = current.parent().text(0).strip()
            file_name = current.text(0).strip()
            file_path = self._client_path(name) / run_name / file_name
            if file_path.exists():
                self._open_path(file_path)
        else:
            # Run-level — open the folder
            run_path = self._selected_run_path()
            if run_path and run_path.exists():
                self._open_path(run_path)

    def _on_delete_run(self, item: Optional[QTreeWidgetItem] = None) -> None:
        run_path = self._selected_run_path()
        if run_path is None:
            return
        reply = QMessageBox.question(
            self, "Delete Run",
            f'Delete run folder "{run_path.name}" and all its files?\nThis cannot be undone.',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel,
            QMessageBox.StandardButton.Cancel,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        try:
            shutil.rmtree(run_path)
            name = self._selected_client()
            if name:
                self._refresh_runs(self._client_path(name))
            self._set_status(f'Deleted run "{run_path.name}".')
        except OSError as e:
            self._set_status(f"Delete run failed: {e}", error=True)

    def _update_compare_button_state(self, client_path: Path) -> None:
        """Enable Compare Runs button only when the client has ≥2 runs with json/ data."""
        if not client_path.exists():
            self.compare_runs_button.setEnabled(False)
            return
        json_runs = [
            d for d in client_path.iterdir()
            if d.is_dir() and _TIMESTAMP_RE.match(d.name)
            and (d / "json").is_dir() and any((d / "json").glob("*.json"))
        ]
        self.compare_runs_button.setEnabled(len(json_runs) >= 2)

    def _on_compare_runs(self) -> None:
        name = self._selected_client()
        if not name:
            return
        dlg = DeviceDiffDialog(name, self._client_path(name), parent=self)
        dlg.exec()

    def _on_delete_cache_entry(self) -> None:
        name = self._selected_client()
        if not name:
            return
        row = self.cache_table.currentRow()
        if row < 0:
            return
        ip = self.cache_table.item(row, 0).text()
        reply = QMessageBox.question(
            self, "Remove Cache Entry",
            f'Remove "{ip}" from the device type cache for "{name}"?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel,
            QMessageBox.StandardButton.Cancel,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        cache_path = self._client_path(name) / DEVICE_CACHE_FILENAME
        with _device_cache_lock:
            cache = _load_device_cache(cache_path)
            if ip in cache:
                del cache[ip]
                _save_device_cache(cache_path, cache)
        self._refresh_cache(self._client_path(name))
        self._set_status(f'Removed cache entry for {ip}.')

    def _on_delete_all_cache(self) -> None:
        name = self._selected_client()
        if not name:
            return
        cache_path = self._client_path(name) / DEVICE_CACHE_FILENAME
        reply = QMessageBox.question(
            self, "Clear Device Cache",
            f'Delete device_cache.json for "{name}"?\nAll devices will be re-detected on the next run.',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel,
            QMessageBox.StandardButton.Cancel,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        try:
            cache_path.unlink()
            self._refresh_cache(self._client_path(name))
            self._set_status(f'Cache cleared for "{name}".')
        except OSError as e:
            self._set_status(f"Failed to clear cache: {e}", error=True)

    def _on_report(self) -> None:
        name = self._selected_client()
        if not name:
            return
        msg = QMessageBox(self)
        msg.setWindowTitle("Client Report")
        msg.setText(f'Export run history report for "{name}"?')
        create_btn = msg.addButton("Create Report", QMessageBox.ButtonRole.AcceptRole)
        msg.addButton("Cancel", QMessageBox.ButtonRole.RejectRole)
        msg.exec()
        if msg.clickedButton() is not create_btn:
            return

        client_path = self._client_path(name)
        runs = sorted(
            (d for d in client_path.iterdir() if d.is_dir() and _TIMESTAMP_RE.match(d.name)),
            key=lambda d: d.name
        )

        run_data = []
        for run in runs:
            info = self._read_run_info(run)
            files = sorted(
                f.name for f in run.iterdir()
                if f.is_file() and f.suffix in (".txt", ".xlsx", ".log", ".json")
            )
            run_data.append({
                "timestamp":    run.name,
                "device_file":  info.get("device_file", ""),
                "command_file": info.get("command_file", ""),
                "files":        files,
            })

        default_name = client_path / f"{name}_Report.xlsx"
        save_path, _ = QFileDialog.getSaveFileName(
            self, "Save Report", str(default_name), "Excel Files (*.xlsx)"
        )
        if not save_path:
            return

        try:
            from openpyxl import Workbook  # type: ignore
            from openpyxl.styles import Font, Alignment  # type: ignore
            wb = Workbook()
            ws = wb.active
            ws.title = "Runs"
            headers = ["Client", "Run Timestamp", "Device File", "Command File", "Files"]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(vertical="center")
            mid_align    = Alignment(vertical="center")
            files_align  = Alignment(wrap_text=True, vertical="center")
            for r in run_data:
                ws.append([name, r["timestamp"], r["device_file"], r["command_file"],
                            "\n".join(r["files"])])
                row_num = ws.max_row
                ws.row_dimensions[row_num].height = max(15, 15 * len(r["files"]))
                for cell in ws[row_num]:
                    cell.alignment = mid_align
                ws.cell(row=row_num, column=5).alignment = files_align
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(
                    60, max(len(str(c.value or "")) for c in col) + 4)
            wb.save(save_path)
            self._set_status(f"Report saved: {Path(save_path).name}")
        except Exception as e:
            self._set_status(f"Report failed: {e}", error=True)


class DeviceDiffDialog(QDialog):
    """Dialog for selecting two runs and generating a Device Diff report."""

    def __init__(self, client_name: str, client_path: Path, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.client_name = client_name
        self.client_path = client_path
        self.setWindowTitle(f'Compare Runs — "{client_name}"')
        self.setMinimumWidth(460)
        self._init_ui()
        self._update_preview()

    def _init_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setSpacing(8)

        # Discover runs with json/ subfolders
        self.runs: List[Path] = []
        if self.client_path.exists():
            self.runs = sorted(
                (d for d in self.client_path.iterdir()
                 if d.is_dir() and _TIMESTAMP_RE.match(d.name)
                 and (d / "json").is_dir() and any((d / "json").glob("*.json"))),
                key=lambda d: d.name,
            )

        run_names = [r.name for r in self.runs]

        # Run A (older)
        layout.addWidget(QLabel("Run A (older):"))
        self.combo_a = QComboBox()
        self.combo_a.addItems(run_names)
        if len(run_names) >= 2:
            self.combo_a.setCurrentIndex(len(run_names) - 2)
        self.combo_a.currentIndexChanged.connect(self._update_preview)
        layout.addWidget(self.combo_a)

        # Run B (newer)
        layout.addWidget(QLabel("Run B (newer):"))
        self.combo_b = QComboBox()
        self.combo_b.addItems(run_names)
        if run_names:
            self.combo_b.setCurrentIndex(len(run_names) - 1)
        self.combo_b.currentIndexChanged.connect(self._update_preview)
        layout.addWidget(self.combo_b)

        # Port Map (optional)
        pm_layout = QHBoxLayout()
        pm_layout.addWidget(QLabel("Port Map (optional):"))
        self.port_map_input = QLineEdit()
        self.port_map_input.setPlaceholderText("CSV file mapping old \u2192 new interface names")
        self.port_map_input.setReadOnly(True)
        pm_layout.addWidget(self.port_map_input, stretch=1)
        browse_pm_btn = QPushButton("Browse\u2026")
        browse_pm_btn.setFixedWidth(80)
        browse_pm_btn.clicked.connect(self._browse_port_map)
        pm_layout.addWidget(browse_pm_btn)
        clear_pm_btn = QPushButton("Clear")
        clear_pm_btn.setFixedWidth(50)
        clear_pm_btn.clicked.connect(lambda: self.port_map_input.clear())
        pm_layout.addWidget(clear_pm_btn)
        edit_pm_btn = QPushButton("Edit\u2026")
        edit_pm_btn.setFixedWidth(55)
        edit_pm_btn.clicked.connect(self._edit_port_map)
        pm_layout.addWidget(edit_pm_btn)
        layout.addLayout(pm_layout)

        # Preview label
        self.preview_label = QLabel("")
        self.preview_label.setWordWrap(True)
        self.preview_label.setStyleSheet("color: #555;")
        layout.addWidget(self.preview_label)

        # Buttons
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        self.compare_btn = QPushButton("Compare")
        self.compare_btn.setDefault(True)
        cancel_btn = QPushButton("Cancel")
        self.compare_btn.clicked.connect(self._on_compare)
        cancel_btn.clicked.connect(self.reject)
        btn_row.addWidget(self.compare_btn)
        btn_row.addWidget(cancel_btn)
        layout.addLayout(btn_row)

    def _update_preview(self) -> None:
        idx_a = self.combo_a.currentIndex()
        idx_b = self.combo_b.currentIndex()
        if idx_a < 0 or idx_b < 0 or not self.runs:
            self.preview_label.setText("No valid runs available.")
            self.compare_btn.setEnabled(False)
            return
        if idx_a == idx_b:
            self.preview_label.setText("Please select two different runs.")
            self.compare_btn.setEnabled(False)
            return
        run_a = self.runs[idx_a]
        run_b = self.runs[idx_b]
        count_a = len(list((run_a / "json").glob("*.json")))
        count_b = len(list((run_b / "json").glob("*.json")))
        self.preview_label.setText(
            f"Run A: {count_a} device(s)  |  Run B: {count_b} device(s)")
        self.compare_btn.setEnabled(True)

    def _browse_port_map(self) -> None:
        """Open a file dialog to select a port map CSV file."""
        start_dir = str(Path(__file__).parent / "input")
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select Port Map CSV", start_dir,
            "CSV Files (*.csv);;All Files (*)",
        )
        if file_path:
            self.port_map_input.setText(file_path)

    def _edit_port_map(self) -> None:
        """Open the port map CSV in the file editor."""
        pm_path = self.port_map_input.text().strip()
        if not pm_path:
            # Default to the sample file
            pm_path = str(Path(__file__).parent / "input" / "sample_port_map.csv")
        path = Path(pm_path)
        if not path.exists():
            QMessageBox.warning(self, "File Not Found",
                                f"Port map file not found:\n{path}")
            return
        editor = FileEditorDialog(path, "port_map", self)
        editor.file_path_updated.connect(
            lambda new_path, _ft: self.port_map_input.setText(str(new_path)))
        editor.exec()

    def _on_compare(self) -> None:
        idx_a = self.combo_a.currentIndex()
        idx_b = self.combo_b.currentIndex()
        run_a = self.runs[idx_a]
        run_b = self.runs[idx_b]

        timestamp = datetime.now().strftime("%m-%d-%Y_%I-%M_%p")
        default_name = f"{self.client_name}_diff_{timestamp}.xlsx"
        save_path, _ = QFileDialog.getSaveFileName(
            self, "Save Diff Report", str(self.client_path / default_name),
            "Excel Workbook (*.xlsx)",
        )
        if not save_path:
            return

        # Load port map if provided
        port_map: Optional[Dict[str, str]] = None
        pm_path = self.port_map_input.text().strip()
        if pm_path:
            try:
                port_map = _load_port_map(Path(pm_path))
            except Exception as e:
                QMessageBox.warning(self, "Port Map Error",
                                    f"Failed to load port map:\n{e}")
                return

        logger = logging.getLogger(APP_NAME)
        result = generate_diff_report(run_a, run_b, Path(save_path), logger, port_map=port_map)
        if result:
            QMessageBox.information(self, "Diff Report", f"Report saved:\n{result.name}")
            self.accept()
        else:
            QMessageBox.warning(self, "Diff Report", "Failed to generate diff report.")


class SettingsDialog(QDialog):
    """Settings dialog for persistent preferences."""

    SETTINGS_FILE = str(Path(__file__).parent / "nddu.ini")

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("Settings")
        self.setFixedWidth(460)
        self.settings = QSettings(self.SETTINGS_FILE, QSettings.Format.IniFormat)
        self._init_ui()
        self._load()

    # ------------------------------------------------------------------
    # UI construction
    # ------------------------------------------------------------------
    def _init_ui(self) -> None:
        outer = QVBoxLayout(self)
        outer.setSpacing(10)
        outer.setContentsMargins(12, 12, 12, 12)

        # ── Excel Report Components ───────────────────────────────────
        comp_group = QGroupBox("Excel Report Components")
        comp_outer = QVBoxLayout()
        comp_outer.setSpacing(4)

        comp_btn_row = QHBoxLayout()
        comp_btn_row.setSpacing(6)
        comp_lbl = QLabel("Include in report:", self)
        comp_lbl.setStyleSheet("font-size: 11px;")
        comp_btn_row.addWidget(comp_lbl)
        self.comp_all_btn = QPushButton("All", self)
        self.comp_all_btn.setFixedWidth(40)
        self.comp_all_btn.setFixedHeight(20)
        self.comp_all_btn.clicked.connect(lambda: self._set_all_components(True))
        comp_btn_row.addWidget(self.comp_all_btn)
        self.comp_none_btn = QPushButton("None", self)
        self.comp_none_btn.setFixedWidth(40)
        self.comp_none_btn.setFixedHeight(20)
        self.comp_none_btn.clicked.connect(lambda: self._set_all_components(False))
        comp_btn_row.addWidget(self.comp_none_btn)
        comp_btn_row.addStretch()
        comp_outer.addLayout(comp_btn_row)

        # Two columns of checkboxes
        self.component_checkboxes: Dict[str, QCheckBox] = {}
        cols_layout = QHBoxLayout()
        cols_layout.setSpacing(16)
        col1 = QVBoxLayout()
        col1.setSpacing(2)
        col2 = QVBoxLayout()
        col2.setSpacing(2)
        for i, name in enumerate(COMPONENT_NAMES):
            cb = QCheckBox(name, self)
            cb.setChecked(True)
            self.component_checkboxes[name] = cb
            (col1 if i < 4 else col2).addWidget(cb)
        cols_layout.addLayout(col1)
        cols_layout.addLayout(col2)
        cols_layout.addStretch()
        comp_outer.addLayout(cols_layout)

        comp_group.setLayout(comp_outer)
        outer.addWidget(comp_group)

        # ── Output ────────────────────────────────────────────────────
        output_group = QGroupBox("Output")
        output_layout = QVBoxLayout()
        output_layout.setSpacing(4)

        self.save_txt_cb = QCheckBox("Save Excel Report structured data in TXT output files", self)
        self.save_txt_cb.setToolTip(
            "When checked, inventory command output (used for Excel Report structured data)\n"
            "is also written to the per-device .txt files alongside regular command output.\n"
            "When unchecked, .txt files contain only the commands from Commands.txt.\n"
            "Has no effect when Excel Report is disabled."
        )
        output_layout.addWidget(self.save_txt_cb)

        self.combined_output_cb = QCheckBox("Combined Output File", self)
        self.combined_output_cb.setToolTip(
            "When checked, all per-device output is merged into a single Combined.txt file\n"
            "in the client output folder after the run completes."
        )
        output_layout.addWidget(self.combined_output_cb)

        output_group.setLayout(output_layout)
        outer.addWidget(output_group)

        # ── Reset ─────────────────────────────────────────────────────
        reset_group = QGroupBox("Reset")
        reset_layout = QVBoxLayout()
        reset_layout.setSpacing(4)

        self.reset_include_files_cb = QCheckBox("Also reset default Input Files to original defaults", self)
        self.reset_include_files_cb.setToolTip(
            "When checked, any custom default Device or Command files set via\n"
            "'Set as Default' in the file editor will also be cleared."
        )
        reset_layout.addWidget(self.reset_include_files_cb)

        self.reset_btn = QPushButton("Reset to Defaults", self)
        self.reset_btn.setToolTip(
            "Select all Excel components, reset the TXT option, and\n"
            "clear all remembered credentials and Client/Job values."
        )
        self.reset_btn.clicked.connect(self._on_reset)
        reset_layout.addWidget(self.reset_btn)

        self.status_label_reset = QLabel("", self)
        self.status_label_reset.setStyleSheet("font-size: 10px; color: gray;")
        reset_layout.addWidget(self.status_label_reset)

        reset_group.setLayout(reset_layout)
        outer.addWidget(reset_group)

        # ── Buttons ───────────────────────────────────────────────────
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        self.ok_btn = QPushButton("OK", self)
        self.ok_btn.setFixedWidth(80)
        self.ok_btn.clicked.connect(self._on_ok)
        self.cancel_btn = QPushButton("Cancel", self)
        self.cancel_btn.setFixedWidth(80)
        self.cancel_btn.clicked.connect(self.reject)
        btn_row.addWidget(self.ok_btn)
        btn_row.addWidget(self.cancel_btn)
        outer.addLayout(btn_row)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    def _on_reset(self) -> None:
        """Reset all settings to defaults in the UI (does not save until OK is clicked)."""
        self.save_txt_cb.setChecked(False)
        self.combined_output_cb.setChecked(False)
        self._set_all_components(True)
        if self.reset_include_files_cb.isChecked():
            self.settings.setValue("defaults/device_file",  str(DEFAULT_DEVICE_FILE))
            self.settings.setValue("defaults/command_file", str(DEFAULT_COMMAND_FILE))
            self.reset_include_files_cb.setChecked(False)
        # Clear remembered credential/client values and their flags
        for key in ("remember/username", "remember/keyring_sys", "remember/client",
                    "remembered/username", "remembered/keyring_system", "remembered/client"):
            self.settings.remove(key)
        self.status_label_reset.setText("Settings reset to defaults — click OK to save.")

    def _set_all_components(self, checked: bool) -> None:
        for cb in self.component_checkboxes.values():
            cb.setChecked(checked)

    def get_selected_components(self) -> Set[str]:
        return {name for name, cb in self.component_checkboxes.items() if cb.isChecked()}

    # ------------------------------------------------------------------
    # Load / Save
    # ------------------------------------------------------------------
    def _load(self) -> None:
        s = self.settings
        self.save_txt_cb.setChecked(s.value("output/save_txt", False, type=bool))
        self.combined_output_cb.setChecked(s.value("output/combined", False, type=bool))
        saved = s.value("excel/components", list(COMPONENT_NAMES))
        if isinstance(saved, str):          # QSettings on some platforms serialises as CSV
            saved = [x.strip() for x in saved.split(",") if x.strip()]
        for name, cb in self.component_checkboxes.items():
            cb.setChecked(name in saved)

    def _save(self) -> None:
        s = self.settings
        s.setValue("output/save_txt",  self.save_txt_cb.isChecked())
        s.setValue("output/combined",  self.combined_output_cb.isChecked())
        s.setValue("excel/components", list(self.get_selected_components()))

    def _on_ok(self) -> None:
        self._save()
        self.accept()

    # ------------------------------------------------------------------
    # Static helpers for reading settings without opening the dialog
    # ------------------------------------------------------------------
    @staticmethod
    def read(key: str, default: Any = None, type: type = str) -> Any:  # noqa: A002
        return QSettings(SettingsDialog.SETTINGS_FILE, QSettings.Format.IniFormat).value(key, default, type=type)

    @staticmethod
    def write(key: str, value: Any) -> None:
        QSettings(SettingsDialog.SETTINGS_FILE, QSettings.Format.IniFormat).setValue(key, value)


class VersionChecker(QObject):
    """Check for updates using GitHub Releases API."""
    update_found = Signal(str, str)  # (new_version, release_url)
    check_complete = Signal(bool)  # Whether check was successful
    
    def check(self) -> None:
        """Check for updates in a non-blocking way."""
        try:
            # Create request with headers (GitHub API likes User-Agent)
            headers = {
                'User-Agent': f'{APP_NAME}/{APP_VERSION}',
                'Accept': 'application/vnd.github.v3+json'
            }
            req = urllib.request.Request(GITHUB_API_LATEST_RELEASE, headers=headers)
            
            with urllib.request.urlopen(req, timeout=3) as response:
                data = json.loads(response.read().decode())
                latest_tag = data.get('tag_name', '')  # e.g., "v1.1.0"
                current_version = APP_VERSION  # e.g., "v1.0.0"
                
                # Remove 'v' prefix for comparison
                latest_version_str = latest_tag.lstrip('v')
                current_version_str = current_version.lstrip('v')
                
                # Compare versions
                latest_ver = version.parse(latest_version_str)
                current_ver = version.parse(current_version_str)
                
                if latest_ver > current_ver:
                    release_url = data.get('html_url', REPO_URL)
                    self.update_found.emit(latest_version_str, release_url)
                
                self.check_complete.emit(True)
                
        except Exception:
            # Silently fail - don't interrupt user
            self.check_complete.emit(False)

class MyWindow(QWidget):
    """Main application window for the Network Device Documentation Utility."""
    
    def __init__(self) -> None:
        """Initialize the main window with default settings."""
        super().__init__()
        self.dark_mode = DARK_MODE_STATE  # Set the dark mode state
        self.enable_was_enabled = False
        self.verbose_was_enabled = False
        self.structured_output_was_enabled = False
        self.stop_requested = False
        self.update_available = False
        self.new_version = ""
        self.release_url = ""
        self.toggle_theme(self.dark_mode)  # Toggle theme according to DARK_MODE_STATE
        self.init_ui()
        self._apply_remembered_settings()

        # Start update check after UI is shown
        QTimer.singleShot(1000, self.check_for_updates)

    def configure_logging(self, log_file: str) -> None:
        """
        Configure logging for the application.
        
        Args:
            log_file: Path to the log file
        """
        # Configure root logger
        self.logger = logging.getLogger()
        self.logger.setLevel(logging.INFO)
        
        # Clear existing handlers
        for handler in self.logger.handlers[:]:
            self.logger.removeHandler(handler)

        # Create formatter
        formatter = CustomFormatter()
        
        # Configure file handler
        file_handler = logging.FileHandler(log_file)
        file_handler.setFormatter(formatter)  # Apply to file handler
        
        # Configure GUI handler
        self.gui_handler = GUILogHandler()
        self.gui_handler.setFormatter(logging.Formatter('%(message)s'))  # Raw for GUI
        
        # Add both handlers
        self.logger.addHandler(self.gui_handler)
        self.logger.addHandler(file_handler)
        
        # Drain the log queue at ~150 ms intervals rather than per-message.
        # This prevents cross-thread signal floods from starving the Qt event loop
        # when many devices are being processed simultaneously.
        self._log_drain_timer = QTimer(self)
        self._log_drain_timer.setInterval(150)
        self._log_drain_timer.timeout.connect(self._drain_log_queue)
        self._log_drain_timer.start()

    def _drain_log_queue(self) -> None:
        """Flush all pending log messages from the queue into the UI in one batch.

        Builds a single HTML string for the entire batch and calls insertHtml()
        once, avoiding N layout recalculations and N scroll resets.
        """
        q = self.gui_handler._queue
        messages = []
        try:
            while True:
                messages.append(q.get_nowait())
        except queue.Empty:
            pass
        if not messages:
            return

        parts = []
        for msg, level in messages:
            color   = self.get_color_for_level(level)
            html    = self.format_message(msg, level)
            parts.append(f'<span style="color:{color}">{html}</span><br>' if color
                         else f'{html}<br>')

        widget = self.log_output
        cursor = widget.textCursor()
        cursor.movePosition(QTextCursor.MoveOperation.End)
        widget.setUpdatesEnabled(False)
        try:
            cursor.insertHtml(''.join(parts))
            widget.setTextCursor(cursor)
        finally:
            widget.setUpdatesEnabled(True)
        widget.ensureCursorVisible()
        widget.horizontalScrollBar().setValue(0)

    def toggle_theme(self, dark_mode: bool = True) -> None:
        """
        Toggle between dark and light themes.
        
        Args:
            dark_mode: Whether to enable dark mode (True) or light mode (False)
        """
        app = QApplication.instance()
        app.setStyle("Fusion")  # Use Fusion style as base for dark theme
        
        if dark_mode:
            dark_palette = QPalette()
            dark_palette.setColor(QPalette.ColorRole.Window, QColor(53, 53, 53))
            dark_palette.setColor(QPalette.ColorRole.WindowText, Qt.GlobalColor.white)
            dark_palette.setColor(QPalette.ColorRole.Base, QColor(35, 35, 35))
            dark_palette.setColor(QPalette.ColorRole.AlternateBase, QColor(53, 53, 53))
            dark_palette.setColor(QPalette.ColorRole.ToolTipBase, QColor(25, 25, 25))
            dark_palette.setColor(QPalette.ColorRole.ToolTipText, Qt.GlobalColor.white)
            dark_palette.setColor(QPalette.ColorRole.Text, Qt.GlobalColor.white)
            dark_palette.setColor(QPalette.ColorRole.Button, QColor(53, 53, 53))
            dark_palette.setColor(QPalette.ColorRole.ButtonText, Qt.GlobalColor.white)
            dark_palette.setColor(QPalette.ColorRole.BrightText, Qt.GlobalColor.red)
            dark_palette.setColor(QPalette.ColorRole.Link, QColor(42, 130, 218))
            dark_palette.setColor(QPalette.ColorRole.Highlight, QColor(42, 130, 218))
            dark_palette.setColor(QPalette.ColorRole.HighlightedText, Qt.GlobalColor.black)
            app.setPalette(dark_palette)
            
            # Additional dark mode styling
            app.setStyleSheet("""
                QToolTip {
                    color: #ffffff;
                    background-color: #2a82da;
                    border: 1px solid white;
                }
                QGroupBox {
                    border: 1px solid gray;
                    border-radius: 3px;
                    margin-top: 0.5em;
                }
                QGroupBox::title {
                    subcontrol-origin: margin;
                    left: 10px;
                    padding: 0 3px;
                }
            """)
            self.dark_mode = True
        else:
            app.setPalette(app.style().standardPalette())
            app.setStyleSheet("")
            self.dark_mode = False

    def init_ui(self) -> None:
        """Initialize all UI components and layouts."""
        # Set window properties
        self.setWindowTitle(f"nddu")
        self.setMaximumSize(730, 1000)  # Width, Height

        # Define the relative path to the "input" folder
        self.input_folder = DEFAULT_INPUT_FOLDER
        self.default_device_list = DEFAULT_DEVICE_FILE
        self.default_command_list = DEFAULT_COMMAND_FILE

        # Create the logo and title section
        logo_title_layout = QHBoxLayout()
        logo_title_layout.setSpacing(5)

        # Load and resize the logo
        self.logo = QLabel(self)
        logo_path = LOGO_PATH
        if logo_path.exists():
            pixmap = QPixmap(str(logo_path))
            pixmap = pixmap.scaledToHeight(80, Qt.TransformationMode.SmoothTransformation)
            self.logo.setPixmap(pixmap)
        else:
            self.logo.setText("Logo not found")
            self.logo.setStyleSheet("color: red;")
        logo_title_layout.addSpacing(10)
        logo_title_layout.addWidget(self.logo)
        logo_title_layout.addSpacing(20)

        # Create a container widget with fixed height
        title_container = QWidget()
        title_container.setFixedHeight(60)  # Enough for title + update indicator
        title_layout = QVBoxLayout(title_container)
        title_layout.setContentsMargins(0, 0, 0, 0)  # No margins
        title_layout.setSpacing(2)

        # Add the script name and version
        self.title_label = QLabel(
            f"<span style='font-size: 18px; font-weight: bold;'>{APP_NAME}</span><br>"
            f"<span style='font-size: 12px;'>{APP_VERSION}</span>", 
            self
        )
        self.title_label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        title_layout.addWidget(self.title_label)

        # Add update indicator (initially empty)
        self.update_indicator = QLabel("", self)
        self.update_indicator.setStyleSheet("color: #4CAF50; font-size: 11px;")
        self.update_indicator.setOpenExternalLinks(True)
        self.update_indicator.setCursor(Qt.PointingHandCursor)
        title_layout.addWidget(self.update_indicator)

        # Now add the container to the logo_title_layout
        logo_title_layout.addWidget(title_container)
        logo_title_layout.addStretch()

        # Create the Input Files section
        input_files_group = QGroupBox("Input Files")
        input_files_layout = QVBoxLayout()
        input_files_layout.setSpacing(5)

        # Device List
        self.device_label = QLabel("Device(s):", self)
        self.device_input = QLineEdit(self)
        self.device_input.setFixedWidth(280)
        self.device_input.setReadOnly(True)
        self.device_button = QPushButton("Browse", self)
        self.device_button.clicked.connect(self.browse_device_list)
        self.device_edit_button = QPushButton("Edit", self)
        self.device_edit_button.clicked.connect(self.edit_device_file)
        device_layout = QHBoxLayout()
        device_layout.setSpacing(5)
        device_layout.addWidget(self.device_input)
        device_layout.addWidget(self.device_edit_button)
        device_layout.addWidget(self.device_button)
        input_files_layout.addWidget(self.device_label)
        input_files_layout.addLayout(device_layout)

        # Command List
        self.command_label = QLabel("Command(s):", self)
        self.command_input = QLineEdit(self)
        self.command_input.setFixedWidth(280)
        self.command_input.setReadOnly(True)
        self.command_button = QPushButton("Browse", self)
        self.command_button.clicked.connect(self.browse_command_list)
        self.command_edit_button = QPushButton("Edit", self)
        self.command_edit_button.clicked.connect(self.edit_command_file)
        command_layout = QHBoxLayout()
        command_layout.setSpacing(5)
        command_layout.addWidget(self.command_input)
        command_layout.addWidget(self.command_edit_button)
        command_layout.addWidget(self.command_button)
        input_files_layout.addWidget(self.command_label)
        input_files_layout.addLayout(command_layout)
        input_files_group.setLayout(input_files_layout)

        # Create the Credential Options section
        credentials_group = QGroupBox("Credential Options")
        credentials_layout = QVBoxLayout()
        credentials_layout.setSpacing(5)

        # Manual Credentials radio row
        manual_radio_row = QHBoxLayout()
        manual_radio_row.setSpacing(8)
        self.manual_radio = QRadioButton("Manual Credentials", self)
        self.manual_radio.setChecked(True)
        self.manual_radio.toggled.connect(self.toggle_credential_options)
        manual_radio_row.addWidget(self.manual_radio)
        self.remember_username_cb = QCheckBox("Remember Username", self)
        self.remember_username_cb.setChecked(
            SettingsDialog.read("remember/username", False, type=bool)
        )
        self.remember_username_cb.setToolTip(
            "Save the username when Go runs successfully.\n"
            "Restores the username and selects Manual Credentials on next launch."
        )
        manual_radio_row.addWidget(self.remember_username_cb)
        manual_radio_row.addStretch()
        credentials_layout.addLayout(manual_radio_row)

        # Manual Credentials sub-group
        self.manual_credentials_group = QGroupBox()
        manual_credentials_layout = QVBoxLayout()
        manual_credentials_layout.setSpacing(5)
        manual_credentials_layout.setContentsMargins(4, 4, 4, 4)

        # Username
        username_layout = QHBoxLayout()
        username_layout.setSpacing(5)
        self.username_label = QLabel("Username:", self)
        self.username_input = QLineEdit(self)
        self.username_input.setFixedWidth(280)
        self.username_input.textChanged.connect(self.validate_fields)
        username_layout.addWidget(self.username_label)
        username_layout.addWidget(self.username_input)
        manual_credentials_layout.addLayout(username_layout)

        # Password
        password_layout = QHBoxLayout()
        password_layout.setSpacing(5)
        self.password_label = QLabel("Password:", self)
        self.password_input = QLineEdit(self)
        self.password_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.password_input.setFixedWidth(280)
        self.password_input.textChanged.connect(self.validate_fields)
        self.password_input.textChanged.connect(self.sync_enable_password)
        password_layout.addWidget(self.password_label)
        password_layout.addWidget(self.password_input)
        manual_credentials_layout.addLayout(password_layout)

        # Enable Checkbox
        self.enable_checkbox = QCheckBox("Enable:", self)
        self.enable_checkbox.setChecked(False)
        self.enable_checkbox.stateChanged.connect(self.toggle_enable_field)

        # Enable
        enable_layout = QHBoxLayout()
        enable_layout.setSpacing(5)
        self.enable_input = QLineEdit(self)
        self.enable_input.setDisabled(True)
        self.enable_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.enable_input.setFixedWidth(280)
        self.enable_input.textChanged.connect(self.validate_fields)
        enable_layout.addWidget(self.enable_checkbox)
        enable_layout.addWidget(self.enable_input)
        manual_credentials_layout.addLayout(enable_layout)

        self.manual_credentials_group.setLayout(manual_credentials_layout)
        manual_indent = QHBoxLayout()
        manual_indent.setContentsMargins(16, 0, 0, 0)
        manual_indent.addWidget(self.manual_credentials_group)
        credentials_layout.addLayout(manual_indent)

        # Keyring Credentials radio row
        keyring_radio_row = QHBoxLayout()
        keyring_radio_row.setSpacing(8)
        self.keyring_radio = QRadioButton("Keyring Credentials", self)
        self.keyring_radio.toggled.connect(self.toggle_credential_options)
        keyring_radio_row.addWidget(self.keyring_radio)
        self.remember_keyring_system_cb = QCheckBox("Remember System Name", self)
        self.remember_keyring_system_cb.setChecked(
            SettingsDialog.read("remember/keyring_sys", False, type=bool)
        )
        self.remember_keyring_system_cb.setToolTip(
            "Save the Keyring System Name when Go runs successfully.\n"
            "Restores the system name and selects Keyring Credentials on next launch."
        )
        keyring_radio_row.addWidget(self.remember_keyring_system_cb)
        keyring_radio_row.addStretch()
        credentials_layout.addLayout(keyring_radio_row)

        # Keyring Credentials sub-group
        self.keyring_credentials_group = QGroupBox()
        keyring_credentials_layout = QVBoxLayout()
        keyring_credentials_layout.setSpacing(5)
        keyring_credentials_layout.setContentsMargins(4, 4, 4, 4)

        # Keyring System Name
        keyring_system_layout = QHBoxLayout()
        keyring_system_layout.setSpacing(5)
        self.keyring_system_label = QLabel("System Name:", self)
        self.keyring_system_input = QLineEdit(self)
        self.keyring_system_input.setFixedWidth(280)
        self.keyring_system_input.textChanged.connect(self.validate_fields)
        keyring_system_layout.addWidget(self.keyring_system_label)
        keyring_system_layout.addWidget(self.keyring_system_input)
        keyring_credentials_layout.addLayout(keyring_system_layout)

        # Keyring User Name
        keyring_user_layout = QHBoxLayout()
        keyring_user_layout.setSpacing(5)
        self.keyring_user_label = QLabel("Username:", self)
        self.keyring_user_input = QLineEdit(self)
        self.keyring_user_input.setFixedWidth(280)
        self.keyring_user_input.textChanged.connect(self.validate_fields)
        keyring_user_layout.addWidget(self.keyring_user_label)
        keyring_user_layout.addWidget(self.keyring_user_input)
        keyring_credentials_layout.addLayout(keyring_user_layout)

        self.keyring_credentials_group.setLayout(keyring_credentials_layout)
        keyring_indent = QHBoxLayout()
        keyring_indent.setContentsMargins(16, 0, 0, 0)
        keyring_indent.addWidget(self.keyring_credentials_group)
        credentials_layout.addLayout(keyring_indent)

        credentials_group.setLayout(credentials_layout)

        # Add the Script Options section
        options_group = QGroupBox("Script Options")
        options_layout = QVBoxLayout()
        options_layout.setSpacing(5)

        # Client / Job field — editable dropdown populated from existing output subfolders
        client_layout = QHBoxLayout()
        client_layout.setSpacing(5)
        self.client_label = QLabel("    Client / Job:", self)
        self.client_input = QComboBox(self)
        self.client_input.setEditable(True)
        self.client_input.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.client_input.setFixedWidth(240)
        self.client_input.addItem("")  # blank = no client subfolder
        for client in get_existing_clients():
            self.client_input.addItem(client)
        self.client_input.setCurrentIndex(0)
        self.client_input.lineEdit().setPlaceholderText('Saves output in "./output/{Client}/"')
        self.client_input.lineEdit().setMaxLength(64)
        self.client_input.lineEdit().editingFinished.connect(self._sanitize_client_input)
        self.remember_client_cb = QCheckBox("Remember", self)
        self.remember_client_cb.setChecked(
            SettingsDialog.read("remember/client", False, type=bool)
        )
        self.remember_client_cb.setToolTip(
            "Save the Client/Job name when Go runs successfully.\n"
            "Restores the selection on next launch."
        )
        client_layout.addWidget(self.client_label)
        client_layout.addWidget(self.client_input)
        client_layout.addWidget(self.remember_client_cb)
        options_layout.addLayout(client_layout)

        # First row of checkboxes
        first_row_layout = QHBoxLayout()
        first_row_layout.setSpacing(40)
        
        # Verbose Output checkbox
        self.verbose_checkbox = QCheckBox("Verbose Output", self)
        self.verbose_checkbox.setChecked(False)
        first_row_layout.addWidget(self.verbose_checkbox)

        # Structured Output checkbox (implies device type auto-detection + TextFSM parsing + JSON)
        self.structured_output_checkbox = QCheckBox("Excel Report", self)
        self.structured_output_checkbox.setChecked(False)
        self.structured_output_checkbox.toggled.connect(self._on_excel_report_toggled)
        first_row_layout.addWidget(self.structured_output_checkbox)

        # Force Re-detect checkbox — visible only when Excel Report is checked
        self.force_redetect_checkbox = QCheckBox("Force Re-detect", self)
        self.force_redetect_checkbox.setChecked(False)
        self.force_redetect_checkbox.setToolTip(
            "Ignore cached device types and re-detect all devices.\n"
            "Use when devices have been replaced or upgraded."
        )
        self.force_redetect_checkbox.setVisible(False)
        first_row_layout.addWidget(self.force_redetect_checkbox)

        # Add stretch to push checkboxes to the left
        first_row_layout.addStretch()
        options_layout.addLayout(first_row_layout)

        # Second row of checkboxes
        second_row_layout = QHBoxLayout()
        second_row_layout.setSpacing(40)

        # Add stretch to push checkboxes to the left
        second_row_layout.addStretch()
        options_layout.addLayout(second_row_layout)
        
        options_group.setLayout(options_layout)

        # Actions section
        actions_group = QGroupBox("Actions")
        actions_layout = QVBoxLayout()
        actions_layout.setSpacing(5)

        # First row of Actions buttons
        top_button_layout = QHBoxLayout()
        top_button_layout.setSpacing(5)
        self.help_button = QPushButton("Help", self)
        self.help_button.clicked.connect(self.show_help)
        self.settings_button = QPushButton("Settings", self)
        self.settings_button.clicked.connect(self.show_settings)
        self.quit_button = QPushButton("Quit", self)
        self.quit_button.clicked.connect(self.close)
        self.go_button = QPushButton("Go", self)
        self.go_button.clicked.connect(self.on_go)
        self.go_button.setEnabled(False)
        self.update_go_button_style()  # New method to handle button styling
        top_button_layout.addWidget(self.help_button)
        top_button_layout.addWidget(self.settings_button)
        top_button_layout.addWidget(self.quit_button)
        top_button_layout.addWidget(self.go_button)

        # Second row of Actions buttons
        bottom_button_layout = QHBoxLayout()
        bottom_button_layout.setSpacing(5)
        self.keyring_tools_button = QPushButton("Keyring Tools", self)
        self.keyring_tools_button.clicked.connect(self.open_keyring_tools)
        self.show_output_button = QPushButton("Open Output Folder", self)
        self.show_output_button.clicked.connect(self.show_output_folder)
        self.manage_clients_button = QPushButton("Manage Clients", self)
        self.manage_clients_button.clicked.connect(self.show_client_manager)
        bottom_button_layout.addWidget(self.keyring_tools_button)
        bottom_button_layout.addWidget(self.manage_clients_button)
        bottom_button_layout.addWidget(self.show_output_button)

        actions_layout.addLayout(top_button_layout)
        actions_layout.addLayout(bottom_button_layout)
        actions_group.setLayout(actions_layout)

        # Add Output section with scroll bars
        output_group = QGroupBox("Output")
        output_layout = QVBoxLayout()
        output_layout.setSpacing(5)
        self.progress_bar = QProgressBar(self)
        self.progress_bar.setValue(0)
        output_layout.addWidget(self.progress_bar)

        # Create a scroll area for the log output
        self.log_output = QTextEdit(self)
        self.log_output.setReadOnly(True)
        self.log_output.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop)
        self.log_output.setLineWrapMode(QTextEdit.LineWrapMode.NoWrap)  # Disable text wrapping
        self.log_output.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)  # Enable horizontal scrollbar
        self.log_output.horizontalScrollBar().setValue(0)  # Set horizontal scrollbar to the left initially
        # Connect the textChanged signal to a custom slot to control scrolling behavior
        # Horizontal scroll is reset in _drain_log_queue after each batch insert.

        scroll_area = QScrollArea(self)
        scroll_area.setWidgetResizable(True)
        scroll_area.setWidget(self.log_output)
        scroll_area.setMinimumHeight(150)
        output_layout.addWidget(scroll_area)
        output_group.setLayout(output_layout)

        # Set up the main layout
        self.main_layout = QVBoxLayout()
        self.main_layout.setSpacing(5)
        self.main_layout.addLayout(logo_title_layout)
        self.main_layout.addSpacing(10)
        self.main_layout.addWidget(input_files_group)
        self.main_layout.addWidget(credentials_group)
        self.main_layout.addWidget(options_group)
        self.main_layout.addWidget(actions_group)
        self.main_layout.addWidget(output_group)
        self.setLayout(self.main_layout)

        # Set default files
        self.set_default_files()

        # Initialize credential options
        self.toggle_credential_options()

        # Reduce margins for the main layout
        self.main_layout.setContentsMargins(15, 5, 15, 15)  # Left, Top, Right, Bottom margins
        input_files_layout.setContentsMargins(5, 5, 5, 5)  # Reduce margins for the Input Files layout
        credentials_layout.setContentsMargins(5, 5, 5, 5)  # Reduce margins for the Credentials layout
        options_layout.setContentsMargins(5, 5, 5, 5)  # Reduce margins for the Options layout
        actions_layout.setContentsMargins(5, 5, 5, 5)  # Reduce margins for the Actions layout
        output_layout.setContentsMargins(5, 5, 5, 5)  # Reduce margins for the Output layout

    def check_for_updates(self) -> None:
        """Start background update check."""
        self.version_checker = VersionChecker()
        self.version_checker.update_found.connect(self.on_update_found)
        self.version_checker.check_complete.connect(self.on_check_complete)
        
        # Run in a thread to avoid blocking UI
        self.check_thread = QThread()
        self.version_checker.moveToThread(self.check_thread)
        self.check_thread.started.connect(self.version_checker.check)
        self.check_thread.start()

    def on_update_found(self, new_version: str, release_url: str) -> None:
        """Handle when an update is found."""
        self.update_available = True
        self.new_version = new_version
        self.release_url = release_url
        
        # Show update indicator
        update_text = f'<a href="{release_url}" style="color: #4CAF50; text-decoration: none;">'
        update_text += f'Update available: v{new_version} ↗</a>'
        self.update_indicator.setText(update_text)
        
        # Also log to output
        # self.append_colored_message(f"Update available: v{new_version} (current: v{APP_VERSION.lstrip('v')})", "INFO")

    def on_check_complete(self, success: bool) -> None:
        """Clean up after update check."""
        if hasattr(self, 'check_thread'):
            self.check_thread.quit()
            self.check_thread.wait()
            
        if not success and not self.update_available:
            # Check failed but that's OK - we don't show errors
            pass

    def keep_horizontal_scroll_left(self) -> None:
        """Ensure the horizontal scrollbar stays on the left side when new text is added."""
        self.log_output.horizontalScrollBar().setValue(0)

    def showEvent(self, event: Any) -> None:
        """Override the showEvent to center the window after it is fully laid out."""
        super().showEvent(event)  # Call the base class implementation
        self.center()  # Center the window after it is shown

    def center(self) -> None:
        """Center the window exactly in the middle of the screen."""
        # Get the screen's geometry
        screen_geometry = QApplication.primaryScreen().geometry()

        # Calculate the center position
        x = (screen_geometry.width() - self.width()) // 2
        y = (screen_geometry.height() - self.height()) // 2

        # Move the window to the calculated position
        self.move(x, y)

    def _resolve_default_file(self, file_type: str, fallback: Path) -> Path:
        """Return saved default for file_type if it exists, otherwise the hardcoded fallback."""
        key = f"defaults/{file_type}_file"
        saved = SettingsDialog.read(key, "", type=str)
        if saved:
            p = Path(saved)
            if p.exists():
                return p
            # Saved path no longer exists — clear it so it won't be retried
            SettingsDialog.write(key, "")
        return fallback

    def set_default_files(self) -> None:
        """Set the default input files in the UI, preferring saved defaults over hardcoded ones."""
        self.default_device_list  = self._resolve_default_file("device",  DEFAULT_DEVICE_FILE)
        self.default_command_list = self._resolve_default_file("command", DEFAULT_COMMAND_FILE)
        SettingsDialog.write("defaults/device_file",  str(self.default_device_list))
        SettingsDialog.write("defaults/command_file", str(self.default_command_list))

        for path, input_widget in [
            (self.default_device_list,  self.device_input),
            (self.default_command_list, self.command_input),
        ]:
            if path.exists():
                try:
                    relative_path = path.relative_to(Path(__file__).parent)
                    input_widget.setText(f"./{relative_path.as_posix()}")
                except ValueError:
                    input_widget.setText(str(path))
                input_widget.setStyleSheet("")
            else:
                try:
                    rel = path.relative_to(Path(__file__).parent)
                    input_widget.setText(f"Default file not found: ./{rel.as_posix()}")
                except ValueError:
                    input_widget.setText(f"Default file not found: {path}")
                input_widget.setStyleSheet("color: red;")

    def browse_device_list(self) -> None:
        """Open a file dialog to browse for the Device(s) input file."""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select Device List File", str(self.input_folder), "Text Files (*.txt);;All Files (*)"
        )
        if file_path:
            file_path = Path(file_path)
            try:
                relative_path = file_path.relative_to(Path(__file__).parent)
                self.device_input.setText(f"./{relative_path.as_posix()}")
                self.device_input.setStyleSheet("")  # Reset to default color
            except ValueError:
                # If the file is not in a subpath of the script's directory, use the absolute path
                self.device_input.setText(str(file_path))
                self.device_input.setStyleSheet("")  # Reset to default color
            self.validate_fields()

    def browse_command_list(self) -> None:
        """Open a file dialog to browse for the Command(s) input file."""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select Command List File", str(self.input_folder), "Text Files (*.txt);;All Files (*)"
        )
        if file_path:
            file_path = Path(file_path)
            try:
                relative_path = file_path.relative_to(Path(__file__).parent)
                self.command_input.setText(f"./{relative_path.as_posix()}")
                self.command_input.setStyleSheet("")  # Reset to default color
            except ValueError:
                # If the file is not in a subpath of the script's directory, use the absolute path
                self.command_input.setText(str(file_path))
                self.command_input.setStyleSheet("")  # Reset to default color
            self.validate_fields()

    def toggle_enable_field(self) -> None:
        """Enable or disable the Enable field based on the checkbox state."""
        if self.enable_checkbox.isChecked():
            self.enable_input.setDisabled(False)
            self.enable_input.clear()
        else:
            self.enable_input.setDisabled(True)
            self.sync_enable_password()
        self.validate_fields()

    def sync_enable_password(self) -> None:
        """Sync the Enable password with the Manual or Keyring password."""
        if not self.enable_checkbox.isChecked():
            if self.manual_radio.isChecked():
                self.enable_input.setText(self.password_input.text())
            else:
                self.enable_input.setText("")  # Clear if Keyring is selected

    def toggle_credential_options(self) -> None:
        """Enable/disable credential sections based on the selected radio button."""
        if self.manual_radio.isChecked():
            self.manual_credentials_group.setEnabled(True)
            self.keyring_credentials_group.setEnabled(False)
            self.remember_username_cb.setEnabled(True)
            self.remember_keyring_system_cb.setEnabled(False)
            self.reset_credentials()  # Reset all credentials when switching to Manual
            self.sync_enable_password()  # Sync Enable password when switching to Manual
        else:
            self.manual_credentials_group.setEnabled(False)
            self.keyring_credentials_group.setEnabled(True)
            self.remember_username_cb.setEnabled(False)
            self.remember_keyring_system_cb.setEnabled(True)
            self.reset_credentials()  # Reset all credentials when switching to Keyring
            self.sync_enable_password()  # Sync Enable password when switching to Keyring
        self.validate_fields()

    def reset_credentials(self) -> None:
        """Reset all credential fields to their default values."""
        self.username_input.clear()
        self.password_input.clear()
        self.enable_input.clear()
        self.keyring_system_input.clear()
        self.keyring_user_input.clear()

    def disable_input_controls(self) -> None:
        """Disable all input controls while the script is running."""
        self.device_button.setEnabled(False)
        self.device_edit_button.setEnabled(False)
        self.command_button.setEnabled(False)
        self.command_edit_button.setEnabled(False)
        # self.go_button.setEnabled(False)  # Remove this line
        self.go_button.setStyleSheet("")  # Reset the button's style to default
        self.quit_button.setEnabled(False)
        self.manual_radio.setEnabled(False)
        self.remember_username_cb.setEnabled(False)
        self.keyring_radio.setEnabled(False)
        self.remember_keyring_system_cb.setEnabled(False)
        self.username_input.setEnabled(False)
        self.password_input.setEnabled(False)
        
        # Store current enable state before disabling
        self.enable_was_enabled = self.enable_checkbox.isEnabled()
        self.enable_checkbox.setEnabled(False)
        self.enable_input.setEnabled(False)
        
        # Store and disable Options checkboxes
        self.verbose_was_enabled = self.verbose_checkbox.isEnabled()
        self.structured_output_was_enabled = self.structured_output_checkbox.isEnabled()
        self.verbose_checkbox.setEnabled(False)
        self.structured_output_checkbox.setEnabled(False)
        self.force_redetect_checkbox.setEnabled(False)
        self.settings_button.setEnabled(False)
        self.manage_clients_button.setEnabled(False)

        self.keyring_system_input.setEnabled(False)
        self.keyring_user_input.setEnabled(False)
        self.client_input.setEnabled(False)
        self.remember_client_cb.setEnabled(False)

    def _sanitize_client_input(self) -> None:
        """Sanitize the Client/Job name on the fly; notify the user if it was changed."""
        raw = self.client_input.currentText()
        sanitized = sanitize_folder_name(raw) if raw.strip() else ""
        if sanitized != raw:
            self.client_input.lineEdit().setText(sanitized)
            self.client_input.lineEdit().setToolTip(
                f'Invalid characters were removed. Folder name: "{sanitized}"'
            )
        else:
            self.client_input.lineEdit().setToolTip("")

    def _on_excel_report_toggled(self, checked: bool) -> None:
        """Show/hide force re-detect when Excel Report checkbox is toggled."""
        self.force_redetect_checkbox.setVisible(checked)

    def _get_selected_components(self) -> Set[str]:
        """Return the set of component names saved in Settings."""
        return SettingsDialog.read("excel/components", list(COMPONENT_NAMES), type=list)  # type: ignore[arg-type]

    def show_settings(self) -> None:
        """Open the Settings dialog. Current field values are preserved on close."""
        SettingsDialog(self).exec()

    def show_client_manager(self) -> None:
        """Open the Client / Job Manager dialog."""
        dlg = ClientManagerDialog(self)
        dlg.exec()
        # Refresh the client dropdown in case folders were added/renamed/deleted
        current = self.client_input.currentText()
        self.client_input.blockSignals(True)
        self.client_input.clear()
        self.client_input.addItem("")
        for client in get_existing_clients():
            self.client_input.addItem(client)
        idx = self.client_input.findText(current)
        self.client_input.setCurrentIndex(idx if idx >= 0 else 0)
        self.client_input.blockSignals(False)

    def _apply_remembered_settings(self) -> None:
        """Restore remembered field values from Settings on launch."""
        s = SettingsDialog
        if s.read("remember/username", False, type=bool):
            saved_user = s.read("remembered/username", "", type=str)
            if saved_user:
                self.manual_radio.blockSignals(True)
                self.manual_radio.setChecked(True)
                self.manual_radio.blockSignals(False)
                self.toggle_credential_options()
                self.username_input.setText(saved_user)
        if s.read("remember/keyring_sys", False, type=bool):
            saved_sys = s.read("remembered/keyring_system", "", type=str)
            if saved_sys:
                self.keyring_radio.blockSignals(True)
                self.keyring_radio.setChecked(True)
                self.keyring_radio.blockSignals(False)
                self.toggle_credential_options()
                self.keyring_system_input.setText(saved_sys)
        if s.read("remember/client", False, type=bool):
            saved_client = s.read("remembered/client", "", type=str)
            if saved_client:
                idx = self.client_input.findText(saved_client)
                if idx >= 0:
                    self.client_input.setCurrentIndex(idx)
                else:
                    self.client_input.lineEdit().setText(saved_client)

    def _save_remembered_values(self) -> None:
        """Persist remembered field values after a successful Go. Also saves the preference flags."""
        s = SettingsDialog
        # Save preference flags from the inline checkboxes
        s.write("remember/username",    self.remember_username_cb.isChecked())
        s.write("remember/keyring_sys", self.remember_keyring_system_cb.isChecked())
        s.write("remember/client",      self.remember_client_cb.isChecked())
        # Save actual values
        if self.remember_username_cb.isChecked():
            s.write("remembered/username", self.username_input.text().strip())
        if self.remember_keyring_system_cb.isChecked():
            s.write("remembered/keyring_system", self.keyring_system_input.text().strip())
        if self.remember_client_cb.isChecked():
            s.write("remembered/client", self.client_input.currentText().strip())

    def enable_input_controls(self) -> None:
        """Enable all input controls after the script completes."""
        self.device_button.setEnabled(True)
        self.device_edit_button.setEnabled(True)
        self.command_button.setEnabled(True)
        self.command_edit_button.setEnabled(True)
        self.go_button.setEnabled(True)
        self.quit_button.setEnabled(True)
        self.manual_radio.setEnabled(True)
        self.remember_username_cb.setEnabled(True)
        self.keyring_radio.setEnabled(True)
        self.remember_keyring_system_cb.setEnabled(True)
        self.username_input.setEnabled(True)
        self.password_input.setEnabled(True)
        
        # Restore enable checkbox and input based on previous state
        if self.enable_was_enabled:
            self.enable_checkbox.setEnabled(True)
            if self.enable_checkbox.isChecked():
                self.enable_input.setEnabled(True)
        
        # Restore Options checkboxes
        if self.verbose_was_enabled:
            self.verbose_checkbox.setEnabled(True)
        if self.structured_output_was_enabled:
            self.structured_output_checkbox.setEnabled(True)
        self.force_redetect_checkbox.setEnabled(True)
        self.settings_button.setEnabled(True)
        self.manage_clients_button.setEnabled(True)

        self.keyring_system_input.setEnabled(True)
        self.keyring_user_input.setEnabled(True)
        self.client_input.setEnabled(True)
        self.remember_client_cb.setEnabled(True)
        self.update_go_button_style(is_stop=False)
        self.validate_fields()  # Revalidate fields

    def show_output_folder(self) -> None:
        """Open the output folder using the OS's native file explorer."""
        client_name = self.client_input.currentText().strip()
        if client_name:
            client_folder = DEFAULT_OUTPUT_FOLDER / sanitize_folder_name(client_name)
            output_folder = client_folder if client_folder.exists() else DEFAULT_OUTPUT_FOLDER
        else:
            output_folder = DEFAULT_OUTPUT_FOLDER

        if not output_folder.exists():
            # Log an error message in the Output section
            self.append_colored_message("Output folder does not exist yet.", "ERROR")
            return

        if platform.system() == "Windows":
            os.startfile(output_folder)
        elif platform.system() == "Darwin":  # macOS
            subprocess.run(["open", output_folder])
        elif platform.system() == "Linux":
            subprocess.run(["xdg-open", output_folder])
        else:
            # Log an error message in the Output section
            self.append_colored_message("Unsupported operating system.", "ERROR")

    def validate_fields(self) -> None:
        """Validate all required fields and enable/disable the Go button."""
        device_file = self.device_input.text().strip()
        command_file = self.command_input.text().strip()

        if self.manual_radio.isChecked():
            username = self.username_input.text().strip()
            password = self.password_input.text().strip()
            enable = self.enable_input.text().strip()
            is_valid = bool(
                device_file
                and command_file
                and username
                and password
                and (not self.enable_checkbox.isChecked() or enable)
            )
        else:
            keyring_system = self.keyring_system_input.text().strip()
            keyring_user = self.keyring_user_input.text().strip()
            is_valid = bool(
                device_file
                and command_file
                and keyring_system
                and keyring_user
            )

        # Enable/disable the Go button and change its color
        self.go_button.setEnabled(is_valid)
        if is_valid:
            self.go_button.setStyleSheet("")  # Reset to default style
            self.go_button.setStyleSheet(
                """
                QPushButton {
                    background-color: #5cb85c;
                    color: white;
                    border-radius: 5px;
                    padding: 1px;
                    border: 1px solid #4cae4c;
                }
                QPushButton:hover {
                    background-color: #449d44;
                }
                QPushButton:pressed {
                    background-color: #398439;
                }
                """
            )
        else:
            self.go_button.setStyleSheet("")  # Reset to default style

    def update_go_button_style(self, is_stop: bool = False) -> None:
        """
        Update the Go/Stop button appearance.
        
        Args:
            is_stop: Whether to show the button in "Stop" mode
        """
        if is_stop:
            self.go_button.setText("Stop")
            self.go_button.setStyleSheet("""
                QPushButton {
                    background-color: #d9534f;
                    color: white;
                    border-radius: 5px;
                    padding: 1px;
                    border: 1px solid #d43f3a;
                }
                QPushButton:hover {
                    background-color: #c9302c;
                }
                QPushButton:pressed {
                    background-color: #ac2925;
                }
            """)
        else:
            self.go_button.setText("Go")
            if self.go_button.isEnabled():
                self.go_button.setStyleSheet("""
                    QPushButton {
                        background-color: #5cb85c;
                        color: white;
                        border-radius: 5px;
                        padding: 1px;
                        border: 1px solid #4cae4c;
                    }
                    QPushButton:hover {
                        background-color: #449d44;
                    }
                    QPushButton:pressed {
                        background-color: #398439;
                    }
                """)
            else:
                self.go_button.setStyleSheet("")  # Default disabled style

    def show_help(self) -> None:
        """Create and show the custom help dialog."""
        help_dialog = HelpDialog(APP_NAME, APP_VERSION, VERSION_DATE, self)
        help_dialog.exec()

    def on_go(self) -> None:
        """Handle Go/Stop button click."""
        # Clear the log output when starting a new run
        if not (hasattr(self, 'worker') and self.worker and self.worker.isRunning()):
            self.log_output.clear()
            # self.log_output.setText("")
        
        if hasattr(self, 'worker') and self.worker and self.worker.isRunning():
            # If worker is running, this click means stop
            self.stop_script()
        else:
            # Normal execution
            self.start_script()

    def start_script(self) -> None:
        """Start script execution."""
        # Validate component selection before doing anything else
        if self.structured_output_checkbox.isChecked() and not self._get_selected_components():
            self.append_colored_message(
                "No report components selected. Please select at least one component or uncheck Excel Report.",
                "ERROR"
            )
            return

        # Clear previous worker if exists
        if hasattr(self, 'worker'):
            self.worker.quit()
            self.worker.wait(100)
            del self.worker

        # Reset state
        self.stop_requested = False
        self.progress_bar.setValue(0)
        self.progress_bar.setStyleSheet("")  # Reset to default color

        # Disable controls and update button, Update UI immediately
        self.disable_input_controls()
        self.update_go_button_style(is_stop=True)
        QApplication.processEvents()  # Force UI update

        # Create new output folder and log file for this run
        client_name = self.client_input.currentText().strip()
        timestamp = datetime.now().strftime("%m-%d-%Y - %I_%M_%p")
        if client_name:
            output_folder = DEFAULT_OUTPUT_FOLDER / sanitize_folder_name(client_name) / timestamp
        else:
            output_folder = DEFAULT_OUTPUT_FOLDER / timestamp
        output_folder.mkdir(parents=True, exist_ok=True)
        log_file = str(output_folder / Path(__file__).stem) + ".log"

        # Configure logging for this run
        self.configure_logging(log_file)

        # Get input values
        device_file = self.device_input.text().strip()
        command_file = self.command_input.text().strip()

        # Write run metadata so the Client Manager can display it without scanning output files
        try:
            run_info = {
                "timestamp": timestamp,
                "device_file": device_file,
                "command_file": command_file,
            }
            with open(output_folder / "run_info.json", "w", encoding="utf-8") as _f:
                json.dump(run_info, _f, indent=2)
        except OSError:
            pass

        if self.manual_radio.isChecked():
            credentials = {
                'username': self.username_input.text().strip(),
                'password': self.password_input.text().strip()
            }
            enable_password = self.enable_input.text().strip()
        else:
            keyring_system = self.keyring_system_input.text().strip()
            keyring_user = self.keyring_user_input.text().strip()
            try:
                password = keyring.get_password(keyring_system, keyring_user)
                if password:
                    credentials = {
                        'username': keyring_user,
                        'password': password
                    }
                    enable_password = password
                else:
                    self.append_colored_message("No credentials found in the keyring.", "ERROR")
                    self.enable_input_controls()  # Re-enable controls on error
                    return
            except Exception as e:
                self.append_colored_message(f"Failed to fetch credentials from the keyring: {e}", "ERROR")
                self.enable_input_controls()  # Re-enable controls on error
                return

        # Persist remembered field values before starting
        self._save_remembered_values()

        # Create and start worker
        self.worker = Worker(
            device_file=device_file,
            command_file=command_file,
            credentials=credentials,
            enable_password=enable_password,
            output_folder=output_folder,
            verbose_enabled=self.verbose_checkbox.isChecked(),
            create_combined_output=SettingsDialog.read("output/combined", False, type=bool),
            enable_structured_output=self.structured_output_checkbox.isChecked(),
            selected_components=set(self._get_selected_components()) if self.structured_output_checkbox.isChecked() else None,
            force_redetect=self.force_redetect_checkbox.isChecked(),
            save_txt=SettingsDialog.read("output/save_txt", False, type=bool)
        )
        self.worker.progress_signal.connect(self.update_progress)
        self.worker.log_signal.connect(self.update_log)
        self.worker.completion_signal.connect(self.on_script_complete)
        self.worker.start()

    def update_progress(self, value: int) -> None:
        """Update the progress bar with the current value."""
        self.progress_bar.setValue(value)

    def append_colored_message(self, message: str, level_name: str = "INFO") -> None:
        """
        Append a colored message to the log output.
        
        Args:
            message: The message text to append
            level_name: Log level name (e.g., "INFO", "ERROR")
        """
        # Only use this for non-logger GUI messages
        cursor = self.log_output.textCursor()
        cursor.movePosition(QTextCursor.MoveOperation.End)
        
        color = self.get_color_for_level(level_name)
        formatted = self.format_message(message, level_name)
        
        cursor.insertHtml(f'<span style="color:{color}">{formatted}</span><br>')
        self.log_output.setTextCursor(cursor)
        self.log_output.ensureCursorVisible()

    def get_color_for_level(self, level_name: str) -> str:
        """
        Get the color associated with a log level.
        
        Args:
            level_name: Log level name
            
        Returns:
            Color string for the log level
        """
        color_map = {
            "DEBUG": "blue",
            "VERBOSE": "green", 
            "INFO": "",
            "WARNING": "orange",
            "ERROR": "red",
            "CRITICAL": "darkred"
        }
        return color_map.get(level_name, "")

    def format_message(self, message: str, level_name: str) -> str:
        """
        Format a message for display with proper HTML escaping.
        
        Args:
            message: The message text
            level_name: Log level name
            
        Returns:
            Formatted HTML string
        """
        # Convert message to HTML with proper line breaks and spacing
        message = (message
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace("\n", "<br>")
            .replace(" ", "&nbsp;")
            .replace("\t", "&nbsp;&nbsp;&nbsp;&nbsp;"))
        
        if level_name != "INFO":
            return f"{level_name}: {message}"
        return message

    def update_log(self, message: str) -> None:
        """Append a colored message to the log output."""
        self.append_colored_message(message)
        self.log_output.ensureCursorVisible()  # Scroll to the bottom
        self.keep_horizontal_scroll_left()  # Reset horizontal scrollbar to the left

    def edit_device_file(self) -> None:
        """Open the device file for editing."""
        file_path = self.device_input.text().strip()
        if file_path:
            file_path = Path(file_path)
            if file_path.exists() and file_path.is_file():
                self.show_file_editor(file_path, "device")  # Pass "device" as the file type
            else:
                # Log an error message in the Output section
                self.append_colored_message("The selected file does not exist or is not a valid file.", "ERROR")
        else:
            # Log an error message in the Output section
            self.append_colored_message("No file selected.", "ERROR")

    def edit_command_file(self) -> None:
        """Open the command file for editing."""
        file_path = self.command_input.text().strip()
        if file_path:
            file_path = Path(file_path)
            if file_path.exists() and file_path.is_file():
                self.show_file_editor(file_path, "command")  # Pass "command" as the file type
            else:
                # Log an error message in the Output section
                self.append_colored_message("The selected file does not exist or is not a valid file.", "ERROR")
        else:
            # Log an error message in the Output section
            self.append_colored_message("No file selected.", "ERROR")

    def show_file_editor(self, file_path: Path, file_type: str) -> None:
        """Create and show the file editor dialog."""
        # Create and show the file editor dialog
        self.file_editor_dialog = FileEditorDialog(file_path, file_type, self)
        # Connect the file_path_updated signal to update the input file path
        self.file_editor_dialog.file_path_updated.connect(self.update_input_file)
        self.file_editor_dialog.exec()

    def update_input_file(self, new_file_path: Path, file_type: str) -> None:
        """
        Update the input file path in the main window based on the file type.
        
        Args:
            new_file_path: New path to the file
            file_type: Type of file ('device' or 'command')
        """
        try:
            relative_path = new_file_path.relative_to(Path(__file__).parent)
            display_path = f"./{relative_path.as_posix()}"
        except ValueError:
            display_path = str(new_file_path)
        
        if file_type == "device":
            self.device_input.setText(display_path)
            self.device_input.setStyleSheet("")  # Reset to default color
        elif file_type == "command":
            self.command_input.setText(display_path)
            self.command_input.setStyleSheet("")  # Reset to default color
        self.validate_fields()  # Revalidate fields in case the file path changed

    def open_keyring_tools(self) -> None:
        """Open the Keyring Tools utility in a separate process."""
        # Define the path to the Keyring Tools script
        keyring_tools_path = SCRIPT_DIR / KEYRING_TOOLS_SCRIPT

        # Check if the Keyring Tools script exists
        if not keyring_tools_path.exists():
            # If the script is not found, log an error message in the Output section
            self.append_colored_message(f"'{KEYRING_TOOLS_SCRIPT}' not found.\n- Please ensure it is in the same folder as this script.", "ERROR")
            return

        try:
            # Import the KeyringApp from the external script
            keyring_tools_path = SCRIPT_DIR / KEYRING_TOOLS_SCRIPT
            if platform.system() == "Windows":
                subprocess.Popen(
                    [sys.executable, str(keyring_tools_path)],
                    creationflags=subprocess.CREATE_NO_WINDOW,
                )
            else:
                subprocess.Popen([sys.executable, str(keyring_tools_path)])
        except Exception as e:
            # If there's an error during import or execution, log it in the Output section
            self.append_colored_message(f"ERROR: Failed to open Keyring Tools - {e}", "ERROR")

    def stop_script(self) -> None:
        """Stop script execution."""
        self.stop_requested = True
        self.logger.warning("Stopping script execution...")  # Will appear once in both places
        self.go_button.setEnabled(False)  # Disable while stopping
        
        if hasattr(self, 'worker'):
            self.worker.cancel()  # Signal the worker to stop

    def on_script_complete(self, output_folder: str, success: bool) -> None:
        """
        Handle script completion.

        Args:
            output_folder: Path to the output folder
            success: True if all devices completed without errors
        """
        if self.stop_requested:
            self.logger.warning("Script execution was cancelled - collected output may be incomplete!")

        # Color the progress bar based on outcome
        if self.stop_requested:
            color = "#f0ad4e"  # amber for cancelled
        elif success:
            color = "#5cb85c"  # green for success
        else:
            color = "#d9534f"  # red for failure
        self.progress_bar.setStyleSheet(
            f"QProgressBar {{ text-align: center; }}"
            f"QProgressBar::chunk {{ background-color: {color}; }}")

        # Rest of your existing completion handling
        self.enable_input_controls()  # Re-enable all input controls
        self.update_go_button_style(is_stop=False)
        
# ─────────────────────────────────────────────────────────────────────────────
# Excel Report Generation
# ─────────────────────────────────────────────────────────────────────────────

def _get(data: Any, *keys: Any, default: Any = 'N/A') -> Any:
    """Safely navigate a nested dict/list, returning default on any miss."""
    for key in keys:
        try:
            data = data[key]
        except (KeyError, IndexError, TypeError):
            return default
    return data if data is not None else default


# Lowercase-prefix → canonical full name table (ordered longest prefix first to prevent
# shorter prefixes like 'eth' matching before 'ethernet').
_INTF_ABBREV: List[tuple] = [
    ('tengigabitethernet',   'TenGigabitEthernet'),
    ('hundredgige',          'HundredGigE'),
    ('fortygigabitethernet', 'FortyGigabitEthernet'),
    ('gigabitethernet',      'GigabitEthernet'),
    ('fastethernet',         'FastEthernet'),
    ('ethernet',             'Ethernet'),
    ('loopback',             'Loopback'),
    ('port-channel',         'Port-channel'),
    ('tunnel',               'Tunnel'),
    ('vlan',                 'Vlan'),
    ('mgmt',                 'mgmt'),
    # Abbreviated forms — after full-name entries so full names match first
    ('tengig',               'TenGigabitEthernet'),
    ('hundgig',              'HundredGigE'),
    ('fortygig',             'FortyGigabitEthernet'),
    ('eth',                  'Ethernet'),
    ('gi',                   'GigabitEthernet'),
    ('fa',                   'FastEthernet'),
    ('te',                   'TenGigabitEthernet'),
    ('hu',                   'HundredGigE'),
    ('fo',                   'FortyGigabitEthernet'),
    ('lo',                   'Loopback'),
    ('po',                   'Port-channel'),
    ('tu',                   'Tunnel'),
    ('vl',                   'Vlan'),
]

# NX-OS interface type keyword values that appear in the 'type' column of
# 'show interface description' — distinguish real type/speed columns from
# Genie misidentifying description words as type/speed.
_NXOS_INTF_TYPE_KEYWORDS = {
    'eth', 'fc', 'mgmt', 'vlan', 'lag', 'svi', 'loopback',
    'tunnel', 'pc', 'nve', 'sup-eth',
}

def _normalize_intf_name(name: str) -> str:
    """Expand abbreviated or mixed-case Cisco interface names to canonical full form.

    Case-insensitive: 'loopback0', 'Lo0', 'Loopback0' all → 'Loopback0'.
    'Eth1/11' → 'Ethernet1/11', 'Lo0' → 'Loopback0', 'Po2' → 'Port-channel2', etc.
    """
    name_lower = name.lower()
    for abbrev, full in _INTF_ABBREV:
        if name_lower.startswith(abbrev):
            rest = name_lower[len(abbrev):]
            # Only treat as abbreviation/full-name if the remainder is a digit, '/', or end
            if not rest or rest[0].isdigit() or rest[0] == '/':
                return full + rest
    return name


def _load_port_map(csv_path: Path) -> Dict[str, str]:
    """Load a port map CSV mapping old (Run A) interface names to new (Run B) names.

    CSV format: old_interface,new_interface
    Lines starting with '#' are comments; blank lines are ignored.
    The first non-comment row may be a header ('old_interface,new_interface') and is skipped.
    Both columns are normalized via _normalize_intf_name().
    """
    port_map: Dict[str, str] = {}
    with open(csv_path, newline='', encoding='utf-8') as f:
        first_data_row = True
        for line in f:
            stripped = line.strip()
            if not stripped or stripped.startswith('#'):
                continue
            parts = stripped.split(',', 1)
            if len(parts) != 2:
                continue
            old_raw, new_raw = parts[0].strip(), parts[1].strip()
            if not old_raw or not new_raw:
                continue
            if first_data_row:
                first_data_row = False
                if old_raw.lower() == 'old_interface':
                    continue
            port_map[_normalize_intf_name(old_raw)] = _normalize_intf_name(new_raw)
    return port_map


def _remap_intf_keys(
    intf_map: Dict[str, Dict], port_map: Optional[Dict[str, str]],
) -> Dict[str, Dict]:
    """Return a new dict with interface keys remapped per port_map.

    Keys present in port_map are translated to the new name.
    Keys not in port_map are kept as-is.
    """
    if not port_map:
        return intf_map
    return {port_map.get(name, name): data for name, data in intf_map.items()}


def _remap_intf_string(intf_str: str, port_map: Optional[Dict[str, str]]) -> str:
    """Remap interface names within a comma-separated string.

    Each name is normalized, looked up in port_map, and replaced.
    Returns the remapped string, sorted for stable comparison.
    """
    if not port_map or not intf_str:
        return intf_str
    parts = [p.strip() for p in intf_str.split(',') if p.strip()]
    remapped = [port_map.get(_normalize_intf_name(p), _normalize_intf_name(p)) for p in parts]
    return ', '.join(sorted(remapped))


def _normalize_intf_string(intf_str: str) -> str:
    """Normalize and sort interface names within a comma-separated string (no remapping)."""
    if not intf_str:
        return intf_str
    parts = [_normalize_intf_name(p.strip()) for p in intf_str.split(',') if p.strip()]
    return ', '.join(sorted(parts))


# Column header names that always receive wrap_text in any sheet they appear in.
_WRAP_TEXT_COLS: frozenset = frozenset({
    'Root Bridge For', 'Interfaces',
    'Root Bridge For (A)', 'Root Bridge For (B)',
    'Interfaces (A)', 'Interfaces (B)',
})


def _apply_header_style(ws: Any) -> None:
    """Excel Table with alternating row stripes, freeze top row, auto-fit columns."""
    from openpyxl.styles import Font, Alignment                      # type: ignore
    from openpyxl.utils import get_column_letter                     # type: ignore
    from openpyxl.worksheet.table import Table, TableStyleInfo       # type: ignore

    # Auto-fit columns before adding the table
    ws.freeze_panes = "A2"
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 60)

    # Bold + centered header text (table style handles the fill colour)
    for cell in ws[1]:
        cell.font      = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Determine which column indices need wrap text (matched by header name)
    wrap_col_idxs = {
        cell.column for cell in ws[1]
        if cell.value in _WRAP_TEXT_COLS
    }

    # Middle-align all data rows vertically; apply wrap text to designated columns.
    # Preserve any existing horizontal alignment and wrap_text already set on the cell
    # (e.g. Summary sheet PID/Serial cells for stacked switches).
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            a = cell.alignment
            needs_wrap = bool(a.wrap_text) or cell.column in wrap_col_idxs
            cell.alignment = Alignment(
                horizontal=a.horizontal,
                vertical='center',
                wrap_text=True if needs_wrap else None,
            )

    # Build the table reference (e.g. A1:H42)
    max_col = ws.max_column
    max_row = ws.max_row
    if max_row < 2:
        return  # no data rows — skip table creation
    end_col = get_column_letter(max_col)
    ref     = f"A1:{end_col}{max_row}"

    # Table names must be unique, contain no spaces, and only [A-Za-z0-9_]
    safe_title = re.sub(r'[^A-Za-z0-9_]', '_', ws.title)
    table_name = f"tbl_{safe_title}"
    style = TableStyleInfo(
        name            = "TableStyleMedium9",
        showFirstColumn = False,
        showLastColumn  = False,
        showRowStripes  = True,
        showColumnStripes = False,
    )
    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = style
    ws.add_table(tbl)


def _sheet_run_info(wb: Any, run_folder: Path, devices_data: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet("Run Info")
    ws.append(["Field", "Value"])
    meta = devices_data[0].get('meta', {}) if devices_data else {}
    ws.append(["Client",             run_folder.parent.name])
    ws.append(["Run Folder",         run_folder.name])
    ws.append(["nddu Version",       _get(meta, 'nddu_version')])
    ws.append(["Devices Documented", len(devices_data)])
    _apply_header_style(ws)


def _parse_inventory_members_raw(raw: str) -> Dict[str, Any]:
    """
    Raw parser for 'show inventory' — extracts per-stack-member chassis PID and SN.

    Scans for NAME: "Switch N" blocks (exact match, case-insensitive) and pulls
    the PID and SN from the following PID: line.  Returns an ordered dict:
        {'1': {'pid': 'C9300-48P', 'sn': 'FOX...'}, '2': {...}, ...}
    Empty dict for non-stacked devices (no "Switch N" entries present).
    """
    members: Dict[str, Any] = {}
    current_num: Optional[str] = None
    for line in raw.splitlines():
        stripped = line.strip()
        m = re.match(r'NAME:\s+"Switch\s+(\d+)"\s*,', stripped, re.IGNORECASE)
        if m:
            current_num = m.group(1)
            continue
        if current_num and stripped.upper().startswith('PID:'):
            pid_m = re.search(r'PID:\s*(\S+)',  stripped, re.IGNORECASE)
            sn_m  = re.search(r'SN:\s*(\S+)',   stripped, re.IGNORECASE)
            pid = pid_m.group(1).rstrip(',') if pid_m else ''
            sn  = sn_m.group(1).rstrip(',')  if sn_m  else ''
            members[current_num] = {'pid': pid, 'sn': sn}
            current_num = None
    return members


def _extract_inventory_pid(inv: Any) -> str:
    """
    Extract the chassis PID from 'show inventory' structured output.

    TextFSM format: List[Dict] with keys: name, descr, pid, vid, sn.
    Wrapped format: {_entries: [...], _stack_members: {...}} when stack data added.
    Prefers entries whose name contains 'Chassis'; falls back to first entry.
    """
    # Unwrap if enriched with _stack_members
    entries = inv
    if isinstance(inv, dict) and '_entries' in inv:
        entries = inv['_entries']

    if isinstance(entries, list) and entries:
        # TextFSM flat list — prefer Chassis entry
        for entry in entries:
            if isinstance(entry, dict) and 'chassis' in entry.get('name', '').lower():
                pid = entry.get('pid', '').strip()
                if pid:
                    return pid
        # Fallback: first entry with a PID
        for entry in entries:
            if isinstance(entry, dict):
                pid = entry.get('pid', '').strip()
                if pid:
                    return pid
    return 'N/A'


def _sheet_summary(wb: Any, devices_data: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet("Summary")
    ws.append(["Hostname", "IP", "Device Type", "Platform / Model",
               "Version", "PID", "Serial Number", "Uptime", "Stack Members"])
    for dev in devices_data:
        meta       = dev.get('meta', {})
        structured = dev.get('structured', {})
        hostname    = meta.get('hostname', 'N/A')
        ip          = meta.get('host',     'N/A')
        device_type = meta.get('device_type', 'N/A')

        # TextFSM show version → List[Dict] with keys:
        #   IOS/IOS-XE: hardware (list), serial (list), version, uptime, hostname, ...
        #   NX-OS:      platform, serial, hostname, uptime, ...
        ver_list = structured.get('show version')
        ver = ver_list[0] if isinstance(ver_list, list) and ver_list else {}

        hw = ver.get('hardware', [])
        platform = hw[0] if isinstance(hw, list) and hw else ver.get('platform', 'N/A')
        sw_ver   = ver.get('version', 'N/A') or 'N/A'
        sn_list  = ver.get('serial', [])
        serial   = sn_list[0] if isinstance(sn_list, list) and sn_list else ver.get('serial', 'N/A')
        uptime   = ver.get('uptime', 'N/A') or 'N/A'

        # PID from show inventory (TextFSM flat list)
        pid = _extract_inventory_pid(structured.get('show inventory'))

        # Stack member PID/SN — populated by _parse_inventory_members_raw() during
        # collection; stored as '_stack_members' dict inside the structured entry.
        inv_data = structured.get('show inventory')
        raw_members = inv_data.get('_stack_members', {}) if isinstance(inv_data, dict) else {}
        stack_member_map: Dict[str, tuple] = {
            num: (d.get('pid', ''), d.get('sn', ''))
            for num, d in raw_members.items()
            if isinstance(d, dict) and d.get('pid')
        }

        # If stack member inventory data exists, replace PID/Serial with per-member
        # newline-separated lists so each switch is visible in order.
        if stack_member_map:
            ordered = sorted(stack_member_map.items(), key=lambda x: int(x[0]))
            pid    = '\n'.join(p for _, (p, _s) in ordered)
            serial = '\n'.join(_s for _, (_p, _s) in ordered)

        # show switch detail — TextFSM: List[Dict] with switch, role, ...
        switch_list = structured.get('show switch detail')
        if isinstance(switch_list, list) and switch_list:
            members = ", ".join(
                f"{entry.get('switch', '?')}:{entry.get('role', '?')}"
                for entry in sorted(switch_list, key=lambda x: x.get('switch', ''))
            )
        else:
            members = 'N/A'

        ws.append([hostname, ip, device_type, platform, sw_ver, pid, serial, uptime, members])

        # Apply wrap text to PID and Serial Number cells when they contain stacked values
        if stack_member_map:
            from openpyxl.styles import Alignment  # type: ignore
            row = ws.max_row
            for col in (6, 7):  # PID = col 6, Serial Number = col 7
                ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')
    _apply_header_style(ws)


def _sheet_interfaces(wb: Any, devices_data: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet("Interfaces")
    ws.append(["Hostname", "IP", "Interface", "Description", "IP Address",
               "Status", "Protocol", "Vlan", "Duplex", "Speed", "Type"])
    for dev in devices_data:
        meta        = dev.get('meta', {})
        structured  = dev.get('structured', {})
        hostname    = meta.get('hostname',    'N/A')
        dev_ip      = meta.get('host',        'N/A')

        # TextFSM show interface status → List[Dict]:
        #   port, name, status, vlan_id, duplex, speed, type
        status_list = structured.get('show interface status', [])
        status_map: Dict[str, Dict] = {}
        if isinstance(status_list, list):
            for entry in status_list:
                if isinstance(entry, dict):
                    status_map[_normalize_intf_name(entry.get('port', ''))] = entry

        # TextFSM show ip interface brief → List[Dict]:
        #   interface, ip_address, status, proto
        ip_list = structured.get('show ip interface brief', [])
        ip_map: Dict[str, Dict] = {}
        if isinstance(ip_list, list):
            for entry in ip_list:
                if isinstance(entry, dict):
                    ip_map[_normalize_intf_name(entry.get('interface', ''))] = entry

        # TextFSM show interfaces description → List[Dict]:
        #   port, status, protocol, description
        desc_key = 'show interfaces description'
        desc_list = structured.get(desc_key) or structured.get('show interface description', [])
        desc_map: Dict[str, Dict] = {}
        if isinstance(desc_list, list):
            for entry in desc_list:
                if isinstance(entry, dict):
                    desc_map[_normalize_intf_name(entry.get('port', ''))] = entry

        # Build merged interface set from all sources
        all_intfs: set = set()
        all_intfs.update(status_map.keys())
        all_intfs.update(ip_map.keys())
        all_intfs.update(desc_map.keys())

        if not all_intfs:
            ws.append([hostname, dev_ip] + ['N/A'] * 9)
            continue

        for intf in sorted(all_intfs):
            st = status_map.get(intf, {})
            ip_info = ip_map.get(intf, {})
            desc_entry = desc_map.get(intf, {})
            # Prefer show interfaces description (full text) over
            # show interface status (truncated to ~25 chars)
            desc = (desc_entry.get('description') or st.get('name') or '')
            ip_addr = ip_info.get('ip_address', 'N/A') or 'N/A'
            # Status: prefer show interface status, fall back to ip brief, then desc
            status = (st.get('status') or ip_info.get('status')
                      or desc_entry.get('status', 'N/A') or 'N/A')
            # Protocol: prefer ip brief, fall back to desc
            protocol = (ip_info.get('proto') or ip_info.get('protocol')
                        or desc_entry.get('protocol', 'N/A') or 'N/A')
            ws.append([hostname, dev_ip, intf, desc, ip_addr, status, protocol,
                        st.get('vlan_id', 'N/A'),
                        st.get('duplex', 'N/A'),
                        st.get('speed', 'N/A'),
                        st.get('type', 'N/A')])
    _apply_header_style(ws)


def _sheet_neighbors(wb: Any, devices_data: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet("Neighbors")
    ws.append(["Hostname", "IP", "Local Interface", "Neighbor",
               "Neighbor IP", "Platform", "Neighbor Port"])
    for dev in devices_data:
        meta       = dev.get('meta', {})
        structured = dev.get('structured', {})
        hostname   = meta.get('hostname', 'N/A')
        ip         = meta.get('host',     'N/A')
        # TextFSM show cdp neighbors detail → List[Dict]:
        #   IOS: neighbor_name, mgmt_address, platform, neighbor_interface, local_interface, ...
        #   NX-OS: dest_host, mgmt_ip, platform, remote_port, local_port, ...
        cdp_list = structured.get('show cdp neighbors detail')
        if not isinstance(cdp_list, list) or not cdp_list:
            continue
        for nbr in cdp_list:
            if not isinstance(nbr, dict):
                continue
            ws.append([hostname, ip,
                        nbr.get('local_interface') or nbr.get('local_port', 'N/A'),
                        nbr.get('neighbor_name') or nbr.get('dest_host', 'N/A'),
                        nbr.get('mgmt_address') or nbr.get('mgmt_ip', 'N/A'),
                        nbr.get('platform', 'N/A'),
                        nbr.get('neighbor_interface') or nbr.get('remote_port', 'N/A')])
    _apply_header_style(ws)


def _sheet_vlans(wb: Any, devices_data: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet("VLANs")
    ws.append(["Hostname", "IP", "VLAN ID", "Name", "State", "Port Count"])
    for dev in devices_data:
        meta       = dev.get('meta', {})
        structured = dev.get('structured', {})
        hostname   = meta.get('hostname', 'N/A')
        ip         = meta.get('host',     'N/A')
        # IOS-XE: show vlan brief → vlan → {vlan1: {vlan_name, vlan_status, vlan_port}}
        # NX-OS:  show vlan       → vlans → {1: {name, state, interfaces}}
        iosxe_vlans = _get(structured, 'show vlan brief', 'vlan',  default=None)
        nxos_vlans  = _get(structured, 'show vlan',       'vlans', default=None)

        if isinstance(iosxe_vlans, dict):
            def _vlan_sort_key(k: str) -> int:
                try:    return int(k[4:] if k.startswith('vlan') else k)
                except: return 9999
            for key, vdata in sorted(iosxe_vlans.items(), key=lambda x: _vlan_sort_key(x[0])):
                vlan_id    = key[4:] if key.startswith('vlan') else key
                ports      = _get(vdata, 'vlan_port', default=[])
                port_count = len(ports) if isinstance(ports, list) else 'N/A'
                ws.append([hostname, ip, vlan_id,
                            _get(vdata, 'vlan_name'),
                            _get(vdata, 'vlan_status'),
                            port_count])
        elif isinstance(nxos_vlans, dict):
            for vlan_id, vdata in sorted(nxos_vlans.items(),
                                         key=lambda x: int(x[0]) if str(x[0]).isdigit() else 9999):
                ports      = _get(vdata, 'interfaces', default=[])
                port_count = len(ports) if isinstance(ports, list) else 'N/A'
                ws.append([hostname, ip, vlan_id,
                            _get(vdata, 'name'),
                            _get(vdata, 'state'),
                            port_count])
        else:
            continue
    _apply_header_style(ws)


def _sheet_routing(wb: Any, devices_data: List[Dict[str, Any]]) -> None:
    """Routing sheet — reads from _parse_ip_route_summary_raw output format."""
    ws = wb.create_sheet("Routing")
    ws.append(["Hostname", "IP", "Protocol", "Instance", "Routes"])
    for dev in devices_data:
        meta       = dev.get('meta', {})
        structured = dev.get('structured', {})
        hostname   = meta.get('hostname', 'N/A')
        ip         = meta.get('host',     'N/A')
        route_data = structured.get('show ip route summary')
        if not isinstance(route_data, dict):
            continue
        for entry in route_data.get('protocols', []):
            ws.append([hostname, ip,
                        entry.get('protocol', 'N/A'),
                        entry.get('instance', '-'),
                        entry.get('routes', 0)])
        total = route_data.get('total', {})
        if total:
            total_routes = total.get('routes', 'N/A')
            total_paths = total.get('paths')
            total_str = f"{total_routes} routes / {total_paths} paths" \
                        if total_paths is not None else str(total_routes)
            ws.append([hostname, ip, 'TOTAL', '-', total_str])
    _apply_header_style(ws)


def _sheet_stp(wb: Any, devices_data: List[Dict[str, Any]]) -> None:
    """STP sheet — reads from _parse_stp_summary_raw output format."""
    ws = wb.create_sheet("STP")
    ws.append(["Hostname", "IP", "Mode", "Root Bridge For", "VLANs", "Blocking", "Forwarding", "STP Active"])
    for dev in devices_data:
        meta       = dev.get('meta', {})
        structured = dev.get('structured', {})
        hostname   = meta.get('hostname', 'N/A')
        ip         = meta.get('host',     'N/A')
        stp        = structured.get('show spanning-tree summary')
        if not isinstance(stp, dict):
            continue
        ws.append([hostname, ip,
                    stp.get('mode', 'N/A'),
                    stp.get('root_bridge_for', 'N/A'),
                    stp.get('num_vlans', 'N/A'),
                    stp.get('blocking', 'N/A'),
                    stp.get('forwarding', 'N/A'),
                    stp.get('stp_active', 'N/A')])
    _apply_header_style(ws)


def _sheet_mac(wb: Any, devices_data: List[Dict[str, Any]]) -> None:
    """
    MAC Address sheet — one row per MAC table entry across all devices.

    TextFSM show mac address-table → List[Dict]:
      IOS:  destination_address, type, vlan_id, destination_port (list)
      NX-OS: mac, type, vlan, ports (str), age, secure, ntfy

    TextFSM show arp / show ip arp → List[Dict]:
      address, hardware_address, interface, age_min, ...
    NX-OS show ip arp: raw-parsed to {mac: ip} dict directly
    """
    ws = wb.create_sheet("MAC Addresses")
    ws.append(["Hostname", "IP", "VLAN", "MAC Address", "Type", "Port", "IP Address"])

    for dev in devices_data:
        hostname   = dev.get('meta', {}).get('hostname', 'N/A') or 'N/A'
        ip         = dev.get('meta', {}).get('host', 'N/A')
        structured = dev.get('structured', {})

        # Build MAC → IP lookup from ARP table
        # TextFSM show arp: [{address, hardware_address, interface, ...}]
        arp_map: Dict[str, str] = {}
        arp_list = structured.get('show arp') or structured.get('show ip arp')
        if isinstance(arp_list, list):
            for entry in arp_list:
                if isinstance(entry, dict):
                    mac = entry.get('hardware_address') or entry.get('mac', '')
                    addr = entry.get('address', '')
                    if mac and addr:
                        arp_map[mac] = addr
        elif isinstance(arp_list, dict):
            # NX-OS raw-parsed {mac: ip} format
            arp_map.update(arp_list)

        # TextFSM show mac address-table → List[Dict]
        mac_list = structured.get('show mac address-table')
        if not isinstance(mac_list, list) or not mac_list:
            continue

        for entry in sorted(mac_list,
                            key=lambda x: int(x.get('vlan_id') or x.get('vlan') or '0')
                            if (x.get('vlan_id') or x.get('vlan') or '0').isdigit() else 0):
            mac_addr = entry.get('destination_address') or entry.get('mac', 'N/A')
            vlan_id = entry.get('vlan_id') or entry.get('vlan', 'N/A')
            mac_type = entry.get('type', 'N/A')
            # destination_port is a list on IOS; ports is a string on NX-OS
            port_val = entry.get('destination_port') or entry.get('ports', 'N/A')
            port = port_val[0] if isinstance(port_val, list) and port_val else (port_val or 'N/A')
            ws.append([
                hostname, ip, vlan_id, mac_addr,
                mac_type, port,
                arp_map.get(mac_addr, ''),
            ])

    _apply_header_style(ws)


def _sheet_mac_summary(wb: Any, devices_data: List[Dict[str, Any]]) -> None:
    """
    MAC Summary sheet — counts from 'show mac address-table count'.

    IOS-XE: one row per VLAN per device + a TOTAL row per device.
    NX-OS:  one TOTAL row per device (no per-VLAN breakdown available).

    Columns: Hostname | IP | VLAN | Dynamic | Static | Total
    """
    ws = wb.create_sheet("MAC Summary")
    ws.append(["Hostname", "IP", "VLAN", "Dynamic", "Static", "Total"])

    for dev in devices_data:
        hostname   = dev.get('meta', {}).get('hostname', 'N/A') or 'N/A'
        ip         = dev.get('meta', {}).get('host', 'N/A')
        structured = dev.get('structured', {})

        counts = structured.get('show mac address-table count')
        if not isinstance(counts, dict):
            continue

        # Per-VLAN rows (IOS-XE only)
        vlans = counts.get('vlans', {})
        for vlan_id, vdata in sorted(vlans.items(), key=lambda x: int(x[0]) if x[0].isdigit() else 0):
            ws.append([
                hostname, ip, vlan_id,
                vdata.get('dynamic', ''),
                vdata.get('static', ''),
                vdata.get('total', ''),
            ])

        # Totals row
        totals = counts.get('totals', {})
        if totals:
            ws.append([
                hostname, ip, 'TOTAL',
                totals.get('dynamic', ''),
                totals.get('static', ''),
                totals.get('total', ''),
            ])

    _apply_header_style(ws)


def _sheet_templates(wb: Any, devices_data: List[Dict[str, Any]]) -> None:
    """
    Templates sheet — IOS-XE interface templates (show template) and
    NX-OS port profiles (show port-profile).
    Columns: Hostname | IP | Name | Source | Status | Port Count | Interfaces
      Source  — 'User' / 'Built-in' (IOS-XE) or profile type e.g. 'Ethernet' (NX-OS)
      Status  — 'Bound' / 'Unbound' (IOS-XE) or 'enabled' / 'disabled' (NX-OS)
      Port Count — number of bound/assigned interfaces (or interface-range entries)
      Interfaces — comma-separated list of bound ports or interface range strings
    """
    ws = wb.create_sheet("Templates")
    ws.append(["Hostname", "IP", "Template Name", "Source", "Status", "Interfaces", "Port Count"])

    for dev in devices_data:
        hostname   = dev.get('meta', {}).get('hostname', 'N/A') or 'N/A'
        ip         = dev.get('meta', {}).get('host', 'N/A')
        structured = dev.get('structured', {})

        # IOS / IOS-XE: show template
        # type field holds 'User' or 'Built-in' (the Source column in show template output)
        templates = structured.get('show template', {})
        if isinstance(templates, dict):
            for name, data in sorted(templates.items()):
                intfs = data.get('interfaces', [])
                ws.append([
                    hostname, ip, name,
                    data.get('type', 'N/A'),          # 'User' or 'Built-in'
                    'Bound' if intfs else 'Unbound',
                    ', '.join(intfs),
                    len(intfs),
                ])

        # NX-OS: show port-profile
        # type field holds profile type e.g. 'Ethernet', 'Pseudowire'
        profiles = structured.get('show port-profile', {})
        if isinstance(profiles, dict):
            for name, data in sorted(profiles.items()):
                intfs = data.get('interfaces', [])
                ws.append([
                    hostname, ip, name,
                    data.get('type', 'N/A'),    # 'Ethernet', 'Pseudowire', etc.
                    data.get('status', 'N/A'),  # 'enabled' or 'disabled'
                    ', '.join(intfs),
                    len(intfs),
                ])

    _apply_header_style(ws)


def generate_excel_report(run_folder: Path, logger: logging.Logger,
                          selected_components: Optional[Set[str]] = None) -> Optional[Path]:
    """
    Generate an Excel workbook from per-device JSON files in the run's json/ subfolder.
    Only sheets for selected components are included (Run Info and Summary always included).
    Returns the saved workbook path, or None on failure.
    """
    try:
        from openpyxl import Workbook  # type: ignore
    except ImportError:
        logger.error("openpyxl is not installed — run: pip install openpyxl")
        return None

    json_folder = run_folder / "json"
    if not json_folder.exists():
        logger.warning("No json/ folder found — Structured Output must be enabled to generate reports.")
        return None

    json_files = sorted(json_folder.glob("*.json"))
    if not json_files:
        logger.warning("No JSON files found in json/ — nothing to report.")
        return None

    devices_data: List[Dict[str, Any]] = []
    for jf in json_files:
        try:
            with open(jf, 'r', encoding='utf-8') as f:
                devices_data.append(json.load(f))
        except Exception as e:
            logger.warning(f"Skipping {jf.name}: {e}")
    if not devices_data:
        return None

    components = selected_components or set(COMPONENT_NAMES)

    wb = Workbook()
    wb.remove(wb.active)  # Remove the default empty sheet

    # Run Info and Summary are always included
    _sheet_run_info(wb, run_folder, devices_data)
    _sheet_summary(wb, devices_data)

    # Component-driven sheets
    _COMPONENT_SHEETS: Dict[str, Any] = {
        'Interfaces':    _sheet_interfaces,
        'Neighbors':     _sheet_neighbors,
        'VLANs':         _sheet_vlans,
        'Routing':       _sheet_routing,
        'STP':           _sheet_stp,
        'MAC Addresses': _sheet_mac,
        'MAC Summary':   _sheet_mac_summary,
        'Templates':     _sheet_templates,
    }
    for comp_name in COMPONENT_NAMES:
        if comp_name in components and comp_name in _COMPONENT_SHEETS:
            _COMPONENT_SHEETS[comp_name](wb, devices_data)

    report_path = run_folder / "nddu_report.xlsx"
    try:
        wb.save(report_path)
        logger.info(f'Excel report saved: "{report_path.name}"')
        return report_path
    except Exception as e:
        logger.error(f"Failed to save Excel report: {e}")
        return None


# ─────────────────────────────────────────────────────────────────────────────
# Device Diff Engine
# ─────────────────────────────────────────────────────────────────────────────

def _load_run_devices(run_folder: Path) -> Dict[str, Dict[str, Any]]:
    """Load json/*.json from a run folder, return dict keyed by meta['host'] (IP)."""
    devices: Dict[str, Dict[str, Any]] = {}
    json_dir = run_folder / "json"
    if not json_dir.is_dir():
        return devices
    for fp in sorted(json_dir.glob("*.json")):
        try:
            data = json.loads(fp.read_text(encoding='utf-8'))
            host = data.get('meta', {}).get('host', '')
            if host:
                devices[host] = data
        except (json.JSONDecodeError, OSError):
            continue
    return devices


def _match_devices(
    devs_a: Dict[str, Dict[str, Any]],
    devs_b: Dict[str, Dict[str, Any]],
) -> Tuple[Dict[str, Tuple], Dict[str, Dict], Dict[str, Dict]]:
    """Two-pass device matching: primary by IP, secondary by hostname for unmatched.

    Returns:
        matched  — {key: (dev_a, dev_b)} where key is IP or hostname
        added    — {key: dev_b}  (in B only)
        removed  — {key: dev_a}  (in A only)
    """
    matched: Dict[str, Tuple] = {}
    added:   Dict[str, Dict]  = {}
    removed: Dict[str, Dict]  = {}

    # Pass 1: match by IP
    ips_a = set(devs_a.keys())
    ips_b = set(devs_b.keys())
    common_ips = ips_a & ips_b
    for ip in common_ips:
        matched[ip] = (devs_a[ip], devs_b[ip])

    unmatched_a = {ip: devs_a[ip] for ip in ips_a - common_ips}
    unmatched_b = {ip: devs_b[ip] for ip in ips_b - common_ips}

    # Pass 2: match remaining by hostname
    hostname_map_a: Dict[str, Tuple[str, Dict]] = {}
    for ip, dev in unmatched_a.items():
        hn = dev.get('meta', {}).get('hostname', '').lower()
        if hn:
            hostname_map_a[hn] = (ip, dev)

    still_unmatched_b: Dict[str, Dict] = {}
    matched_a_hostnames: set = set()
    for ip_b, dev_b in unmatched_b.items():
        hn = dev_b.get('meta', {}).get('hostname', '').lower()
        if hn and hn in hostname_map_a:
            ip_a, dev_a = hostname_map_a[hn]
            # Use hostname as key to indicate IP changed
            key = dev_b.get('meta', {}).get('hostname', hn)
            matched[key] = (dev_a, dev_b)
            matched_a_hostnames.add(hn)
        else:
            still_unmatched_b[ip_b] = dev_b

    # Remaining unmatched
    for hn, (ip, dev) in hostname_map_a.items():
        if hn not in matched_a_hostnames:
            removed[ip] = dev
    added.update(still_unmatched_b)

    return matched, added, removed


def _diff_list_by_key(
    list_a: List[Dict], list_b: List[Dict],
    key_field: str, compare_fields: Optional[List[str]] = None,
) -> List[Tuple[str, Optional[Dict], Optional[Dict], str]]:
    """Compare two lists of dicts matched by key_field.

    Returns [(key, row_a_or_None, row_b_or_None, status), ...]
    status is one of: 'added', 'removed', 'changed', 'unchanged'
    """
    map_a = {str(row.get(key_field, '')): row for row in (list_a or [])}
    map_b = {str(row.get(key_field, '')): row for row in (list_b or [])}
    all_keys = list(dict.fromkeys(list(map_a.keys()) + list(map_b.keys())))

    results: List[Tuple[str, Optional[Dict], Optional[Dict], str]] = []
    for key in all_keys:
        a = map_a.get(key)
        b = map_b.get(key)
        if a and not b:
            results.append((key, a, None, 'removed'))
        elif b and not a:
            results.append((key, None, b, 'added'))
        else:
            fields = compare_fields or [k for k in set(list((a or {}).keys()) + list((b or {}).keys())) if k != key_field]
            changed = any(str(a.get(f, '')) != str(b.get(f, '')) for f in fields)
            results.append((key, a, b, 'changed' if changed else 'unchanged'))
    return results


def _diff_dict(
    dict_a: Dict[str, Any], dict_b: Dict[str, Any],
    ignore_keys: Optional[set] = None,
) -> Dict[str, Tuple[Any, Any]]:
    """Compare two flat dicts. Returns {field: (old, new)} for changed fields."""
    ignore = ignore_keys or set()
    changes: Dict[str, Tuple[Any, Any]] = {}
    all_keys = set(dict_a.keys()) | set(dict_b.keys())
    for k in all_keys:
        if k in ignore:
            continue
        va = dict_a.get(k, 'N/A')
        vb = dict_b.get(k, 'N/A')
        if str(va) != str(vb):
            changes[k] = (va, vb)
    return changes


def _diff_fills() -> Dict[str, Any]:
    """Return dict of PatternFill objects for diff status coloring."""
    from openpyxl.styles import PatternFill  # type: ignore
    return {
        'added':   PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),  # green
        'removed': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),  # red
        'changed': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),  # yellow
    }


def _apply_diff_fill(ws: Any, row_num: int, fill: Any) -> None:
    """Apply fill to every cell in a row."""
    for cell in ws[row_num]:
        cell.fill = fill


def _append_no_changes(ws: Any, num_cols: int) -> None:
    """If a diff sheet has only the header row, append a status message."""
    if ws.max_row <= 1:
        row = ["No changes detected between runs."] + [""] * (num_cols - 1)
        ws.append(row)


def _diff_sheet_run_info(
    wb: Any, run_a: Path, run_b: Path,
    devs_a: Dict[str, Dict], devs_b: Dict[str, Dict],
    port_map: Optional[Dict[str, str]] = None,
) -> None:
    """Side-by-side run metadata comparison."""
    ws = wb.active
    ws.title = "Diff - Run Info"
    ws.append(["Field", "Run A", "Run B"])

    meta_a = next(iter(devs_a.values()), {}).get('meta', {}) if devs_a else {}
    meta_b = next(iter(devs_b.values()), {}).get('meta', {}) if devs_b else {}

    ws.append(["Client",             run_a.parent.name, run_b.parent.name])
    ws.append(["Run Folder",         run_a.name, run_b.name])
    ws.append(["nddu Version",       meta_a.get('nddu_version', 'N/A'), meta_b.get('nddu_version', 'N/A')])
    ws.append(["Devices Documented", len(devs_a), len(devs_b)])
    if port_map:
        ws.append(["Port Map", f"{len(port_map)} interface mapping(s) applied", ""])
    else:
        ws.append(["Port Map", "None", ""])
    _apply_header_style(ws)


def _diff_sheet_summary(
    wb: Any,
    matched: Dict[str, Tuple],
    added: Dict[str, Dict],
    removed: Dict[str, Dict],
    fills: Dict[str, Any],
) -> None:
    """One row per device showing status and key field changes."""
    ws = wb.create_sheet("Diff - Summary")
    ws.append(["Status", "Hostname (A)", "Hostname (B)", "IP (A)", "IP (B)",
               "Device Type", "Platform (A)", "Platform (B)",
               "Version (A)", "Version (B)", "Uptime (A)", "Uptime (B)"])

    def _extract_summary(dev: Dict) -> Dict[str, str]:
        meta = dev.get('meta', {})
        structured = dev.get('structured', {})
        ver_list = structured.get('show version')
        ver = ver_list[0] if isinstance(ver_list, list) and ver_list else {}
        hw = ver.get('hardware', [])
        platform = hw[0] if isinstance(hw, list) and hw else ver.get('platform', 'N/A')
        return {
            'hostname':    meta.get('hostname', 'N/A'),
            'ip':          meta.get('host', 'N/A'),
            'device_type': meta.get('device_type', 'N/A'),
            'platform':    platform,
            'version':     ver.get('version', 'N/A') or 'N/A',
            'uptime':      ver.get('uptime', 'N/A') or 'N/A',
        }

    # Pairs of (A_col_index, B_col_index, field_name) — 1-based Excel columns.
    # Device Type is a single column (index 6) shared between runs.
    pair_cols = [
        (2, 3, 'hostname'),
        (4, 5, 'ip'),
        (7, 8, 'platform'),
        (9, 10, 'version'),
        (11, 12, 'uptime'),
    ]

    # Matched devices — highlight only the cell pairs that differ.
    for key, (dev_a, dev_b) in sorted(matched.items()):
        sa = _extract_summary(dev_a)
        sb = _extract_summary(dev_b)
        field_changed = {f: sa[f] != sb[f] for f in
                         ('hostname', 'ip', 'device_type', 'platform', 'version', 'uptime')}
        status = 'Changed' if any(field_changed.values()) else 'Unchanged'
        ws.append([status, sa['hostname'], sb['hostname'], sa['ip'], sb['ip'],
                   sa['device_type'], sa['platform'], sb['platform'],
                   sa['version'], sb['version'], sa['uptime'], sb['uptime']])
        row_num = ws.max_row
        if status == 'Changed':
            ws.cell(row=row_num, column=1).fill = fills['changed']
            for col_a, col_b, field in pair_cols:
                if field_changed[field]:
                    ws.cell(row=row_num, column=col_a).fill = fills['changed']
                    ws.cell(row=row_num, column=col_b).fill = fills['changed']
            if field_changed['device_type']:
                ws.cell(row=row_num, column=6).fill = fills['changed']

    # Added devices (in B only)
    for key, dev in sorted(added.items()):
        sb = _extract_summary(dev)
        ws.append(['Added', '', sb['hostname'], '', sb['ip'],
                   sb['device_type'], '', sb['platform'],
                   '', sb['version'], '', sb['uptime']])
        _apply_diff_fill(ws, ws.max_row, fills['added'])

    # Removed devices (in A only)
    for key, dev in sorted(removed.items()):
        sa = _extract_summary(dev)
        ws.append(['Removed', sa['hostname'], '', sa['ip'], '',
                   sa['device_type'], sa['platform'], '',
                   sa['version'], '', sa['uptime'], ''])
        _apply_diff_fill(ws, ws.max_row, fills['removed'])

    _append_no_changes(ws, 12)
    _apply_header_style(ws)


def _diff_sheet_interfaces(
    wb: Any,
    matched: Dict[str, Tuple],
    added: Dict[str, Dict],
    removed: Dict[str, Dict],
    fills: Dict[str, Any],
    port_map: Optional[Dict[str, str]] = None,
) -> None:
    """Per-device interface comparison: added/removed/changed interfaces."""
    ws = wb.create_sheet("Diff - Interfaces")
    ws.append(["Hostname", "IP", "Interface", "Status",
               "Description (A)", "Description (B)",
               "IP Address (A)", "IP Address (B)",
               "Vlan (A)", "Vlan (B)",
               "Speed (A)", "Speed (B)",
               "Duplex (A)", "Duplex (B)"])

    def _get_intf_map(dev: Dict) -> Dict[str, Dict]:
        structured = dev.get('structured', {})
        status_list = structured.get('show interface status', [])
        brief_list = structured.get('show ip interface brief', [])
        desc_list = structured.get('show interfaces description', [])

        intf_map: Dict[str, Dict] = {}
        for row in (status_list if isinstance(status_list, list) else []):
            name = _normalize_intf_name(row.get('port', row.get('interface', '')))
            if name:
                intf_map[name] = {
                    'description': row.get('name', row.get('description', '')),
                    'status': row.get('status', ''),
                    'vlan': row.get('vlan', ''),
                    'speed': row.get('speed', ''),
                    'duplex': row.get('duplex', ''),
                    'type': row.get('type', ''),
                    'ip_address': '',
                }
        for row in (brief_list if isinstance(brief_list, list) else []):
            name = _normalize_intf_name(row.get('intf', row.get('interface', '')))
            if name and name in intf_map:
                intf_map[name]['ip_address'] = row.get('ipaddr', row.get('ip_address', ''))
            elif name:
                intf_map[name] = {
                    'description': '', 'status': row.get('status', ''),
                    'vlan': '', 'speed': '', 'duplex': '', 'type': '',
                    'ip_address': row.get('ipaddr', row.get('ip_address', '')),
                }
        for row in (desc_list if isinstance(desc_list, list) else []):
            name = _normalize_intf_name(row.get('port', row.get('interface', '')))
            if name and name in intf_map and not intf_map[name]['description']:
                intf_map[name]['description'] = row.get('descrip', row.get('description', ''))
        return intf_map

    def _write_device_intfs(hostname, ip, intf_a, intf_b):
        all_names = sorted(set(list(intf_a.keys()) + list(intf_b.keys())))
        for name in all_names:
            a = intf_a.get(name)
            b = intf_b.get(name)
            if a and not b:
                ws.append([hostname, ip, name, 'Removed',
                           a.get('description', ''), '',
                           a.get('ip_address', ''), '',
                           a.get('vlan', ''), '',
                           a.get('speed', ''), '',
                           a.get('duplex', ''), ''])
                _apply_diff_fill(ws, ws.max_row, fills['removed'])
            elif b and not a:
                ws.append([hostname, ip, name, 'Added',
                           '', b.get('description', ''),
                           '', b.get('ip_address', ''),
                           '', b.get('vlan', ''),
                           '', b.get('speed', ''),
                           '', b.get('duplex', '')])
                _apply_diff_fill(ws, ws.max_row, fills['added'])
            else:
                cmp_fields = ['description', 'ip_address', 'vlan', 'speed', 'duplex']
                changed = any(str(a.get(f, '')) != str(b.get(f, '')) for f in cmp_fields)
                if changed:
                    ws.append([hostname, ip, name, 'Changed',
                               a.get('description', ''), b.get('description', ''),
                               a.get('ip_address', ''), b.get('ip_address', ''),
                               a.get('vlan', ''), b.get('vlan', ''),
                               a.get('speed', ''), b.get('speed', ''),
                               a.get('duplex', ''), b.get('duplex', '')])
                    _apply_diff_fill(ws, ws.max_row, fills['changed'])

    for key, (dev_a, dev_b) in sorted(matched.items()):
        hostname = dev_b.get('meta', {}).get('hostname', key)
        ip = dev_b.get('meta', {}).get('host', key)
        _write_device_intfs(hostname, ip,
                            _remap_intf_keys(_get_intf_map(dev_a), port_map),
                            _get_intf_map(dev_b))

    for key, dev in sorted(added.items()):
        hostname = dev.get('meta', {}).get('hostname', key)
        ip = dev.get('meta', {}).get('host', key)
        _write_device_intfs(hostname, ip, {}, _get_intf_map(dev))

    for key, dev in sorted(removed.items()):
        hostname = dev.get('meta', {}).get('hostname', key)
        ip = dev.get('meta', {}).get('host', key)
        _write_device_intfs(hostname, ip, _get_intf_map(dev), {})

    _append_no_changes(ws, 14)
    _apply_header_style(ws)


def _diff_sheet_neighbors(
    wb: Any,
    matched: Dict[str, Tuple],
    added: Dict[str, Dict],
    removed: Dict[str, Dict],
    fills: Dict[str, Any],
    port_map: Optional[Dict[str, str]] = None,
) -> None:
    """CDP neighbor comparison per device."""
    ws = wb.create_sheet("Diff - Neighbors")
    ws.append(["Hostname", "IP", "Local Interface", "Status",
               "Neighbor (A)", "Neighbor (B)",
               "Neighbor IP (A)", "Neighbor IP (B)",
               "Platform (A)", "Platform (B)"])

    def _get_neighbor_map(dev: Dict) -> Dict[str, Dict]:
        structured = dev.get('structured', {})
        cdp = structured.get('show cdp neighbors detail', [])
        nbr_map: Dict[str, Dict] = {}
        for row in (cdp if isinstance(cdp, list) else []):
            local_intf = _normalize_intf_name(
                row.get('local_interface', row.get('local_port', '')))
            neighbor = row.get('destination_host', row.get('neighbor', ''))
            key = f"{local_intf}|{neighbor}"
            nbr_map[key] = {
                'local_intf': local_intf,
                'neighbor': neighbor,
                'neighbor_ip': row.get('management_ip', row.get('neighbor_ip', '')),
                'platform': row.get('platform', ''),
            }
        return nbr_map

    def _write_device_neighbors(hostname, ip, nbr_a, nbr_b):
        all_keys = sorted(set(list(nbr_a.keys()) + list(nbr_b.keys())))
        for key in all_keys:
            a = nbr_a.get(key)
            b = nbr_b.get(key)
            local_intf = (b or a or {}).get('local_intf', key.split('|')[0])
            if a and not b:
                ws.append([hostname, ip, local_intf, 'Removed',
                           a['neighbor'], '', a['neighbor_ip'], '', a['platform'], ''])
                _apply_diff_fill(ws, ws.max_row, fills['removed'])
            elif b and not a:
                ws.append([hostname, ip, local_intf, 'Added',
                           '', b['neighbor'], '', b['neighbor_ip'], '', b['platform']])
                _apply_diff_fill(ws, ws.max_row, fills['added'])
            else:
                changed = (a['neighbor_ip'] != b['neighbor_ip'] or
                           a['platform'] != b['platform'])
                if changed:
                    ws.append([hostname, ip, local_intf, 'Changed',
                               a['neighbor'], b['neighbor'],
                               a['neighbor_ip'], b['neighbor_ip'],
                               a['platform'], b['platform']])
                    _apply_diff_fill(ws, ws.max_row, fills['changed'])

    for key, (dev_a, dev_b) in sorted(matched.items()):
        hostname = dev_b.get('meta', {}).get('hostname', key)
        ip = dev_b.get('meta', {}).get('host', key)
        nbr_a = _get_neighbor_map(dev_a)
        if port_map:
            remapped_a: Dict[str, Dict] = {}
            for nbr_key, nbr_data in nbr_a.items():
                old_intf = nbr_data['local_intf']
                new_intf = port_map.get(old_intf, old_intf)
                new_key = f"{new_intf}|{nbr_data['neighbor']}"
                nbr_data_copy = dict(nbr_data)
                nbr_data_copy['local_intf'] = new_intf
                remapped_a[new_key] = nbr_data_copy
            nbr_a = remapped_a
        _write_device_neighbors(hostname, ip, nbr_a, _get_neighbor_map(dev_b))

    for key, dev in sorted(added.items()):
        hostname = dev.get('meta', {}).get('hostname', key)
        ip = dev.get('meta', {}).get('host', key)
        _write_device_neighbors(hostname, ip, {}, _get_neighbor_map(dev))

    for key, dev in sorted(removed.items()):
        hostname = dev.get('meta', {}).get('hostname', key)
        ip = dev.get('meta', {}).get('host', key)
        _write_device_neighbors(hostname, ip, _get_neighbor_map(dev), {})

    _append_no_changes(ws, 10)
    _apply_header_style(ws)


def _diff_sheet_vlans(
    wb: Any,
    matched: Dict[str, Tuple],
    added: Dict[str, Dict],
    removed: Dict[str, Dict],
    fills: Dict[str, Any],
    port_map: Optional[Dict[str, str]] = None,
) -> None:
    """VLAN comparison per device."""
    ws = wb.create_sheet("Diff - VLANs")
    ws.append(["Hostname", "IP", "VLAN ID", "Status",
               "Name (A)", "Name (B)", "Status (A)", "Status (B)",
               "Interfaces (A)", "Interfaces (B)"])

    def _get_vlan_map(dev: Dict) -> Dict[str, Dict]:
        structured = dev.get('structured', {})
        # IOS-XE: show vlan brief → dict with 'vlans' (dict-of-dicts) or list
        vlan_data = structured.get('show vlan brief', {})
        if isinstance(vlan_data, dict):
            # Prefer 'vlans' key (clean numeric IDs); fall back to 'vlan' key
            vlan_data = vlan_data.get('vlans', vlan_data.get('vlan', {}))
        # NX-OS: show vlan
        if not vlan_data:
            vlan_data = structured.get('show vlan', {})
            if isinstance(vlan_data, dict):
                vlan_data = vlan_data.get('vlans', {})

        vlan_map: Dict[str, Dict] = {}
        # Handle dict-of-dicts (keyed by VLAN ID) or list-of-dicts
        items: list = []
        if isinstance(vlan_data, dict):
            items = [(str(vid), info) for vid, info in vlan_data.items()
                     if isinstance(info, dict)]
        elif isinstance(vlan_data, list):
            items = [(str(row.get('vlan_id', '')), row) for row in vlan_data]

        for vid, info in items:
            if not vid:
                continue
            intfs = info.get('interfaces', info.get('vlan_port', []))
            if isinstance(intfs, list):
                intfs = ', '.join(intfs)
            vlan_map[vid] = {
                'name': info.get('name', info.get('vlan_name', '')),
                'status': info.get('status', info.get('vlan_status', '')),
                'interfaces': str(intfs),
            }
        return vlan_map

    def _write_device_vlans(hostname, ip, vlan_a, vlan_b):
        all_vids = sorted(set(list(vlan_a.keys()) + list(vlan_b.keys())),
                          key=lambda x: int(x) if x.isdigit() else 0)
        for vid in all_vids:
            a = vlan_a.get(vid)
            b = vlan_b.get(vid)
            if a and not b:
                ws.append([hostname, ip, vid, 'Removed',
                           a['name'], '', a['status'], '', a['interfaces'], ''])
                _apply_diff_fill(ws, ws.max_row, fills['removed'])
            elif b and not a:
                ws.append([hostname, ip, vid, 'Added',
                           '', b['name'], '', b['status'], '', b['interfaces']])
                _apply_diff_fill(ws, ws.max_row, fills['added'])
            else:
                changed = (a['name'] != b['name'] or a['status'] != b['status'] or
                           a['interfaces'] != b['interfaces'])
                if changed:
                    ws.append([hostname, ip, vid, 'Changed',
                               a['name'], b['name'], a['status'], b['status'],
                               a['interfaces'], b['interfaces']])
                    _apply_diff_fill(ws, ws.max_row, fills['changed'])

    for key, (dev_a, dev_b) in sorted(matched.items()):
        hostname = dev_b.get('meta', {}).get('hostname', key)
        ip = dev_b.get('meta', {}).get('host', key)
        vlan_a = _get_vlan_map(dev_a)
        vlan_b = _get_vlan_map(dev_b)
        if port_map:
            for info in vlan_a.values():
                info['interfaces'] = _remap_intf_string(info['interfaces'], port_map)
            for info in vlan_b.values():
                info['interfaces'] = _normalize_intf_string(info['interfaces'])
        _write_device_vlans(hostname, ip, vlan_a, vlan_b)

    for key, dev in sorted(added.items()):
        hostname = dev.get('meta', {}).get('hostname', key)
        ip = dev.get('meta', {}).get('host', key)
        _write_device_vlans(hostname, ip, {}, _get_vlan_map(dev))

    for key, dev in sorted(removed.items()):
        hostname = dev.get('meta', {}).get('hostname', key)
        ip = dev.get('meta', {}).get('host', key)
        _write_device_vlans(hostname, ip, _get_vlan_map(dev), {})

    _append_no_changes(ws, 10)
    _apply_header_style(ws)


def _diff_sheet_routing(
    wb: Any,
    matched: Dict[str, Tuple],
    added: Dict[str, Dict],
    removed: Dict[str, Dict],
    fills: Dict[str, Any],
) -> None:
    """Route count changes from show ip route summary."""
    ws = wb.create_sheet("Diff - Routing")
    ws.append(["Hostname", "IP", "Protocol", "Status",
               "Routes (A)", "Routes (B)", "Subnets (A)", "Subnets (B)"])

    def _get_route_map(dev: Dict) -> Dict[str, Dict]:
        structured = dev.get('structured', {})
        route_data = structured.get('show ip route summary', {})
        if isinstance(route_data, list) and route_data:
            route_data = route_data[0] if route_data else {}
        protocols = route_data.get('protocols', []) if isinstance(route_data, dict) else []
        route_map: Dict[str, Dict] = {}
        for proto in (protocols if isinstance(protocols, list) else []):
            name = proto.get('protocol', proto.get('name', ''))
            if name:
                route_map[name] = {
                    'routes': str(proto.get('networks', proto.get('routes', ''))),
                    'subnets': str(proto.get('subnets', '')),
                }
        # Also add total if available
        total = route_data.get('total', {}) if isinstance(route_data, dict) else {}
        if total:
            route_map['TOTAL'] = {
                'routes': str(total.get('networks', total.get('routes', ''))),
                'subnets': str(total.get('subnets', '')),
            }
        return route_map

    def _write_device_routes(hostname, ip, route_a, route_b):
        all_protos = list(dict.fromkeys(list(route_a.keys()) + list(route_b.keys())))
        for proto in all_protos:
            a = route_a.get(proto)
            b = route_b.get(proto)
            if a and not b:
                ws.append([hostname, ip, proto, 'Removed',
                           a['routes'], '', a['subnets'], ''])
                _apply_diff_fill(ws, ws.max_row, fills['removed'])
            elif b and not a:
                ws.append([hostname, ip, proto, 'Added',
                           '', b['routes'], '', b['subnets']])
                _apply_diff_fill(ws, ws.max_row, fills['added'])
            else:
                changed = a['routes'] != b['routes'] or a['subnets'] != b['subnets']
                if changed:
                    ws.append([hostname, ip, proto, 'Changed',
                               a['routes'], b['routes'], a['subnets'], b['subnets']])
                    _apply_diff_fill(ws, ws.max_row, fills['changed'])

    for key, (dev_a, dev_b) in sorted(matched.items()):
        hostname = dev_b.get('meta', {}).get('hostname', key)
        ip = dev_b.get('meta', {}).get('host', key)
        _write_device_routes(hostname, ip, _get_route_map(dev_a), _get_route_map(dev_b))

    for key, dev in sorted(added.items()):
        hostname = dev.get('meta', {}).get('hostname', key)
        ip = dev.get('meta', {}).get('host', key)
        _write_device_routes(hostname, ip, {}, _get_route_map(dev))

    for key, dev in sorted(removed.items()):
        hostname = dev.get('meta', {}).get('hostname', key)
        ip = dev.get('meta', {}).get('host', key)
        _write_device_routes(hostname, ip, _get_route_map(dev), {})

    _append_no_changes(ws, 8)
    _apply_header_style(ws)


def _diff_sheet_stp(
    wb: Any,
    matched: Dict[str, Tuple],
    added: Dict[str, Dict],
    removed: Dict[str, Dict],
    fills: Dict[str, Any],
) -> None:
    """STP summary comparison per device."""
    ws = wb.create_sheet("Diff - STP")
    ws.append(["Hostname", "IP", "Status",
               "Mode (A)", "Mode (B)",
               "Root Bridge For (A)", "Root Bridge For (B)",
               "VLANs (A)", "VLANs (B)",
               "Blocking (A)", "Blocking (B)",
               "Forwarding (A)", "Forwarding (B)"])

    def _get_stp(dev: Dict) -> Optional[Dict]:
        structured = dev.get('structured', {})
        stp_data = structured.get('show spanning-tree summary', {})
        if isinstance(stp_data, list) and stp_data:
            stp_data = stp_data[0]
        if not isinstance(stp_data, dict) or not stp_data:
            return None
        return {
            'mode': stp_data.get('mode', 'N/A'),
            'root_bridge_for': stp_data.get('root_bridge_for', 'N/A'),
            'num_vlans': str(stp_data.get('num_vlans', stp_data.get('vlans', 'N/A'))),
            'blocking': str(stp_data.get('blocking', 'N/A')),
            'forwarding': str(stp_data.get('forwarding', 'N/A')),
        }

    def _write_stp_row(hostname, ip, stp_a, stp_b, status):
        a = stp_a or {'mode': '', 'root_bridge_for': '', 'num_vlans': '', 'blocking': '', 'forwarding': ''}
        b = stp_b or {'mode': '', 'root_bridge_for': '', 'num_vlans': '', 'blocking': '', 'forwarding': ''}
        ws.append([hostname, ip, status,
                   a['mode'], b['mode'],
                   a['root_bridge_for'], b['root_bridge_for'],
                   a['num_vlans'], b['num_vlans'],
                   a['blocking'], b['blocking'],
                   a['forwarding'], b['forwarding']])
        if status in fills:
            _apply_diff_fill(ws, ws.max_row, fills[status.lower()])

    for key, (dev_a, dev_b) in sorted(matched.items()):
        hostname = dev_b.get('meta', {}).get('hostname', key)
        ip = dev_b.get('meta', {}).get('host', key)
        stp_a, stp_b = _get_stp(dev_a), _get_stp(dev_b)
        if stp_a or stp_b:
            changed = stp_a != stp_b
            _write_stp_row(hostname, ip, stp_a, stp_b, 'Changed' if changed else 'Unchanged')
            if changed:
                _apply_diff_fill(ws, ws.max_row, fills['changed'])

    for key, dev in sorted(added.items()):
        hostname = dev.get('meta', {}).get('hostname', key)
        ip = dev.get('meta', {}).get('host', key)
        stp = _get_stp(dev)
        if stp:
            _write_stp_row(hostname, ip, None, stp, 'Added')
            _apply_diff_fill(ws, ws.max_row, fills['added'])

    for key, dev in sorted(removed.items()):
        hostname = dev.get('meta', {}).get('hostname', key)
        ip = dev.get('meta', {}).get('host', key)
        stp = _get_stp(dev)
        if stp:
            _write_stp_row(hostname, ip, stp, None, 'Removed')
            _apply_diff_fill(ws, ws.max_row, fills['removed'])

    _append_no_changes(ws, 13)
    _apply_header_style(ws)


def _diff_sheet_mac_summary(
    wb: Any,
    matched: Dict[str, Tuple],
    added: Dict[str, Dict],
    removed: Dict[str, Dict],
    fills: Dict[str, Any],
) -> None:
    """MAC address table count comparison per device/VLAN."""
    ws = wb.create_sheet("Diff - MAC Summary")
    ws.append(["Hostname", "IP", "VLAN", "Status",
               "Dynamic (A)", "Dynamic (B)",
               "Static (A)", "Static (B)",
               "Total (A)", "Total (B)"])

    def _get_mac_summary(dev: Dict) -> Dict[str, Dict]:
        structured = dev.get('structured', {})
        mac_data = structured.get('show mac address-table count', {})
        if isinstance(mac_data, list) and mac_data:
            mac_data = mac_data[0] if mac_data else {}
        if not isinstance(mac_data, dict):
            return {}
        vlans = mac_data.get('vlans', {})
        totals = mac_data.get('totals', {})
        result: Dict[str, Dict] = {}
        # Handle dict-of-dicts (keyed by VLAN ID) or list-of-dicts
        if isinstance(vlans, dict):
            for vid, info in vlans.items():
                if isinstance(info, dict):
                    result[str(vid)] = {
                        'dynamic': str(info.get('dynamic', '')),
                        'static': str(info.get('static', '')),
                        'total': str(info.get('total', '')),
                    }
        elif isinstance(vlans, list):
            for v in vlans:
                vid = str(v.get('vlan', v.get('vlan_id', '')))
                if vid:
                    result[vid] = {
                        'dynamic': str(v.get('dynamic', '')),
                        'static': str(v.get('static', '')),
                        'total': str(v.get('total', '')),
                    }
        if isinstance(totals, dict) and totals:
            result['TOTAL'] = {
                'dynamic': str(totals.get('dynamic', '')),
                'static': str(totals.get('static', '')),
                'total': str(totals.get('total', '')),
            }
        return result

    def _write_mac_summary(hostname, ip, ms_a, ms_b):
        all_vlans = sorted(set(list(ms_a.keys()) + list(ms_b.keys())),
                           key=lambda x: (0 if x.isdigit() else 1, int(x) if x.isdigit() else 0, x))
        for vid in all_vlans:
            a = ms_a.get(vid)
            b = ms_b.get(vid)
            if a and not b:
                ws.append([hostname, ip, vid, 'Removed',
                           a['dynamic'], '', a['static'], '', a['total'], ''])
                _apply_diff_fill(ws, ws.max_row, fills['removed'])
            elif b and not a:
                ws.append([hostname, ip, vid, 'Added',
                           '', b['dynamic'], '', b['static'], '', b['total']])
                _apply_diff_fill(ws, ws.max_row, fills['added'])
            else:
                changed = a != b
                if changed:
                    ws.append([hostname, ip, vid, 'Changed',
                               a['dynamic'], b['dynamic'], a['static'], b['static'],
                               a['total'], b['total']])
                    _apply_diff_fill(ws, ws.max_row, fills['changed'])

    for key, (dev_a, dev_b) in sorted(matched.items()):
        hostname = dev_b.get('meta', {}).get('hostname', key)
        ip = dev_b.get('meta', {}).get('host', key)
        _write_mac_summary(hostname, ip, _get_mac_summary(dev_a), _get_mac_summary(dev_b))

    for key, dev in sorted(added.items()):
        hostname = dev.get('meta', {}).get('hostname', key)
        ip = dev.get('meta', {}).get('host', key)
        _write_mac_summary(hostname, ip, {}, _get_mac_summary(dev))

    for key, dev in sorted(removed.items()):
        hostname = dev.get('meta', {}).get('hostname', key)
        ip = dev.get('meta', {}).get('host', key)
        _write_mac_summary(hostname, ip, _get_mac_summary(dev), {})

    _append_no_changes(ws, 10)
    _apply_header_style(ws)


def _diff_sheet_mac_addresses(
    wb: Any,
    matched: Dict[str, Tuple],
    added: Dict[str, Dict],
    removed: Dict[str, Dict],
    fills: Dict[str, Any],
    port_map: Optional[Dict[str, str]] = None,
) -> None:
    """MAC address table comparison — only shows added/removed MACs (skip unchanged)."""
    ws = wb.create_sheet("Diff - MAC Addresses")
    ws.append(["Hostname", "IP", "VLAN", "MAC Address", "Status",
               "Type (A)", "Type (B)", "Interface (A)", "Interface (B)"])

    def _get_mac_map(dev: Dict) -> Dict[str, Dict]:
        structured = dev.get('structured', {})
        mac_list = structured.get('show mac address-table', [])
        if not isinstance(mac_list, list):
            mac_list = []
        mac_map: Dict[str, Dict] = {}
        for row in mac_list:
            vlan = str(row.get('vlan', row.get('vlan_id', '')))
            mac = row.get('destination_address', row.get('mac_address', row.get('mac', '')))
            if vlan and mac:
                key = f"{vlan}|{mac.lower()}"
                intf = row.get('destination_port',
                              row.get('ports', row.get('interface', '')))
                if isinstance(intf, list):
                    intf = ', '.join(intf)
                mac_type = row.get('type', '')
                if isinstance(mac_type, list):
                    mac_type = ', '.join(mac_type)
                mac_map[key] = {
                    'vlan': vlan,
                    'mac': mac,
                    'type': str(mac_type),
                    'interface': str(intf),
                }
        return mac_map

    def _write_mac_diff(hostname, ip, mac_a, mac_b):
        all_keys = sorted(set(list(mac_a.keys()) + list(mac_b.keys())))
        for key in all_keys:
            a = mac_a.get(key)
            b = mac_b.get(key)
            if a and not b:
                ws.append([hostname, ip, a['vlan'], a['mac'], 'Removed',
                           a['type'], '', a['interface'], ''])
                _apply_diff_fill(ws, ws.max_row, fills['removed'])
            elif b and not a:
                ws.append([hostname, ip, b['vlan'], b['mac'], 'Added',
                           '', b['type'], '', b['interface']])
                _apply_diff_fill(ws, ws.max_row, fills['added'])
            else:
                changed = a['type'] != b['type'] or a['interface'] != b['interface']
                if changed:
                    ws.append([hostname, ip, b['vlan'], b['mac'], 'Changed',
                               a['type'], b['type'], a['interface'], b['interface']])
                    _apply_diff_fill(ws, ws.max_row, fills['changed'])

    for key, (dev_a, dev_b) in sorted(matched.items()):
        hostname = dev_b.get('meta', {}).get('hostname', key)
        ip = dev_b.get('meta', {}).get('host', key)
        mac_a = _get_mac_map(dev_a)
        if port_map:
            for info in mac_a.values():
                canonical = _normalize_intf_name(info['interface'])
                info['interface'] = port_map.get(canonical, info['interface'])
        _write_mac_diff(hostname, ip, mac_a, _get_mac_map(dev_b))

    for key, dev in sorted(added.items()):
        hostname = dev.get('meta', {}).get('hostname', key)
        ip = dev.get('meta', {}).get('host', key)
        _write_mac_diff(hostname, ip, {}, _get_mac_map(dev))

    for key, dev in sorted(removed.items()):
        hostname = dev.get('meta', {}).get('hostname', key)
        ip = dev.get('meta', {}).get('host', key)
        _write_mac_diff(hostname, ip, _get_mac_map(dev), {})

    _append_no_changes(ws, 9)
    _apply_header_style(ws)


def _diff_sheet_templates(
    wb: Any,
    matched: Dict[str, Tuple],
    added: Dict[str, Dict],
    removed: Dict[str, Dict],
    fills: Dict[str, Any],
    port_map: Optional[Dict[str, str]] = None,
) -> None:
    """Template/port-profile assignment comparison per device."""
    ws = wb.create_sheet("Diff - Templates")
    ws.append(["Hostname", "IP", "Template", "Status",
               "Interfaces (A)", "Interfaces (B)"])

    def _get_template_map(dev: Dict) -> Dict[str, Dict]:
        structured = dev.get('structured', {})
        # IOS: show template → dict with template names as keys
        tmpl_data = structured.get('show template', {})
        # NX-OS: show port-profile → dict with profile names as keys
        if not tmpl_data:
            tmpl_data = structured.get('show port-profile', {})
        if isinstance(tmpl_data, list) and tmpl_data:
            tmpl_data = tmpl_data[0] if tmpl_data else {}
        if not isinstance(tmpl_data, dict):
            return {}
        tmpl_map: Dict[str, Dict] = {}
        for name, info in tmpl_data.items():
            if name.startswith('_'):
                continue
            intfs = ''
            if isinstance(info, dict):
                intfs = info.get('interfaces', info.get('bound_to', ''))
                if isinstance(intfs, list):
                    intfs = ', '.join(intfs)
            elif isinstance(info, str):
                intfs = info
            tmpl_map[name] = {'interfaces': str(intfs)}
        return tmpl_map

    def _write_template_diff(hostname, ip, tmpl_a, tmpl_b):
        all_names = sorted(set(list(tmpl_a.keys()) + list(tmpl_b.keys())))
        for name in all_names:
            a = tmpl_a.get(name)
            b = tmpl_b.get(name)
            if a and not b:
                ws.append([hostname, ip, name, 'Removed', a['interfaces'], ''])
                _apply_diff_fill(ws, ws.max_row, fills['removed'])
            elif b and not a:
                ws.append([hostname, ip, name, 'Added', '', b['interfaces']])
                _apply_diff_fill(ws, ws.max_row, fills['added'])
            else:
                if a['interfaces'] != b['interfaces']:
                    ws.append([hostname, ip, name, 'Changed',
                               a['interfaces'], b['interfaces']])
                    _apply_diff_fill(ws, ws.max_row, fills['changed'])

    for key, (dev_a, dev_b) in sorted(matched.items()):
        hostname = dev_b.get('meta', {}).get('hostname', key)
        ip = dev_b.get('meta', {}).get('host', key)
        tmpl_a = _get_template_map(dev_a)
        tmpl_b = _get_template_map(dev_b)
        if port_map:
            for info in tmpl_a.values():
                info['interfaces'] = _remap_intf_string(info['interfaces'], port_map)
            for info in tmpl_b.values():
                info['interfaces'] = _normalize_intf_string(info['interfaces'])
        _write_template_diff(hostname, ip, tmpl_a, tmpl_b)

    for key, dev in sorted(added.items()):
        hostname = dev.get('meta', {}).get('hostname', key)
        ip = dev.get('meta', {}).get('host', key)
        _write_template_diff(hostname, ip, {}, _get_template_map(dev))

    for key, dev in sorted(removed.items()):
        hostname = dev.get('meta', {}).get('hostname', key)
        ip = dev.get('meta', {}).get('host', key)
        _write_template_diff(hostname, ip, _get_template_map(dev), {})

    _append_no_changes(ws, 6)
    _apply_header_style(ws)


def generate_diff_report(
    run_a: Path, run_b: Path, output_path: Path,
    logger: logging.Logger,
    port_map: Optional[Dict[str, str]] = None,
) -> Optional[Path]:
    """Generate a Device Diff Excel report comparing two runs.

    Args:
        run_a: Path to the older run folder
        run_b: Path to the newer run folder
        output_path: Full path for the output .xlsx file
        logger: Logger instance
        port_map: Optional mapping of old→new interface names for switch replacements

    Returns:
        Path to the saved report, or None on failure.
    """
    from openpyxl import Workbook  # type: ignore

    devs_a = _load_run_devices(run_a)
    devs_b = _load_run_devices(run_b)

    if not devs_a and not devs_b:
        logger.error("No device data found in either run.")
        return None

    matched, added, removed = _match_devices(devs_a, devs_b)
    fills = _diff_fills()

    wb = Workbook()

    # Sheet 1: Run Info (uses the default sheet)
    _diff_sheet_run_info(wb, run_a, run_b, devs_a, devs_b, port_map)

    # Sheet 2-10: Component diff sheets
    _diff_sheet_summary(wb, matched, added, removed, fills)
    _diff_sheet_interfaces(wb, matched, added, removed, fills, port_map)
    _diff_sheet_neighbors(wb, matched, added, removed, fills, port_map)
    _diff_sheet_vlans(wb, matched, added, removed, fills, port_map)
    _diff_sheet_routing(wb, matched, added, removed, fills)
    _diff_sheet_stp(wb, matched, added, removed, fills)
    _diff_sheet_mac_summary(wb, matched, added, removed, fills)
    _diff_sheet_mac_addresses(wb, matched, added, removed, fills, port_map)
    _diff_sheet_templates(wb, matched, added, removed, fills, port_map)

    try:
        wb.save(output_path)
        logger.info(f'Diff report saved: "{output_path.name}"')
        return output_path
    except Exception as e:
        logger.error(f"Failed to save diff report: {e}")
        return None


# --- Main Execution ---
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MyWindow()
    window.show()
    sys.exit(app.exec())
